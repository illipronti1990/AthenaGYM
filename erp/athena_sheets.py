"""Sprint 12.0 — ATHENA AI + Central de Recomendações + Automação."""

from __future__ import annotations

from datetime import date

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()

RECOM_HEADERS = ["ID", "Tipo", "Descrição", "Responsável", "Situação", "Prioridade", "Categoria", "Data"]
RECOM_SEED = [
    (1, "Cobrança", "Cobrar 5 alunos com mensalidade em atraso", "Financeiro", "Pendente", "🔴", "Financeiro", HOJE),
    (2, "Estoque", "Comprar Creatina (giro alto / estoque baixo)", "Estoque", "Pendente", "🟠", "Estoque", HOJE),
    (3, "Treino", "Atualizar 8 treinos com mais de 45 dias", "Professor", "Pendente", "🟡", "Treinos", HOJE),
    (4, "Retenção", "Campanha de retenção para alunos em alto risco", "CRM", "Pendente", "🟢", "Retenção", HOJE),
    (5, "Avaliação", "Marcar 12 reavaliações físicas", "Professor", "Pendente", "🔵", "Avaliação", HOJE),
]

ATHENA_CHAT_HEADERS = ["ID", "Data", "Hora", "Pergunta", "Resposta", "Usuário", "Módulo"]
ATHENA_CHAT_SEED = [
    (
        1,
        HOJE,
        "08:00",
        "Quanto faturei este mês?",
        "Receita do mês em análise. Use Perguntar na tela ATHENA AI para atualizar.",
        "admin",
        "Financeiro",
    ),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def _bd_sheet(wb, name, headers, rows, table, widths, date_cols=()) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).fill = fill_gold
        ws.cell(row=1, column=i).font = font_header
        ws.cell(row=1, column=i).border = border_thin
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx in date_cols:
                _date_fmt(cell)
    last = max(1 + len(rows), 2)
    _make_table(ws, table, f"A1:{get_column_letter(len(headers))}{last}")
    set_column_widths(ws, widths)
    ws.sheet_state = "hidden"


def build_bd_recomendacoes(wb) -> None:
    _bd_sheet(
        wb,
        "BD_RECOMENDACOES",
        RECOM_HEADERS,
        RECOM_SEED,
        "tbRecomendacoes",
        {1: 6, 2: 12, 3: 48, 4: 14, 5: 12, 6: 10, 7: 12, 8: 12},
        date_cols={8},
    )


def build_bd_athena_chat(wb) -> None:
    _bd_sheet(
        wb,
        "BD_ATHENA_CHAT",
        ATHENA_CHAT_HEADERS,
        ATHENA_CHAT_SEED,
        "tbAthenaChat",
        {1: 6, 2: 12, 3: 8, 4: 32, 5: 55, 6: 12, 7: 12},
        date_cols={2},
    )


def build_athena_ai_ui(wb) -> None:
    """36_ATHENA_AI — assistente inteligente (perguntas + briefing + previsões)."""
    ws = wb.create_sheet("36_ATHENA_AI")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=58, cols=13)
    add_sidebar(ws, active="36_ATHENA_AI", rows=58, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 14, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 12, 11: 12},
    )
    _title(ws, "ATHENA AI — ASSISTENTE INTELIGENTE")

    ws.merge_cells("C7:K7")
    ws["C7"] = "Bom dia. Analise os dados da academia, faça perguntas e receba recomendações."
    ws["C7"].font = Font(name="Calibri", size=11, italic=True, color="555555")

    paint_kpi_card(ws, 9, 3, "Cobranças", 0, False, 2)
    paint_kpi_card(ws, 9, 5, "Avaliações", 0, False, 2)
    paint_kpi_card(ws, 9, 7, "Falta estoque", 0, False, 2)
    paint_kpi_card(ws, 9, 9, "Receita prev. hoje", "R$ 0", False, 2)

    ws.merge_cells("C14:K14")
    ws["C14"] = "FAÇA UMA PERGUNTA"
    ws["C14"].font = font_section
    ws["C14"].fill = fill_brand

    ws["C15"] = "Pergunta"
    ws["C15"].fill = fill_panel
    ws["C15"].border = border_thin
    ws.merge_cells("D15:K15")
    ws["D15"] = "Quanto faturei este mês?"
    ws["D15"].fill = fill_panel
    ws["D15"].border = border_thin
    ws["D15"].font = Font(name="Georgia", size=12, color=BRAND_RED)

    ws.merge_cells("C17:K17")
    ws["C17"] = "RESPOSTA ATHENA"
    ws["C17"].font = font_section
    ws["C17"].fill = fill_brand
    for i in range(6):
        r = 18 + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
        ws.cell(row=r, column=3).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
    ws["C18"] = "Clique em Perguntar para analisar os dados do ERP."
    ws["C18"].font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)

    ws.merge_cells("C25:F25")
    ws["C25"] = "EXEMPLOS"
    ws["C25"].font = font_section
    ws["C25"].fill = fill_brand
    exemplos = [
        "Quanto faturei este mês?",
        "Quem está em risco de cancelar?",
        "Quais produtos devo comprar?",
        "Qual o horário de pico?",
        "Como está o churn?",
        "Previsão de receita em 90 dias?",
    ]
    for i, ex in enumerate(exemplos):
        r = 26 + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.cell(row=r, column=3, value=ex).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin

    ws.merge_cells("H25:K25")
    ws["H25"] = "PREVISÃO FINANCEIRA"
    ws["H25"].font = font_section
    ws["H25"].fill = fill_brand
    for i, lab in enumerate(["Hoje / atual", "30 dias", "60 dias", "90 dias"]):
        r = 26 + i
        ws.cell(row=r, column=8, value=lab).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
        ws.cell(row=r, column=9, value="R$ 0").fill = fill_panel
        ws.cell(row=r, column=9).border = border_thin
        ws.cell(row=r, column=9).font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)

    ws.merge_cells("C33:K33")
    ws["C33"] = "INSIGHTS AUTOMÁTICOS (FINANCEIRO · COMERCIAL · OPERACIONAL)"
    ws["C33"].font = font_section
    ws["C33"].fill = fill_brand
    for i in range(8):
        r = 34 + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
        ws.cell(row=r, column=3).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin

    ws.merge_cells("C43:K43")
    ws["C43"] = "AUTOMACÕES — WhatsApp / E-mail / Relatório (filas geradas)"
    ws["C43"].font = font_section
    ws["C43"].fill = fill_brand
    for i, h in enumerate(["Canal", "Destino", "Mensagem / Assunto", "Status"], start=3):
        ws.cell(row=44, column=i, value=h).fill = fill_gold
        ws.cell(row=44, column=i).font = font_header
    for i in range(5):
        r = 45 + i
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C51:K51")
    ws["C51"] = (
        "Motor rule-based sobre o ERP · reutiliza BI/CRM/Estoque/Acesso · "
        "BD_INSIGHTS · BD_PREVISOES · BD_RECOMENDACOES · SQL cloud na evolução"
    )
    ws["C51"].font = Font(name="Calibri", size=9, color="666666")
    ws.row_dimensions[53].height = 36


def build_recomendacoes_ui(wb) -> None:
    """37_RECOMENDACOES — central de ações priorizadas."""
    ws = wb.create_sheet("37_RECOMENDACOES")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=48, cols=13)
    add_sidebar(ws, active="37_RECOMENDACOES", rows=48, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 10, 4: 14, 5: 36, 6: 14, 7: 12, 8: 12, 9: 12},
    )
    _title(ws, "CENTRAL DE RECOMENDAÇÕES")

    paint_kpi_card(ws, 7, 3, "Pendentes", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Alta prioridade", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Concluídas", 0, False, 2)
    paint_kpi_card(ws, 7, 9, "Módulos", 0, False, 2)

    ws.merge_cells("C12:I12")
    ws["C12"] = "RECOMENDAÇÕES DO DIA"
    ws["C12"].font = font_section
    ws["C12"].fill = fill_brand

    headers = ["Pri.", "Tipo", "Descrição", "Responsável", "Situação", "Categoria"]
    for i, h in enumerate(headers, start=3):
        ws.cell(row=13, column=i, value=h).fill = fill_gold
        ws.cell(row=13, column=i).font = font_header
    for i in range(12):
        r = 14 + i
        for c in range(3, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C28:I28")
    ws["C28"] = "LEGENDA DE PRIORIDADE"
    ws["C28"].font = font_section
    ws["C28"].fill = fill_brand
    for i, txt in enumerate(
        [
            "🔴 Cobranças / financeiro crítico",
            "🟠 Estoque / ruptura",
            "🟡 Treinos desatualizados",
            "🟢 Retenção / campanhas",
            "🔵 Reavaliações / agenda",
        ]
    ):
        r = 29 + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        ws.cell(row=r, column=3, value=txt).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin

    ws.merge_cells("C35:I35")
    ws["C35"] = "Atualize com AtualizarRecomendacoes · execute ações no módulo responsável · marque Situação no BD."
    ws["C35"].font = Font(name="Calibri", size=9, color="666666")
    ws.row_dimensions[37].height = 36
