"""Épico 3 — Franquias: franqueadoras, franqueados, contratos, royalties, UI 41."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()

FRANQUEADORAS_HEADERS = [
    "FranqueadoraID", "Nome", "CNPJ", "Razão Social", "CEO", "E-mail", "Telefone", "Site", "Status",
]
FRANQUEADORAS_SEED = [
    (
        1, "ATHENA FRANCHISE", "00.111.222/0001-33", "ATHENA FRANCHISE HOLDING LTDA",
        "Renan Athena", "franchise@athena.gym", "(11) 3000-0100", "https://franchise.athena.gym", "Ativa",
    ),
]

FRANQUEADOS_HEADERS = [
    "FranqueadoID", "FranqueadoraID", "EmpresaID", "Nome", "CPF/CNPJ", "Cidade", "Estado",
    "Contrato", "Data Início", "Data Fim", "Status",
]
FRANQUEADOS_SEED = [
    (1, 1, 1, "Franquia São Paulo", "12.345.678/0001-90", "São Paulo", "SP",
     "CTR-SP-001", HOJE - timedelta(days=400), HOJE + timedelta(days=965), "Ativo"),
    (2, 1, 2, "Franquia Campinas", "23.456.789/0001-01", "Campinas", "SP",
     "CTR-CP-002", HOJE - timedelta(days=200), HOJE + timedelta(days=1165), "Ativo"),
    (3, 1, 3, "Franquia Santos", "34.567.890/0001-12", "Santos", "SP",
     "CTR-ST-003", HOJE - timedelta(days=120), HOJE + timedelta(days=1245), "Ativo"),
]

CONTRATOS_HEADERS = [
    "ID", "Número", "FranqueadoID", "Franqueado", "Vigência Início", "Vigência Fim",
    "Taxa Inicial", "Royalty %", "Fundo Marketing %", "Status",
]
CONTRATOS_SEED = [
    (1, "CTR-SP-001", 1, "Franquia São Paulo", HOJE - timedelta(days=400), HOJE + timedelta(days=965),
     50000.0, 6.0, 2.0, "Ativo"),
    (2, "CTR-CP-002", 2, "Franquia Campinas", HOJE - timedelta(days=200), HOJE + timedelta(days=1165),
     45000.0, 6.0, 2.0, "Ativo"),
    (3, "CTR-ST-003", 3, "Franquia Santos", HOJE - timedelta(days=120), HOJE + timedelta(days=1245),
     40000.0, 6.0, 2.0, "Ativo"),
]

ROYALTIES_HEADERS = [
    "RoyaltyID", "FranqueadoID", "Franqueado", "Competência", "Receita Base",
    "Percentual", "Valor Royalty", "PercentualMarketing", "ValorMarketing", "Status",
]
# Competência = 1º dia do mês atual
_COMP = date(HOJE.year, HOJE.month, 1)
ROYALTIES_SEED = [
    (1, 1, "Franquia São Paulo", _COMP, 920000.0, 6.0, 55200.0, 2.0, 18400.0, "Em Aberto"),
    (2, 2, "Franquia Campinas", _COMP, 780000.0, 6.0, 46800.0, 2.0, 15600.0, "Em Aberto"),
    (3, 3, "Franquia Santos", _COMP, 590000.0, 6.0, 35400.0, 2.0, 11800.0, "Pago"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _bd_sheet(wb, name, headers, rows, table, widths, date_cols=()) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).fill = fill_gold
        ws.cell(row=1, column=i).font = font_header
        ws.cell(row=1, column=i).border = border_thin
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx in date_cols:
                _date_fmt(cell)
    last = max(1 + len(rows), 2)
    _make_table(ws, table, f"A1:{get_column_letter(len(headers))}{last}")
    set_column_widths(ws, widths)
    ws.sheet_state = "hidden"


def build_bd_franqueadoras(wb) -> None:
    _bd_sheet(
        wb, "BD_FRANQUEADORAS", FRANQUEADORAS_HEADERS, FRANQUEADORAS_SEED, "tbFranqueadoras",
        {1: 14, 2: 20, 3: 18, 4: 28, 5: 16, 6: 24, 7: 16, 8: 28, 9: 10},
    )


def build_bd_franqueados(wb) -> None:
    _bd_sheet(
        wb, "BD_FRANQUEADOS", FRANQUEADOS_HEADERS, FRANQUEADOS_SEED, "tbFranqueados",
        {1: 12, 2: 14, 3: 10, 4: 20, 5: 18, 6: 14, 7: 8, 8: 14, 9: 12, 10: 12, 11: 10},
        date_cols={9, 10},
    )


def build_bd_contratos_franquia(wb) -> None:
    _bd_sheet(
        wb, "BD_CONTRATOS_FRANQUIA", CONTRATOS_HEADERS, CONTRATOS_SEED, "tbContratosFranquia",
        {1: 6, 2: 14, 3: 12, 4: 20, 5: 14, 6: 14, 7: 12, 8: 10, 9: 16, 10: 10},
        date_cols={5, 6},
    )


def build_bd_royalties(wb) -> None:
    _bd_sheet(
        wb, "BD_ROYALTIES", ROYALTIES_HEADERS, ROYALTIES_SEED, "tbRoyalties",
        {1: 10, 2: 12, 3: 20, 4: 12, 5: 12, 6: 10, 7: 12, 8: 16, 9: 14, 10: 12},
        date_cols={4},
    )


def build_franqueadora_ui(wb) -> None:
    """41_FRANQUEADORA — portal da franqueadora (KPIs + ranking + cadastro)."""
    ws = wb.create_sheet("41_FRANQUEADORA")
    style_sheet_tab(ws, BRAND_RED)
    paint_canvas(ws, rows=52, cols=12)
    add_sidebar(ws, active="41_FRANQUEADORA", rows=52, labels=False)
    add_top_bar(ws, start_col=2, end_col=11)
    set_column_widths(ws, {1: 24, 2: 2, 3: 16, 4: 16, 5: 14, 6: 14, 7: 14, 8: 14, 9: 12})

    ws.merge_cells("C5:H5")
    ws["C5"] = "ATHENA FRANCHISE"
    ws["C5"].font = Font(name="Georgia", size=20, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light

    ws.merge_cells("C6:H6")
    ws["C6"] = "Épico 3 — Portal da Franqueadora · rede · royalties · ranking"
    ws["C6"].font = Font(name="Calibri", size=10, color="666666")

    paint_kpi_card(ws, 8, 3, "Franquias", "3", False, 2)
    paint_kpi_card(ws, 8, 5, "Unidades", "0", False, 2)
    paint_kpi_card(ws, 8, 7, "Alunos", "0", False, 2)
    paint_kpi_card(ws, 11, 3, "Receita rede", "R$ 0", False, 2)
    paint_kpi_card(ws, 11, 5, "Royalties", "R$ 0", False, 2)

    ws.merge_cells("C15:H15")
    ws["C15"] = "RANKING DAS FRANQUIAS"
    ws["C15"].font = font_section
    ws["C15"].fill = fill_brand
    for i, h in enumerate(["Franquia", "Receita", "Crescimento", "Royalties", "Status"], start=3):
        ws.cell(row=16, column=i, value=h).fill = fill_gold
        ws.cell(row=16, column=i).font = font_header
        ws.cell(row=16, column=i).border = border_thin
    for r in range(17, 23):
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C24:H24")
    ws["C24"] = "CADASTRO RÁPIDO — FRANQUEADO"
    ws["C24"].font = font_section
    ws["C24"].fill = fill_brand

    fields = [
        (25, "Nome", "D25"),
        (26, "EmpresaID", "D26"),
        (27, "CPF/CNPJ", "D27"),
        (28, "Cidade", "D28"),
        (29, "Estado", "D29"),
        (30, "Contrato", "D30"),
    ]
    for row, lab, _ in fields:
        ws.cell(row=row, column=3, value=lab).fill = fill_panel
        ws.cell(row=row, column=3).border = border_thin
        ws.cell(row=row, column=3).font = Font(name="Calibri", size=10, color="FFFFFF")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.cell(row=row, column=4).fill = fill_light
        ws.cell(row=row, column=4).border = border_thin

    ws["D26"] = 2
    ws["D29"] = "SP"

    ws.merge_cells("C32:H32")
    ws["C32"] = "Sessão franquia"
    ws["C32"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    ws["C33"] = "FranqueadoraID"
    ws["D33"] = 1
    ws["C34"] = "FranqueadoID"
    ws["D34"] = 0
    ws["C35"] = "Nome rede"
    ws["D35"] = "ATHENA FRANCHISE"

    ws.merge_cells("C37:H37")
    ws["C37"] = "Relatório / competência"
    ws["C37"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    ws.merge_cells("C38:H42")
    ws["C38"] = "Use GerarRelatorioFranqueadora para preencher este bloco."
    ws["C38"].fill = fill_light
    ws["C38"].border = border_thin

    ws.merge_cells("C44:H44")
    ws["C44"] = (
        "Ações: CadastrarFranqueado · CalcularRoyalties · AtualizarDashboardFranqueadora · "
        "GerarRelatorioFranqueadora"
    )
    ws["C44"].font = Font(name="Calibri", size=9, color="666666")
