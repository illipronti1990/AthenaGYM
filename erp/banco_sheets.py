"""Abas de banco de dados (ListObjects) — Sprint 3.5 Motor de Configuração."""

from __future__ import annotations

from datetime import date

from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from config_data import ALUNOS, CONTAS_PAGAR, gerar_mensalidades
from styles import (
    BRAND_RED,
    GOLD,
    apply_header_row,
    border_thin,
    font_normal,
    set_column_widths,
    style_sheet_tab,
)

BD_ALUNOS_HEADERS = [
    "ID",
    "EmpresaID",
    "UnidadeID",
    "Matrícula",
    "Nome",
    "CPF",
    "RG",
    "DataNascimento",
    "Sexo",
    "Telefone",
    "WhatsApp",
    "Email",
    "CEP",
    "Endereço",
    "Número",
    "Bairro",
    "Cidade",
    "Plano",
    "Professor",
    "ValorPlano",
    "FormaPagamento",
    "DataCadastro",
    "Status",
]

# Grupo | Parâmetro | Valor | Descrição
PARAMETROS_SEED = [
    ("Sistema", "PrefixoMatricula", "ATH", "Prefixo da matrícula"),
    ("Sistema", "AnoAtual", date.today().year, "Ano corrente da matrícula"),
    ("Sistema", "Versao", "5.0.0", "Versão do ERP"),
    ("CRM", "DiasSemPresenca", 15, "Dias sem check-in = aluno em risco"),
    ("CRM", "DiasPlanoVencendo", 30, "Dias antes do vencimento do plano"),
    ("CRM", "FrequenciaMinimaMes", 4, "Check-ins mínimos no mês"),
    ("CRM", "BonusIndicacao", 50, "Bônus R$ quando indicação vira aluno"),
    ("Treinos", "DiasReavaliacao", 60, "Dias para próxima avaliação física"),
    ("Acesso", "BloquearInadimplente", "SIM", "SIM ou NAO — bloquear entrada se atrasado"),
    ("Acesso", "DiasTolerancia", 5, "Dias de tolerancia financeira (referencia)"),
    ("Acesso", "PermitirLiberacaoManual", "SIM", "SIM permite override na recepcao"),
    ("Acesso", "RegistrarSaida", "SIM", "SIM registra saida e tempo"),
    ("Acesso", "TempoMaximoPermanencia", 4, "Horas maximas de permanencia"),
    ("Acesso", "DiasSemAcessoRisco", 15, "Dias sem acesso = risco/CRM"),
    ("Integracoes", "QRCode", "NAO", "Habilitar leitor QR"),
    ("Integracoes", "Biometria", "NAO", "Habilitar biometria"),
    ("Integracoes", "RFID", "NAO", "Habilitar cartao RFID/NFC"),
    ("Integracoes", "AppAluno", "NAO", "Habilitar app do aluno"),
    ("Estoque", "UnidadePadrao", "ATHENA GYM Matriz", "Unidade operacional padrão"),
    ("Estoque", "DiasAlertaValidade", 30, "Dias para alertar lote vencendo"),
    ("Estoque", "PerguntarMensalidade", "SIM", "PDV pergunta se lança na mensalidade"),
    ("Estoque", "EstoqueMinimoPadrao", 5, "Mínimo padrão ao cadastrar"),
    ("BI", "CACEstimadoMes", 2000, "Marketing mensal estimado para CAC"),
    ("BI", "DiasSemAcessoRisco", 15, "Dias sem acesso = risco retenção BI"),
    ("BI", "AbrirExecutivoLogin", "NAO", "SIM abre 31_BI_EXECUTIVO após login"),
    ("Portal", "ApiUrl", "http://127.0.0.1:8002", "URL da API cloud (Supabase)"),
    ("Portal", "SyncUser", "admin", "Usuário para POST /sync/push"),
    ("Portal", "SyncPass", "123456", "Senha para POST /sync/push"),
    ("Portal", "SyncAutoPush", "SIM", "SIM = após exportar JSON, envia à API"),
    ("Portal", "MatriculaDemo", f"ATH-{date.today().year}-000001", "Matrícula demo do portal aluno"),
    ("Portal", "AppAtivo", "SIM", "Habilita espelho portal no Excel"),
    ("Athena", "DiasTreinoDesatualizado", 45, "Dias para sugerir atualização de treino"),
    ("Athena", "AbrirAthenaLogin", "NAO", "SIM abre 36_ATHENA_AI após login"),
    ("Financeiro", "Multa", 2, "% multa por atraso"),
    ("Financeiro", "Juros", 1, "% juros ao mês"),
    ("Financeiro", "DiasTolerancia", 5, "Dias sem multa após vencimento"),
    ("Financeiro", "DiaVencimentoPadrao", 10, "Dia padrão de vencimento"),
    ("Financeiro", "MetaReceitaMes", 50000, "Meta de receita do mês"),
    ("Financeiro", "BloquearEmailDuplicado", "NÃO", "SIM ou NÃO"),
    ("Financeiro", "BloquearTelefoneDuplicado", "NÃO", "SIM ou NÃO"),
    ("Academia", "Nome", "ATHENA GYM", "Nome da academia"),
    ("Academia", "Cidade", "São Paulo", "Cidade"),
    ("Academia", "Telefone", "(11) 99999-9999", "Contato"),
    ("Academia", "Email", "contato@athenagym.com.br", "E-mail de contato"),
    ("Academia", "Moeda", "R$", "Símbolo de moeda"),
]

PLANOS_BD_SEED = [
    (1, "Mensal", 149.0, 50.0, 10, 0, "Ativo"),
    (2, "Premium", 199.0, 80.0, 10, 12, "Ativo"),
    (3, "VIP", 299.0, 100.0, 5, 12, "Ativo"),
    (4, "Trimestral", 109.9, 50.0, 10, 3, "Ativo"),
    (5, "Semestral", 99.9, 50.0, 10, 6, "Ativo"),
    (6, "Anual", 89.9, 50.0, 10, 12, "Ativo"),
    (7, "Personal", 350.0, 80.0, 10, 0, "Ativo"),
    (8, "Day Use", 40.0, 0.0, 10, 0, "Ativo"),
]

FORMAS_BD_SEED = [
    (1, "PIX", 0.0, 0),
    (2, "Dinheiro", 0.0, 0),
    (3, "Cartão Débito", 1.5, 1),
    (4, "Cartão Crédito", 3.0, 30),
    (5, "Boleto", 2.0, 2),
    (6, "Transferência", 0.0, 1),
]

STATUS_BD_SEED = [
    ("Aluno", "Ativo"),
    ("Aluno", "Inativo"),
    ("Aluno", "Congelado"),
    ("Aluno", "Cancelado"),
    ("Aluno", "Inadimplente"),
    ("Mensalidade", "Pendente"),
    ("Mensalidade", "Pago"),
    ("Mensalidade", "Atrasado"),
    ("Mensalidade", "Cancelado"),
    ("Mensalidade", "Vence hoje"),
    ("Professor", "Ativo"),
    ("Professor", "Férias"),
    ("Professor", "Inativo"),
    ("Evento", "Pendente"),
    ("Evento", "Concluído"),
    ("Evento", "Cancelado"),
]

PERMISSOES_SEED = [
    ("Administrador", 1, 1, 1, 1, 1, 1, 1, 1),
    ("Financeiro", 1, 1, 0, 0, 1, 1, 1, 0),
    ("Recepção", 1, 0, 1, 0, 1, 1, 1, 0),
    ("Professor", 1, 0, 0, 0, 1, 0, 1, 0),
    ("Franqueadora", 1, 1, 1, 1, 1, 1, 1, 0),
    ("Franqueado", 1, 1, 1, 0, 1, 1, 1, 0),
]

CORES_SEED = [
    ("Menu", "#A3001E"),
    ("Botão", "#D4AF37"),
    ("Fundo", "#FFFFFF"),
    ("Título", "#000000"),
    ("Canvas", "#F3F3F3"),
    ("Painel", "#5C0010"),
]

VERSAO_SEED = [
    ("2.0.0", date(2026, 7, 21), "Sprint 3.1–3.4 — Login, alunos, arquitetura MVC"),
    ("2.0.1", date(2026, 7, 21), "Sprint 3.5 — Motor de configuração (BD_PARAMETROS e mestras)"),
    ("2.1.0", date(2026, 7, 21), "Sprint 4.0 — Motor financeiro + livro-razão BD_LANCAMENTOS"),
    ("2.2.0", date(2026, 7, 21), "Sprint 5.0 — Dashboard Executivo 360° + BI + alertas + metas"),
    ("2.2.1", date(2026, 7, 21), "Sprint 5.1 — Agenda Inteligente / Central Operacional"),
    ("2.2.2", date(2026, 7, 21), "Sprint 5.1.1 — Operation Center / Painel de Ações do Dia"),
    ("2.3.0", date(2026, 7, 22), "Sprint 6.0 — CRM Inteligente (leads, funil, retenção, campanhas)"),
    ("2.4.0", date(2026, 7, 22), "Sprint 7.0 — Treinos + Avaliação Física"),
    ("2.5.0", date(2026, 7, 22), "Sprint 8.0 — Controle de Acesso, Presença e Frequência"),
    ("2.6.0", date(2026, 7, 22), "Sprint 9.0 — PDV Inteligente + Gestão de Estoque"),
    ("2.7.0", date(2026, 7, 22), "Sprint 10.0 — Business Intelligence + Inteligência Analítica"),
    ("2.8.0", date(2026, 7, 22), "Sprint 11.0 — Portal do Aluno + API Cloud + App Mobile"),
    ("2.9.0", date(2026, 7, 22), "Sprint 12.0 — ATHENA AI + Automação + Central de Recomendações"),
    ("3.0.0", date(2026, 7, 22), "Épico 1 — Multi-Tenant SaaS (EmpresaID, licenças, Master)"),
    ("5.0.0", date(2026, 7, 22), "Épico 3 A+B — Franquias, contratos, royalties e portal franqueadora"),
    ("4.1.0", date(2026, 7, 22), "Épico 2 C/D — Transferências, professores multiunidade, KPIs e permissões"),
    ("4.0.0", date(2026, 7, 22), "Épico 2 — Multiunidade (UnidadeID, filiais, ATH-CODIGO)"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def _money(cell) -> None:
    cell.number_format = 'R$ #,##0.00'


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _aluno_row(aluno: dict, idx: int) -> list:
    return [
        idx,
        1,  # EmpresaID — ATHENA GYM tenant demo
        1,  # UnidadeID — Matriz
        aluno["matricula"],
        aluno["nome"],
        aluno["cpf"],
        aluno["rg"],
        aluno["nasc"],
        aluno["sexo"],
        aluno["telefone"],
        aluno["whatsapp"],
        aluno["email"],
        aluno["cep"],
        aluno["rua"],
        aluno["numero"],
        aluno["bairro"],
        aluno["cidade"],
        aluno["plano"],
        aluno["professor"],
        aluno["valor"],
        "PIX",
        aluno["inicio"],
        aluno["status"],
    ]


def build_bd_alunos(wb) -> None:
    ws = wb.create_sheet("BD_ALUNOS")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False

    for i, h in enumerate(BD_ALUNOS_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(BD_ALUNOS_HEADERS))

    text_cols = {6, 7, 10, 11, 13, 15}
    for r_idx, aluno in enumerate(ALUNOS, start=2):
        values = _aluno_row(aluno, r_idx - 1)
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if c_idx in text_cols:
                cell.number_format = "@"
                if val is not None:
                    cell.value = str(val)
            if isinstance(val, date) and c_idx not in text_cols:
                _date_fmt(cell)
            if c_idx == 20:
                _money(cell)

    last = 1 + len(ALUNOS)
    for c in range(1, len(BD_ALUNOS_HEADERS) + 1):
        cell = ws.cell(row=last + 1, column=c)
        cell.value = None
        if c in text_cols:
            cell.number_format = "@"
    _make_table(ws, "tbAlunos", f"A1:W{last + 1}")
    for col in text_cols:
        for r in range(2, last + 2):
            ws.cell(row=r, column=col).number_format = "@"

    ws.cell(row=last + 3, column=1, value="Banco canônico — EmpresaID + UnidadeID (Épico 2) · cadastro via frmAluno.")
    ws.cell(row=last + 3, column=1).font = Font(name="Calibri", size=9, color=BRAND_RED)
    set_column_widths(ws, {1: 6, 2: 10, 3: 10, 4: 16, 5: 22, 6: 15, 7: 12, 8: 13, 9: 10, 10: 15, 11: 15, 12: 22, 13: 11, 14: 16, 15: 8, 16: 12, 17: 12, 18: 12, 19: 16, 20: 11, 21: 14, 22: 12, 23: 12})
    ws.freeze_panes = "A2"


_MESES_PT = (
    "",
    "Janeiro",
    "Fevereiro",
    "Março",
    "Abril",
    "Maio",
    "Junho",
    "Julho",
    "Agosto",
    "Setembro",
    "Outubro",
    "Novembro",
    "Dezembro",
)


def _competencia_texto(d: date) -> str:
    return f"{_MESES_PT[d.month]}/{d.year}"


def _encargos(valor: float, venc: date, status: str, hoje: date | None = None) -> tuple[float, float, float]:
    """Retorna (multa, juros, valor_final) — espelha regra VBA (tolerância 5, multa 2%, juros 1%/mês)."""
    hoje = hoje or date.today()
    if status == "Pago":
        return 0.0, 0.0, valor
    dias = max(0, (hoje - venc).days)
    if dias <= 5:
        return 0.0, 0.0, valor
    multa = round(valor * 0.02, 2)
    juros = round(valor * 0.01 * (dias / 30.0), 2)
    return multa, juros, round(valor + multa + juros, 2)


def build_bd_contas_receber(wb) -> None:
    """BD_CONTAS_RECEBER — motor de contas a receber (Sprint 4.0)."""
    ws = wb.create_sheet("BD_CONTAS_RECEBER")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False
    headers = [
        "ID",
        "EmpresaID",
        "UnidadeID",
        "Matrícula",
        "Nome",
        "Competência",
        "Valor Original",
        "Desconto",
        "Multa",
        "Juros",
        "Valor Final",
        "Forma Pagamento",
        "Data Vencimento",
        "Data Pagamento",
        "Situação",
        "Observação",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    rows = gerar_mensalidades()
    rid = 0
    for m in rows:
        rid += 1
        multa, juros, vfinal = _encargos(m["valor"], m["vencimento"], m["status"])
        desconto = 0.0
        if m["status"] == "Pago":
            vfinal = m["valor"]
            multa = 0.0
            juros = 0.0
        vals = [
            rid,
            1,
            1,
            m["codigo"],
            m["aluno"],
            _competencia_texto(m["competencia"]),
            m["valor"],
            desconto,
            multa,
            juros,
            vfinal,
            m["forma"] or "",
            m["vencimento"],
            m["pagamento"],
            m["status"] if m["status"] != "Vence hoje" else "Pendente",
            "",
        ]
        r_idx = rid + 1
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if c_idx in (7, 8, 9, 10, 11):
                _money(cell)
            if c_idx in (13, 14) and isinstance(val, date):
                _date_fmt(cell)

    last = 1 + max(rid, 1)
    if rid == 0:
        for c in range(1, len(headers) + 1):
            ws.cell(row=2, column=c).value = None
        last = 2
    else:
        for c in range(1, len(headers) + 1):
            ws.cell(row=last + 1, column=c).value = None
        last = last + 1

    _make_table(ws, "tbContasReceberBD", f"A1:P{last}")
    set_column_widths(
        ws,
        {1: 6, 2: 10, 3: 10, 4: 14, 5: 22, 6: 14, 7: 12, 8: 10, 9: 10, 10: 10, 11: 12, 12: 14, 13: 12, 14: 12, 15: 12, 16: 20},
    )
    ws.freeze_panes = "A2"


def build_bd_lancamentos(wb) -> None:
    """Livro-razão — eventos financeiros imutáveis (Sprint 4.0)."""
    ws = wb.create_sheet("BD_LANCAMENTOS")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    headers = ["ID", "EmpresaID", "UnidadeID", "Data", "Tipo", "Origem", "Categoria", "Documento", "Débito", "Crédito", "Usuário"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    # Seed a partir de contas pagas + despesas típicas
    seed = []
    lid = 0
    for m in gerar_mensalidades():
        if m["status"] != "Pago" or m["pagamento"] is None:
            continue
        lid += 1
        seed.append(
            (
                lid,
                1,
                1,
                m["pagamento"],
                "Recebimento",
                "Mensalidade",
                "Receita",
                f"REC-{lid:06d}",
                0.0,
                m["valor"],
                "admin",
            )
        )
        if lid >= 12:
            break

    despesas_seed = [
        (date.today().replace(day=5), "Pagamento", "Aluguel", "Despesa", 3500.0),
        (date.today().replace(day=8), "Pagamento", "Energia", "Despesa", 890.0),
        (date.today().replace(day=10), "Pagamento", "Funcionários", "Despesa", 6200.0),
        (date.today().replace(day=12), "Pagamento", "Internet", "Despesa", 199.0),
        (date.today().replace(day=15), "Pagamento", "Limpeza", "Despesa", 450.0),
    ]
    for dt, tipo, cat, origem, valor in despesas_seed:
        lid += 1
        seed.append((lid, 1, 1, dt, tipo, origem, cat, f"PAG-{lid:06d}", valor, 0.0, "financeiro"))

    for r_idx, row in enumerate(seed, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx == 4 and isinstance(val, date):
                _date_fmt(cell)
            if c_idx in (9, 10):
                _money(cell)

    last = 1 + max(len(seed), 1)
    if not seed:
        for c in range(1, 12):
            ws.cell(row=2, column=c).value = None
        last = 2
    else:
        for c in range(1, 12):
            ws.cell(row=last + 1, column=c).value = None
        last = last + 1

    _make_table(ws, "tbLancamentos", f"A1:K{last}")
    set_column_widths(ws, {1: 6, 2: 10, 3: 10, 4: 12, 5: 12, 6: 14, 7: 12, 8: 14, 9: 12, 10: 12, 11: 12})
    ws.freeze_panes = "A2"


def build_bd_fluxo_caixa(wb) -> None:
    """Fluxo append-only — nunca altera linhas antigas."""
    ws = wb.create_sheet("BD_FLUXO_CAIXA")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False
    headers = ["Data", "Tipo", "Categoria", "Descrição", "Entrada", "Saída", "Saldo"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, 7)

    saldo = 25000.0  # alinhado ao saldo inicial típico da config
    movs = []
    for m in gerar_mensalidades():
        if m["status"] == "Pago" and m["pagamento"] is not None:
            movs.append((m["pagamento"], "Entrada", "Mensalidades", f"Receb. {m['aluno']}", m["valor"], 0.0))
    movs.extend(
        [
            (date.today().replace(day=5), "Saída", "Aluguel", "Aluguel mensal", 0.0, 3500.0),
            (date.today().replace(day=8), "Saída", "Energia", "Conta energia", 0.0, 890.0),
            (date.today().replace(day=10), "Saída", "Funcionários", "Folha", 0.0, 6200.0),
        ]
    )
    movs.sort(key=lambda x: x[0])

    r_idx = 1
    for dt, tipo, cat, desc, entrada, saida in movs[:40]:
        r_idx += 1
        saldo = saldo + entrada - saida
        vals = (dt, tipo, cat, desc, entrada, saida, saldo)
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx == 1:
                _date_fmt(cell)
            if c_idx in (5, 6, 7):
                _money(cell)

    last = max(r_idx, 2)
    if r_idx == 1:
        for c in range(1, 8):
            ws.cell(row=2, column=c).value = None
        last = 2
    else:
        for c in range(1, 8):
            ws.cell(row=last + 1, column=c).value = None
        last = last + 1

    _make_table(ws, "tbFluxoCaixaBD", f"A1:G{last}")
    set_column_widths(ws, {1: 12, 2: 10, 3: 14, 4: 28, 5: 12, 6: 12, 7: 12})
    ws.freeze_panes = "A2"


def build_bd_contas_pagar(wb) -> None:
    ws = wb.create_sheet("BD_CONTAS_PAGAR")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False
    headers = ["ID", "EmpresaID", "UnidadeID", "Fornecedor", "Categoria", "Valor", "Vencimento", "Pagamento", "Situação", "Observação"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    for r_idx, c in enumerate(CONTAS_PAGAR, start=2):
        vals = (
            r_idx - 1,
            1,
            1,
            c["fornecedor"],
            c["categoria"],
            c["valor"],
            c["vencimento"],
            None,
            c["situacao"],
            c.get("descricao", ""),
        )
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx == 6:
                _money(cell)
            if c_idx in (7, 8) and isinstance(val, date):
                _date_fmt(cell)

    last = 1 + max(len(CONTAS_PAGAR), 1)
    for c in range(1, 11):
        ws.cell(row=last + 1, column=c).value = None
    last = last + 1
    _make_table(ws, "tbContasPagarBD", f"A1:J{last}")
    set_column_widths(ws, {1: 6, 2: 10, 3: 10, 4: 22, 5: 14, 6: 12, 7: 12, 8: 12, 9: 12, 10: 24})
    ws.freeze_panes = "A2"


def build_bd_parametros(wb) -> None:
    """Painel de configuração — Grupo | Parâmetro | Valor | Descrição."""
    ws = wb.create_sheet("BD_PARAMETROS")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False

    headers = ["Grupo", "Parâmetro", "Valor", "Descrição"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, 4)

    for r_idx, (grupo, param, valor, desc) in enumerate(PARAMETROS_SEED, start=2):
        ws.cell(row=r_idx, column=1, value=grupo).border = border_thin
        ws.cell(row=r_idx, column=2, value=param).border = border_thin
        cell = ws.cell(row=r_idx, column=3, value=valor)
        cell.border = border_thin
        cell.font = font_normal
        if param == "MetaReceitaMes":
            _money(cell)
        ws.cell(row=r_idx, column=4, value=desc).border = border_thin
        ws.cell(row=r_idx, column=4).font = Font(name="Calibri", size=9, color="666666")

    last = 1 + len(PARAMETROS_SEED)
    _make_table(ws, "tbParametros", f"A1:D{last}")
    ws.cell(row=last + 2, column=1, value="ObterParametro(\"Grupo\",\"Parâmetro\") — altere valores sem editar VBA.")
    ws.cell(row=last + 2, column=1).font = Font(name="Calibri", size=9, color=BRAND_RED)
    set_column_widths(ws, {1: 12, 2: 26, 3: 16, 4: 40})
    ws.freeze_panes = "A2"


def build_bd_planos(wb) -> None:
    ws = wb.create_sheet("BD_PLANOS")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False
    headers = ["ID", "Plano", "Valor", "Matrícula", "Vencimento", "Fidelidade", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, 7)

    for r_idx, row in enumerate(PLANOS_BD_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx in (3, 4):
                _money(cell)

    last = 1 + len(PLANOS_BD_SEED)
    _make_table(ws, "tbPlanos", f"A1:G{last}")
    set_column_widths(ws, {1: 6, 2: 14, 3: 12, 4: 12, 5: 12, 6: 12, 7: 10})
    ws.freeze_panes = "A2"


def build_bd_formas_pagamento(wb) -> None:
    ws = wb.create_sheet("BD_FORMAS_PAGAMENTO")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False
    headers = ["ID", "Forma", "Taxa", "CompensaEmDias"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, 4)

    for r_idx, row in enumerate(FORMAS_BD_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx == 3:
                cell.number_format = "0.0"

    last = 1 + len(FORMAS_BD_SEED)
    _make_table(ws, "tbFormasPagamento", f"A1:D{last}")
    set_column_widths(ws, {1: 6, 2: 16, 3: 10, 4: 16})
    ws.freeze_panes = "A2"


def build_bd_status(wb) -> None:
    ws = wb.create_sheet("BD_STATUS")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False
    headers = ["Tipo", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, 2)

    for r_idx, (tipo, status) in enumerate(STATUS_BD_SEED, start=2):
        ws.cell(row=r_idx, column=1, value=tipo).border = border_thin
        ws.cell(row=r_idx, column=2, value=status).border = border_thin

    last = 1 + len(STATUS_BD_SEED)
    _make_table(ws, "tbStatus", f"A1:B{last}")
    set_column_widths(ws, {1: 14, 2: 14})
    ws.freeze_panes = "A2"


def build_bd_permissoes(wb) -> None:
    ws = wb.create_sheet("BD_PERMISSOES")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    headers = [
        "Perfil",
        "Dashboard",
        "Financeiro",
        "Estoque",
        "Configuração",
        "Alunos",
        "Mensalidades",
        "Relatorios",
        "Excluir",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    for r_idx, row in enumerate(PERMISSOES_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin

    last = 1 + len(PERMISSOES_SEED)
    _make_table(ws, "tbPermissoes", f"A1:I{last}")
    set_column_widths(ws, {1: 14, 2: 12, 3: 12, 4: 10, 5: 14, 6: 10, 7: 14, 8: 12, 9: 10})
    ws.freeze_panes = "A2"


def build_bd_cores(wb) -> None:
    ws = wb.create_sheet("BD_CORES")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    headers = ["Item", "Cor"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, 2)

    for r_idx, (item, cor) in enumerate(CORES_SEED, start=2):
        ws.cell(row=r_idx, column=1, value=item).border = border_thin
        ws.cell(row=r_idx, column=2, value=cor).border = border_thin
        ws.cell(row=r_idx, column=2).number_format = "@"

    last = 1 + len(CORES_SEED)
    _make_table(ws, "tbCores", f"A1:B{last}")
    set_column_widths(ws, {1: 12, 2: 12})
    ws.freeze_panes = "A2"


def build_versao(wb) -> None:
    ws = wb.create_sheet("VERSAO")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    headers = ["Versão", "Data", "Alterações"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, 3)

    for r_idx, (ver, dt, alt) in enumerate(VERSAO_SEED, start=2):
        ws.cell(row=r_idx, column=1, value=ver).border = border_thin
        cell = ws.cell(row=r_idx, column=2, value=dt)
        cell.border = border_thin
        _date_fmt(cell)
        ws.cell(row=r_idx, column=3, value=alt).border = border_thin

    last = 1 + len(VERSAO_SEED)
    _make_table(ws, "tbVersao", f"A1:C{last}")
    set_column_widths(ws, {1: 10, 2: 12, 3: 60})
    ws.freeze_panes = "A2"
