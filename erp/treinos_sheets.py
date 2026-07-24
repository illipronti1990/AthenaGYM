"""Sprint 7.0 — Treinos + Avaliacao Fisica (BD e telas)."""

from __future__ import annotations

from datetime import date

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    apply_header_row,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_normal,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

ANO = date.today().year

AVAL_HEADERS = [
    "ID", "Matrícula", "Nome", "Data", "Professor", "Peso", "Altura", "IMC",
    "Gordura Corporal", "Massa Magra", "Massa Gorda", "Objetivo", "Observações",
]
MEDIDAS_HEADERS = [
    "ID", "AvaliacaoID", "Peitoral", "Cintura", "Abdômen", "Quadril",
    "Bíceps Direito", "Bíceps Esquerdo", "Antebraço", "Coxa", "Panturrilha", "Ombros",
]
TREINO_HEADERS = [
    "ID", "Matrícula", "Nome", "Professor", "Tipo", "Objetivo", "Divisão",
    "Data Início", "Data Fim", "Versão", "Status",
]
EXERC_HEADERS = ["Código", "Exercício", "Grupo Muscular", "Equipamento"]
FICHA_HEADERS = [
    "ID", "TreinoID", "Dia", "Grupo", "ExercicioCodigo", "Exercício",
    "Séries", "Repetições", "Ordem", "Observação",
]
FOTOS_HEADERS = ["ID", "Matrícula", "Data", "Frente", "Lado", "Costas"]

EXERCICIOS_SEED = [
    ("EX001", "Supino Reto", "Peito", "Banco"),
    ("EX002", "Supino Inclinado", "Peito", "Banco"),
    ("EX003", "Crucifixo", "Peito", "Halteres"),
    ("EX004", "Puxada Frontal", "Costas", "Pulley"),
    ("EX005", "Remada Curvada", "Costas", "Barra"),
    ("EX006", "Agachamento Livre", "Pernas", "Barra"),
    ("EX007", "Leg Press", "Pernas", "Leg Press"),
    ("EX008", "Extensão de Joelho", "Pernas", "Cadeira Extensora"),
    ("EX009", "Desenvolvimento", "Ombros", "Halteres"),
    ("EX010", "Elevação Lateral", "Ombros", "Halteres"),
    ("EX011", "Rosca Direta", "Bíceps", "Barra"),
    ("EX012", "Rosca Alternada", "Bíceps", "Halteres"),
    ("EX013", "Tríceps Pulley", "Tríceps", "Pulley"),
    ("EX014", "Tríceps Francês", "Tríceps", "Halteres"),
    ("EX015", "Prancha", "Core", "Peso Corporal"),
    ("EX016", "Abdominal supra", "Core", "Peso Corporal"),
    ("EX017", "Stiff", "Posterior", "Barra"),
    ("EX018", "Panturrilha em pé", "Panturrilha", "Máquina"),
    ("EX019", "Flexão de Braço", "Peito", "Peso Corporal"),
    ("EX020", "Barra Fixa", "Costas", "Barra"),
]

AVAL_SEED = [
    (1, f"ATH-{ANO}-000001", "Mariana Oliveira", date(2026, 1, 10), "Juliana Costa",
     68.5, 1.68, round(68.5 / (1.68 ** 2), 1), 24.2, 48.5, round(68.5 - 48.5, 1), "Emagrecimento", "Avaliacao inicial"),
    (2, f"ATH-{ANO}-000001", "Mariana Oliveira", date(2026, 4, 10), "Juliana Costa",
     66.2, 1.68, round(66.2 / (1.68 ** 2), 1), 22.8, 50.1, round(66.2 - 50.1, 1), "Emagrecimento", "Reducao de gordura"),
    (3, f"ATH-{ANO}-000002", "Pedro Henrique Alves", date(2026, 2, 5), "Juliana Costa",
     82.0, 1.78, round(82.0 / (1.78 ** 2), 1), 18.5, 66.0, round(82.0 - 66.0, 1), "Hipertrofia", "Foco hipertrofia"),
    (4, f"ATH-{ANO}-000003", "Fernanda Ribeiro", date(2026, 3, 1), "Juliana Costa",
     59.0, 1.62, round(59.0 / (1.62 ** 2), 1), 26.0, 42.0, round(59.0 - 42.0, 1), "Saude", "Inicio protocolo"),
    (5, f"ATH-{ANO}-000004", "Thiago Ferreira", date(2026, 1, 25), "Rafael Lima",
     90.5, 1.80, round(90.5 / (1.80 ** 2), 1), 16.0, 74.0, round(90.5 - 74.0, 1), "Hipertrofia", "Personal 3x"),
    (6, f"ATH-{ANO}-000005", "Beatriz Nogueira", date(2026, 3, 8), "Juliana Costa",
     71.0, 1.70, round(71.0 / (1.70 ** 2), 1), 27.5, 49.0, round(71.0 - 49.0, 1), "Emagrecimento", "Primeira avaliacao"),
]

MEDIDAS_SEED = [
    (1, 1, 92, 74, 82, 98, 28, 27, 24, 56, 34, 108),
    (2, 2, 90, 70, 78, 96, 28, 27, 24, 55, 34, 106),
    (3, 3, 102, 82, 88, 100, 36, 35, 28, 60, 38, 118),
    (4, 4, 88, 68, 76, 94, 26, 25, 22, 52, 32, 102),
    (5, 5, 110, 86, 92, 104, 38, 37, 30, 64, 40, 122),
    (6, 6, 94, 78, 86, 100, 29, 28, 25, 57, 35, 110),
]

TREINO_SEED = [
    (1, f"ATH-{ANO}-000001", "Mariana Oliveira", "Juliana Costa", "Emagrecimento", "Perder gordura", "ABC",
     date(2026, 1, 15), date(2026, 4, 15), 1, "Inativo"),
    (2, f"ATH-{ANO}-000001", "Mariana Oliveira", "Juliana Costa", "Emagrecimento", "Manter evolucao", "ABC",
     date(2026, 4, 16), None, 2, "Ativo"),
    (3, f"ATH-{ANO}-000002", "Pedro Henrique Alves", "Juliana Costa", "Hipertrofia", "Ganho de massa", "ABCD",
     date(2026, 2, 10), None, 1, "Ativo"),
    (4, f"ATH-{ANO}-000004", "Thiago Ferreira", "Rafael Lima", "Hipertrofia", "Performance", "ABCDE",
     date(2026, 2, 1), None, 1, "Ativo"),
]

FICHA_SEED = [
    (1, 2, "A", "Peito", "EX001", "Supino Reto", 4, "12", 1, ""),
    (2, 2, "A", "Peito", "EX002", "Supino Inclinado", 3, "10", 2, ""),
    (3, 2, "A", "Peito", "EX003", "Crucifixo", 3, "15", 3, ""),
    (4, 2, "A", "Tríceps", "EX013", "Tríceps Pulley", 3, "12", 4, ""),
    (5, 2, "A", "Tríceps", "EX014", "Tríceps Francês", 3, "10", 5, ""),
    (6, 2, "B", "Costas", "EX004", "Puxada Frontal", 4, "12", 1, ""),
    (7, 2, "B", "Costas", "EX005", "Remada Curvada", 3, "10", 2, ""),
    (8, 2, "B", "Bíceps", "EX011", "Rosca Direta", 3, "12", 3, ""),
    (9, 2, "C", "Pernas", "EX006", "Agachamento Livre", 4, "10", 1, ""),
    (10, 2, "C", "Pernas", "EX007", "Leg Press", 3, "12", 2, ""),
    (11, 2, "C", "Pernas", "EX008", "Extensão de Joelho", 3, "15", 3, ""),
    (12, 3, "A", "Peito", "EX001", "Supino Reto", 4, "8-10", 1, "Carga progressiva"),
    (13, 3, "A", "Ombros", "EX009", "Desenvolvimento", 3, "10", 2, ""),
    (14, 3, "B", "Costas", "EX020", "Barra Fixa", 4, "max", 1, ""),
    (15, 3, "C", "Pernas", "EX006", "Agachamento Livre", 5, "5", 1, "Forca"),
]

FOTOS_SEED = [
    (1, f"ATH-{ANO}-000001", date(2026, 1, 10),
     r"C:\ATHENAS\Fotos\ATH2026001\2026-01-10_Frente.jpg",
     r"C:\ATHENAS\Fotos\ATH2026001\2026-01-10_Lado.jpg",
     r"C:\ATHENAS\Fotos\ATH2026001\2026-01-10_Costas.jpg"),
    (2, f"ATH-{ANO}-000001", date(2026, 4, 10),
     r"C:\ATHENAS\Fotos\ATH2026001\2026-04-10_Frente.jpg",
     r"C:\ATHENAS\Fotos\ATH2026001\2026-04-10_Lado.jpg",
     r"C:\ATHENAS\Fotos\ATH2026001\2026-04-10_Costas.jpg"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def _bd_sheet(wb, name: str, headers: list, rows: list, table: str, widths: dict, date_cols=()) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
            cell.border = border_thin
            cell.font = font_normal
            if c_idx in date_cols and val:
                _date_fmt(cell)
    last = 1 + len(rows)
    for c in range(1, len(headers) + 1):
        ws.cell(row=last + 1, column=c).border = border_thin
    end = get_column_letter(len(headers))
    _make_table(ws, table, f"A1:{end}{last + 1}")
    set_column_widths(ws, widths)
    ws.freeze_panes = "A2"


def build_bd_avaliacoes(wb) -> None:
    _bd_sheet(
        wb, "BD_AVALIACOES", AVAL_HEADERS, AVAL_SEED, "tbAvaliacoes",
        {1: 6, 2: 16, 3: 22, 4: 12, 5: 16, 6: 8, 7: 8, 8: 8, 9: 12, 10: 10, 11: 10, 12: 14, 13: 24},
        date_cols={4},
    )


def build_bd_medidas(wb) -> None:
    _bd_sheet(
        wb, "BD_MEDIDAS", MEDIDAS_HEADERS, MEDIDAS_SEED, "tbMedidas",
        {1: 6, 2: 10, 3: 10, 4: 10, 5: 10, 6: 10, 7: 12, 8: 12, 9: 10, 10: 8, 11: 11, 12: 10},
    )


def build_bd_treinos(wb) -> None:
    _bd_sheet(
        wb, "BD_TREINOS", TREINO_HEADERS, TREINO_SEED, "tbTreinos",
        {1: 6, 2: 16, 3: 22, 4: 16, 5: 14, 6: 18, 7: 10, 8: 12, 9: 12, 10: 8, 11: 10},
        date_cols={8, 9},
    )


def build_bd_exercicios(wb) -> None:
    _bd_sheet(
        wb, "BD_EXERCICIOS", EXERC_HEADERS, EXERCICIOS_SEED, "tbExercicios",
        {1: 10, 2: 22, 3: 16, 4: 16},
    )


def build_bd_treino_itens(wb) -> None:
    _bd_sheet(
        wb, "BD_TREINO_ITENS", FICHA_HEADERS, FICHA_SEED, "tbTreinoItens",
        {1: 6, 2: 10, 3: 6, 4: 12, 5: 12, 6: 20, 7: 8, 8: 10, 9: 8, 10: 18},
    )


def build_bd_fotos(wb) -> None:
    _bd_sheet(
        wb, "BD_FOTOS", FOTOS_HEADERS, FOTOS_SEED, "tbFotos",
        {1: 6, 2: 16, 3: 12, 4: 42, 5: 42, 6: 42},
        date_cols={3},
    )


def build_avaliacao_ui(wb) -> None:
    """24_AVALIACAO — centro de avaliacao fisica."""
    ws = wb.create_sheet("24_AVALIACAO")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=58, cols=14)
    add_sidebar(ws, active="24_AVALIACAO", rows=58, labels=False)
    add_top_bar(ws, start_col=2, end_col=13)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 12, 4: 14, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 14, 11: 12, 12: 12},
    )
    _title(ws, "AVALIACAO FISICA")

    paint_kpi_card(ws, 7, 3, "Avaliacoes mes", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Reavaliacoes", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Alunos avaliados", 0, False, 2)
    paint_kpi_card(ws, 7, 9, "Delta peso medio", "0 kg", False, 2)

    ws.merge_cells("C11:F11")
    ws["C11"] = "FICHA DA AVALIACAO"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    labels = [
        (12, "Matricula"), (13, "Aluno"), (14, "Professor"), (15, "Data"),
        (16, "Peso (kg)"), (17, "Altura (m)"), (18, "IMC"), (19, "% Gordura"),
        (20, "Massa Magra"), (21, "Massa Gorda"), (22, "Objetivo"), (23, "Observacoes"),
    ]
    for r, lab in labels:
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).font = font_header
        ws.cell(row=r, column=4).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin

    ws.merge_cells("G11:K11")
    ws["G11"] = "MEDIDAS (cm)"
    ws["G11"].font = font_section
    ws["G11"].fill = fill_brand
    meds = [
        "Peitoral", "Cintura", "Abdomen", "Quadril", "Biceps D", "Biceps E",
        "Antebraco", "Coxa", "Panturrilha", "Ombros",
    ]
    for i, m in enumerate(meds):
        r = 12 + i
        ws.cell(row=r, column=7, value=m).fill = fill_panel
        ws.cell(row=r, column=7).border = border_thin
        ws.cell(row=r, column=8).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin

    ws.merge_cells("C25:K25")
    ws["C25"] = "HISTORICO DE AVALIACOES (selecione a linha)"
    ws["C25"].font = font_section
    ws["C25"].fill = fill_brand
    for i, h in enumerate(["ID", "Data", "Aluno", "Peso", "IMC", "Gordura", "Professor", "Objetivo"], start=3):
        ws.cell(row=26, column=i, value=h).fill = fill_gold
        ws.cell(row=26, column=i).font = font_header
    for i in range(10):
        r = 27 + i
        for c in range(3, 11):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C38:F38")
    ws["C38"] = "EVOLUCAO — PESO"
    ws["C38"].font = font_section
    ws["C38"].fill = fill_brand
    for i, h in enumerate(["Mes", "Peso", "Barra"], start=3):
        ws.cell(row=39, column=i, value=h).fill = fill_gold
        ws.cell(row=39, column=i).font = font_header
    for i in range(8):
        r = 40 + i
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("G38:K38")
    ws["G38"] = "COMPARACAO (Avaliacao A x B)"
    ws["G38"].font = font_section
    ws["G38"].fill = fill_brand
    ws["G39"] = "ID Aval. A"
    ws["H39"] = ""
    ws["I39"] = "ID Aval. B"
    ws["J39"] = ""
    for c in range(7, 11):
        ws.cell(row=39, column=c).fill = fill_panel
        ws.cell(row=39, column=c).border = border_thin
    for i, h in enumerate(["Indicador", "A", "B", "Delta"]):
        ws.cell(row=40, column=7 + i, value=h).fill = fill_gold
        ws.cell(row=40, column=7 + i).font = font_header
    for i in range(6):
        r = 41 + i
        for c in range(7, 11):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C50:K50")
    ws["C50"] = "Dados append-only · Reavaliacao automatica em 60 dias · Fotos = caminho do arquivo"
    ws["C50"].font = Font(name="Calibri", size=9, color="666666")


def build_treinos_ui(wb) -> None:
    """25_TREINOS — fichas e biblioteca."""
    ws = wb.create_sheet("25_TREINOS")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=66, cols=13)
    add_sidebar(ws, active="25_TREINOS", rows=66, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 8, 4: 18, 5: 12, 6: 14, 7: 10, 8: 10, 9: 10, 10: 12, 11: 12},
    )
    _title(ws, "FICHAS DE TREINO")

    paint_kpi_card(ws, 7, 3, "Treinos ativos", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Biblioteca", len(EXERCICIOS_SEED), False, 2)
    paint_kpi_card(ws, 7, 7, "Versoes", 0, False, 2)
    paint_kpi_card(ws, 7, 9, "Alunos c/ ficha", 0, False, 2)

    ws.merge_cells("C11:G11")
    ws["C11"] = "TREINOS (selecione)"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    for i, h in enumerate(["ID", "Aluno", "Tipo", "Divisao", "Versao", "Status", "Inicio"], start=3):
        ws.cell(row=12, column=i, value=h).fill = fill_gold
        ws.cell(row=12, column=i).font = font_header
    for i in range(10):
        r = 13 + i
        for c in range(3, 10):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C24:J24")
    ws["C24"] = "FICHA DO TREINO SELECIONADO"
    ws["C24"].font = font_section
    ws["C24"].fill = fill_brand
    for i, h in enumerate(["Dia", "Grupo", "Exercicio", "Series", "Reps", "Obs"], start=3):
        ws.cell(row=25, column=i, value=h).fill = fill_gold
        ws.cell(row=25, column=i).font = font_header
    for i in range(15):
        r = 26 + i
        for c in range(3, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    # Espaço reservado para botões (C41–C42) — evita sobrepor a biblioteca
    ws.merge_cells("C41:J41")
    ws["C41"] = ""
    ws.row_dimensions[41].height = 36
    ws.row_dimensions[42].height = 36

    ws.merge_cells("C44:G44")
    ws["C44"] = "BIBLIOTECA DE EXERCICIOS"
    ws["C44"].font = font_section
    ws["C44"].fill = fill_brand
    for i, h in enumerate(["Codigo", "Exercicio", "Grupo", "Equipamento"], start=3):
        ws.cell(row=45, column=i, value=h).fill = fill_gold
        ws.cell(row=45, column=i).font = font_header
    for i, ex in enumerate(EXERCICIOS_SEED[:12]):
        r = 46 + i
        for c, v in enumerate(ex, start=3):
            ws.cell(row=r, column=c, value=v).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C59:J59")
    ws["C59"] = "Divisoes: Full Body · AB · ABC · ABCD · ABCDE · Personalizado · Cada alteracao gera nova versao"
    ws["C59"].font = Font(name="Calibri", size=9, color="666666")
