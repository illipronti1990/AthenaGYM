"""Sprint 6.0 — CRM Inteligente: BD + telas operacionais e dashboard."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    apply_header_row,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_normal,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()

LEADS_HEADERS = [
    "ID",
    "Nome",
    "Telefone",
    "Email",
    "Origem",
    "Interesse",
    "Data Cadastro",
    "Responsável",
    "Status",
    "Observações",
    "Próxima Ação",
    "Data Próxima",
    "Matrícula",
    "UnidadeID",
]

HIST_HEADERS = ["ID", "LeadID", "Data", "Hora", "Tipo", "Descrição", "Usuário"]
RETENCAO_HEADERS = ["Matrícula", "Nome", "Motivo", "Data", "Responsável", "Status"]
CAMPANHAS_HEADERS = ["ID", "Nome", "Público", "Data Início", "Data Fim", "Resultado", "Status"]
INDICACOES_HEADERS = [
    "ID",
    "Aluno",
    "Matrícula Aluno",
    "Indicado",
    "Telefone Indicado",
    "Status",
    "Bônus",
    "Data",
]

LEADS_SEED = [
    (1, "João Pedro Silva", "(11) 98888-1001", "joao.p@email.com", "Instagram", "Musculação",
     HOJE - timedelta(days=8), "Lucas", "Novo", "Pediu info de planos", "Ligar", HOJE, "", 1),
    (2, "Maria Fernanda Costa", "(11) 97777-2002", "maria.f@email.com", "Indicação", "Funcional",
     HOJE - timedelta(days=6), "Maria", "Contatado", "Indicada pela aluna Ana", "Confirmar aula", HOJE, "", 1),
    (3, "Carlos Eduardo Lima", "(11) 96666-3003", "carlos.e@email.com", "Google", "Personal",
     HOJE - timedelta(days=5), "Lucas", "Agendado", "Aula experimental marcada", "Aula experimental", HOJE, "", 2),
    (4, "Ana Beatriz Rocha", "(11) 95555-4004", "ana.b@email.com", "Instagram", "Musculação",
     HOJE - timedelta(days=12), "João", "Convertido", "Matriculou Premium", "", None, "ATH-2026-000001", 1),
    (5, "Pedro Henrique Alves", "(11) 94444-5005", "pedro.h@email.com", "Facebook", "Musculação",
     HOJE - timedelta(days=4), "Maria", "Contatado", "Aguardando proposta", "Enviar proposta", HOJE, "", 1),
    (6, "Juliana Souza", "(11) 93333-6006", "ju.souza@email.com", "Site", "Funcional",
     HOJE - timedelta(days=10), "Lucas", "Perdido", "Escolheu concorrente", "", None, "", 2),
    (7, "Rafael Mendes", "(11) 92222-7007", "rafa.m@email.com", "Instagram", "Musculação",
     HOJE - timedelta(days=2), "João", "Novo", "WhatsApp inbound", "Ligar", HOJE, "", 1),
    (8, "Camila Dias", "(11) 91111-8008", "camila.d@email.com", "Indicação", "VIP",
     HOJE - timedelta(days=3), "Maria", "Agendado", "Experimental 18h", "Confirmar aula", HOJE, "", 2),
    (9, "Bruno Carvalho", "(11) 90000-9009", "bruno.c@email.com", "Google", "Musculação",
     HOJE - timedelta(days=15), "Lucas", "Contatado", "Retorno marcado", "Retorno", HOJE, "", 1),
    (10, "Larissa Nunes", "(11) 98989-1010", "lari.n@email.com", "Instagram", "Personal",
     HOJE - timedelta(days=1), "João", "Novo", "Interessada em personal", "Ligar", HOJE, "", 2),
]

HIST_SEED = [
    (1, 1, HOJE - timedelta(days=8), "09:00", "Cadastro", "Lead cadastrado via Instagram", "Sistema"),
    (2, 2, HOJE - timedelta(days=6), "10:15", "Cadastro", "Lead por indicação", "Sistema"),
    (3, 2, HOJE - timedelta(days=5), "14:00", "Ligação", "Ligação realizada — interessada", "Maria"),
    (4, 3, HOJE - timedelta(days=5), "11:00", "Cadastro", "Lead via Google Ads", "Sistema"),
    (5, 3, HOJE - timedelta(days=4), "16:30", "Agendamento", "Aula experimental agendada", "Lucas"),
    (6, 4, HOJE - timedelta(days=12), "08:40", "Cadastro", "Lead Instagram", "Sistema"),
    (7, 4, HOJE - timedelta(days=10), "15:00", "Aula Experimental", "Compareceu à experimental", "João"),
    (8, 4, HOJE - timedelta(days=9), "17:00", "Matrícula", "Convertido em aluno", "João"),
    (9, 5, HOJE - timedelta(days=4), "09:20", "Cadastro", "Lead Facebook", "Sistema"),
    (10, 5, HOJE - timedelta(days=3), "11:00", "Ligação", "Solicitou proposta Premium", "Maria"),
]

RETENCAO_SEED = [
    ("ATH-2025-000088", "Paulo Vieira", "Financeiro", HOJE - timedelta(days=20), "Maria", "Em risco"),
    ("ATH-2025-000092", "Fernanda Lopes", "Falta de tempo", HOJE - timedelta(days=12), "Lucas", "Acompanhamento"),
    ("ATH-2025-000101", "Diego Santos", "Desmotivação", HOJE - timedelta(days=5), "João", "Em risco"),
]

CAMPANHAS_SEED = [
    (1, "Volte a Treinar", "Inativos 30+ dias", HOJE - timedelta(days=10), HOJE + timedelta(days=20), "12 retornos", "Ativa"),
    (2, "Black Friday", "Todos leads + alunos", HOJE - timedelta(days=5), HOJE + timedelta(days=5), "8 matrículas", "Ativa"),
    (3, "Plano Família", "Indicações", HOJE - timedelta(days=40), HOJE - timedelta(days=10), "5 matrículas", "Encerrada"),
    (4, "Indique um Amigo", "Alunos ativos", HOJE - timedelta(days=15), HOJE + timedelta(days=45), "9 indicações", "Ativa"),
]

INDICACOES_SEED = [
    (1, "Ana Beatriz Rocha", "ATH-2026-000001", "Maria Fernanda Costa", "(11) 97777-2002", "Pendente", 50.0, HOJE - timedelta(days=6)),
    (2, "Ana Beatriz Rocha", "ATH-2026-000001", "Camila Dias", "(11) 91111-8008", "Pendente", 50.0, HOJE - timedelta(days=3)),
    (3, "Pedro Henrique", "ATH-2025-000050", "Rafael Mendes", "(11) 92222-7007", "Convertido", 50.0, HOJE - timedelta(days=20)),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _money(cell) -> None:
    cell.number_format = 'R$ #,##0.00'


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def build_bd_leads(wb) -> None:
    ws = wb.create_sheet("BD_LEADS")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(LEADS_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(LEADS_HEADERS))

    for r_idx, row in enumerate(LEADS_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
            cell.border = border_thin
            cell.font = font_normal
            if c_idx in (7, 12) and val:
                _date_fmt(cell)
            if c_idx in (3, 4, 13):
                cell.number_format = "@"

    last = 1 + len(LEADS_SEED)
    for c in range(1, len(LEADS_HEADERS) + 1):
        ws.cell(row=last + 1, column=c).border = border_thin
    _make_table(ws, "tbLeads", f"A1:N{last + 1}")
    set_column_widths(
        ws,
        {1: 6, 2: 22, 3: 16, 4: 22, 5: 12, 6: 14, 7: 12, 8: 12, 9: 12, 10: 28, 11: 16, 12: 12, 13: 16, 14: 10},
    )
    ws.freeze_panes = "A2"


def build_bd_crm_historico(wb) -> None:
    ws = wb.create_sheet("BD_CRM_HISTORICO")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(HIST_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(HIST_HEADERS))

    for r_idx, row in enumerate(HIST_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if c_idx == 3:
                _date_fmt(cell)

    last = 1 + len(HIST_SEED)
    for c in range(1, len(HIST_HEADERS) + 1):
        ws.cell(row=last + 1, column=c).border = border_thin
    _make_table(ws, "tbCrmHistorico", f"A1:G{last + 1}")
    set_column_widths(ws, {1: 6, 2: 8, 3: 12, 4: 8, 5: 16, 6: 36, 7: 12})
    ws.freeze_panes = "A2"


def build_bd_retencao(wb) -> None:
    ws = wb.create_sheet("BD_RETENCAO")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(RETENCAO_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(RETENCAO_HEADERS))

    for r_idx, row in enumerate(RETENCAO_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if c_idx == 4:
                _date_fmt(cell)

    last = 1 + len(RETENCAO_SEED)
    for c in range(1, len(RETENCAO_HEADERS) + 1):
        ws.cell(row=last + 1, column=c).border = border_thin
    _make_table(ws, "tbRetencao", f"A1:F{last + 1}")
    set_column_widths(ws, {1: 16, 2: 20, 3: 16, 4: 12, 5: 12, 6: 16})
    ws.freeze_panes = "A2"


def build_bd_campanhas(wb) -> None:
    ws = wb.create_sheet("BD_CAMPANHAS")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(CAMPANHAS_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(CAMPANHAS_HEADERS))

    for r_idx, row in enumerate(CAMPANHAS_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if c_idx in (4, 5):
                _date_fmt(cell)

    last = 1 + len(CAMPANHAS_SEED)
    for c in range(1, len(CAMPANHAS_HEADERS) + 1):
        ws.cell(row=last + 1, column=c).border = border_thin
    _make_table(ws, "tbCampanhas", f"A1:G{last + 1}")
    set_column_widths(ws, {1: 6, 2: 18, 3: 20, 4: 12, 5: 12, 6: 14, 7: 12})
    ws.freeze_panes = "A2"


def build_bd_indicacoes(wb) -> None:
    ws = wb.create_sheet("BD_INDICACOES")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(INDICACOES_HEADERS, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(INDICACOES_HEADERS))

    for r_idx, row in enumerate(INDICACOES_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if c_idx == 7:
                _money(cell)
            if c_idx == 8:
                _date_fmt(cell)

    last = 1 + len(INDICACOES_SEED)
    for c in range(1, len(INDICACOES_HEADERS) + 1):
        ws.cell(row=last + 1, column=c).border = border_thin
    _make_table(ws, "tbIndicacoes", f"A1:H{last + 1}")
    set_column_widths(ws, {1: 6, 2: 20, 3: 16, 4: 20, 5: 16, 6: 12, 7: 10, 8: 12})
    ws.freeze_panes = "A2"


def build_crm(wb) -> None:
    """22_CRM — operação comercial (leads, funil, agenda, histórico)."""
    ws = wb.create_sheet("22_CRM")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=58, cols=14)
    add_sidebar(ws, active="22_CRM", rows=58, labels=False)
    add_top_bar(ws, start_col=2, end_col=13)
    set_column_widths(
        ws,
        {
            1: 24, 2: 2, 3: 8, 4: 20, 5: 14, 6: 18, 7: 12, 8: 12,
            9: 12, 10: 14, 11: 14, 12: 12, 13: 12,
        },
    )
    _title(ws, "CRM INTELIGENTE — LEADS & FUNIL")

    paint_kpi_card(ws, 7, 3, "Leads mês", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Conversão", "0%", False, 2)
    paint_kpi_card(ws, 7, 7, "Experimentais", 0, False, 2)
    paint_kpi_card(ws, 7, 9, "Em risco", 0, False, 2)

    ws.merge_cells("C11:E11")
    ws["C11"] = "FUNIL DE VENDAS"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    for i, fase in enumerate(["Leads", "Contatados", "Experimentais", "Matrículas", "Ativos"]):
        r = 13 + i
        ws.cell(row=r, column=3, value=fase).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).font = font_header
        cell = ws.cell(row=r, column=4, value=0)
        cell.fill = fill_panel
        cell.border = border_thin
        cell.font = Font(name="Georgia", size=14, bold=True, color=BRAND_RED)
        ws.cell(row=r, column=5, value="").fill = fill_panel
        ws.cell(row=r, column=5).border = border_thin

    ws.merge_cells("G11:K11")
    ws["G11"] = "AGENDA COMERCIAL — HOJE"
    ws["G11"].font = font_section
    ws["G11"].fill = fill_brand
    for i, h in enumerate(["Hora", "Ação", "Lead", "Resp."], start=7):
        ws.cell(row=12, column=i, value=h).fill = fill_gold
        ws.cell(row=12, column=i).font = font_header
    for i in range(8):
        r = 13 + i
        for c in range(7, 11):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C19:K19")
    ws["C19"] = "LEADS (selecione a linha · use os botões)"
    ws["C19"].font = font_section
    ws["C19"].fill = fill_brand

    headers = ["ID", "Nome", "Telefone", "Origem", "Interesse", "Status", "Responsável", "Próxima Ação", "Data"]
    for i, h in enumerate(headers, start=3):
        ws.cell(row=20, column=i, value=h).fill = fill_gold
        ws.cell(row=20, column=i).font = font_header
    for i in range(15):
        r = 21 + i
        for c in range(3, 12):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin
        ws.row_dimensions[r].height = 18

    ws.merge_cells("C37:G37")
    ws["C37"] = "HISTÓRICO DO LEAD"
    ws["C37"].font = font_section
    ws["C37"].fill = fill_brand
    for i, h in enumerate(["Data", "Tipo", "Descrição"], start=3):
        ws.cell(row=38, column=i, value=h).fill = fill_gold
        ws.cell(row=38, column=i).font = font_header
    for i in range(8):
        r = 39 + i
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("H37:K37")
    ws["H37"] = "ALUNOS EM RISCO"
    ws["H37"].font = font_section
    ws["H37"].fill = fill_brand
    for i, h in enumerate(["Aluno", "Motivo", "Status"], start=8):
        ws.cell(row=38, column=i, value=h).fill = fill_gold
        ws.cell(row=38, column=i).font = font_header
    for i in range(8):
        r = 39 + i
        for c in range(8, 11):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C48:K48")
    ws["C48"] = "Fluxo: Lead → Contato → Experimental → Proposta → Matrícula · Conversão gera aluno + financeiro automaticamente"
    ws["C48"].font = Font(name="Calibri", size=9, color="666666")


def build_dash_crm(wb) -> None:
    """23_DASH_CRM — indicadores, origem, ranking, campanhas."""
    ws = wb.create_sheet("23_DASH_CRM")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=48, cols=13)
    add_sidebar(ws, active="22_CRM", rows=48, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14, 10: 12, 11: 12},
    )
    _title(ws, "DASHBOARD COMERCIAL (CRM)")

    paint_kpi_card(ws, 7, 3, "Leads do mês", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Conversão", "0%", False, 2)
    paint_kpi_card(ws, 7, 7, "Experimentais", 0, False, 2)
    paint_kpi_card(ws, 7, 9, "Matrículas", 0, False, 2)

    ws.merge_cells("C11:E11")
    ws["C11"] = "KPIs COMPLEMENTARES"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    for i, lab in enumerate(["Perdas", "Melhor origem", "Indicações", "Campanhas ativas"]):
        r = 12 + i
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=4, value=0 if i != 1 else "").fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
        ws.cell(row=r, column=4).font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)

    ws.merge_cells("G11:K11")
    ws["G11"] = "ORIGEM DOS LEADS"
    ws["G11"].font = font_section
    ws["G11"].fill = fill_brand
    for i, h in enumerate(["Origem", "Qtd", "Barra"], start=7):
        ws.cell(row=12, column=i, value=h).fill = fill_gold
        ws.cell(row=12, column=i).font = font_header
    for i in range(8):
        r = 13 + i
        for c in range(7, 10):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Leads por origem"
    data = Reference(ws, min_col=8, min_row=12, max_row=20)
    cats = Reference(ws, min_col=7, min_row=13, max_row=20)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 10
    ws.add_chart(chart, "G22")

    ws.merge_cells("C18:E18")
    ws["C18"] = "RANKING COMERCIAL"
    ws["C18"].font = font_section
    ws["C18"].fill = fill_brand
    for i, h in enumerate(["Consultor", "Matrículas", "Barra"], start=3):
        ws.cell(row=19, column=i, value=h).fill = fill_gold
        ws.cell(row=19, column=i).font = font_header
    for i in range(6):
        r = 20 + i
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C28:G28")
    ws["C28"] = "CAMPANHAS"
    ws["C28"].font = font_section
    ws["C28"].fill = fill_brand
    for i, h in enumerate(["Nome", "Público", "Resultado", "Status"], start=3):
        ws.cell(row=29, column=i, value=h).fill = fill_gold
        ws.cell(row=29, column=i).font = font_header
    for i in range(6):
        r = 30 + i
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("H28:K28")
    ws["H28"] = "FUNIL (MÊS)"
    ws["H28"].font = font_section
    ws["H28"].fill = fill_brand
    for i, fase in enumerate(["Leads", "Contatados", "Experimentais", "Matrículas", "Ativos"]):
        r = 30 + i
        ws.cell(row=r, column=8, value=fase).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin
        ws.cell(row=r, column=9, value=0).fill = fill_panel
        ws.cell(row=r, column=9).border = border_thin
        ws.cell(row=r, column=9).font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)
