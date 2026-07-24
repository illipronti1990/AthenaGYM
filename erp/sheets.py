"""Construção das abas do ERP ATHENAS GYM."""

from __future__ import annotations

from datetime import date, datetime, time

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table, TableStyleInfo

from config_data import (
    ALUNOS,
    AVALIACOES,
    BANCOS,
    CATEGORIAS_DESPESA,
    CATEGORIAS_RECEITA,
    CONTAS_PAGAR,
    EQUIPAMENTOS,
    ESPECIALIDADES,
    FORMAS_PAGAMENTO,
    PLANOS,
    PRESENCAS,
    PRODUTOS,
    PROFESSORES,
    SEXOS,
    STATUS_ALUNO,
    STATUS_PAGAMENTO,
    STATUS_PROFESSOR,
    gerar_financeiro,
    gerar_mensalidades,
)
from auth_sheets import build_bd_sessao, build_bd_usuarios, build_log as build_auth_log
from banco_sheets import (
    BD_ALUNOS_HEADERS,
    build_bd_alunos,
    build_bd_contas_pagar,
    build_bd_contas_receber,
    build_bd_cores,
    build_bd_fluxo_caixa,
    build_bd_formas_pagamento,
    build_bd_lancamentos,
    build_bd_parametros,
    build_bd_permissoes,
    build_bd_planos,
    build_bd_status,
    build_versao,
)
from agenda_sheets import build_agenda, build_bd_eventos
from bi_sheets import build_bd_metas, build_bd_notificacoes, build_bi_base
from painel_sheets import build_bd_prioridades, build_home
from crm_sheets import (
    build_bd_campanhas,
    build_bd_crm_historico,
    build_bd_indicacoes,
    build_bd_leads,
    build_bd_retencao,
    build_crm,
    build_dash_crm,
)
from treinos_sheets import (
    build_avaliacao_ui,
    build_bd_avaliacoes,
    build_bd_exercicios,
    build_bd_fotos,
    build_bd_medidas,
    build_bd_treino_itens,
    build_bd_treinos,
    build_treinos_ui,
)
from acesso_sheets import (
    build_acesso_ui,
    build_bd_acessos,
    build_bd_presencas,
    build_dash_frequencia,
)
from estoque_sheets import (
    build_bd_compras,
    build_bd_fornecedores,
    build_bd_kit_itens,
    build_bd_kits,
    build_bd_lotes,
    build_bd_movimentacao_estoque,
    build_bd_produtos,
    build_bd_unidades,
    build_bd_venda_itens,
    build_bd_vendas,
    build_dash_pdv,
    build_inventario_ui,
    build_pdv_ui,
)
from bi_analytics_sheets import (
    build_bd_indicadores,
    build_bd_insights,
    build_bd_previsoes,
    build_bd_risco_retencao,
    build_bi_executivo,
    build_insights_ui,
)
from portal_sheets import (
    build_bd_chat,
    build_bd_desafios,
    build_bd_metas_aluno,
    build_bd_portal_tokens,
    build_bd_push,
    build_portal_aluno,
    build_portal_ops,
    build_portal_professor,
)
from athena_sheets import (
    build_athena_ai_ui,
    build_bd_athena_chat,
    build_bd_recomendacoes,
    build_recomendacoes_ui,
)
from empresa_sheets import (
    build_bd_config_empresa,
    build_bd_empresas,
    build_bd_licencas,
    build_master_ui,
    build_nova_academia_ui,
)
from franquia_sheets import (
    build_bd_contratos_franquia,
    build_bd_franqueadoras,
    build_bd_franqueados,
    build_bd_royalties,
    build_franqueadora_ui,
)
from unidade_sheets import (
    build_bd_parametros_unidade,
    build_bd_professor_unidade,
    build_bd_transferencias,
    build_bd_usuario_unidade,
    build_unidades_ui,
)
from dashboards_bi import (
    build_dash_equipamentos_bi,
    build_dash_estoque_bi,
    build_dash_professores,
)
from premium_ui import (
    build_dashboard,
    build_form_aluno,
    build_login,
    build_relatorios,
)
from styles import (
    BLACK,
    GOLD,
    GREEN_BG,
    ICON,
    RED_BG,
    WHITE,
    YELLOW_BG,
    add_sidebar,
    add_top_bar,
    apply_header_row,
    border_thin,
    center,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    fill_red,
    fill_yellow,
    font_header,
    font_kpi,
    font_kpi_label,
    font_normal,
    font_section,
    paint_canvas,
    paint_kpi_card,
    paint_title_bar,
    set_column_widths,
    style_sheet_tab,
)


def _table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def _dv_list(ws, formula: str, ranges: str, allow_blank: bool = True) -> None:
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=allow_blank,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valor inválido",
        error="Selecione um valor da lista.",
    )
    dv.add(ranges)
    ws.add_data_validation(dv)


def _money(cell) -> None:
    cell.number_format = 'R$ #,##0.00'


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


# ---------------------------------------------------------------------------
# 15 CONFIG
# ---------------------------------------------------------------------------
def build_config(wb) -> None:
    ws = wb.create_sheet("15_CONFIG", 0)
    style_sheet_tab(ws)
    paint_title_bar(ws, "CONFIGURAÇÕES", "Cadastros mestres que alimentam as listas suspensas", 8)
    # Sem add_nav_home — o botão shape "Menu" é único (evita Menu duplicado)

    sections = [
        (1, "PLANOS", ["Plano", "Valor"], [(p, v) for p, v in PLANOS]),
        (3, "FORMAS_PAGAMENTO", ["Forma"], [(f,) for f in FORMAS_PAGAMENTO]),
        (5, "BANCOS", ["Banco"], [(b,) for b in BANCOS]),
        (7, "CAT_RECEITA", ["Categoria"], [(c,) for c in CATEGORIAS_RECEITA]),
        (9, "CAT_DESPESA", ["Categoria"], [(c,) for c in CATEGORIAS_DESPESA]),
        (11, "STATUS_ALUNO", ["Status"], [(s,) for s in STATUS_ALUNO]),
        (13, "STATUS_PAGAMENTO", ["Status"], [(s,) for s in STATUS_PAGAMENTO]),
        (15, "SEXOS", ["Sexo"], [(s,) for s in SEXOS]),
        (17, "ESPECIALIDADES", ["Especialidade"], [(e,) for e in ESPECIALIDADES]),
        (19, "STATUS_PROFESSOR", ["Status"], [(s,) for s in STATUS_PROFESSOR]),
    ]

    ws.cell(row=4, column=1, value="Preencha as listas abaixo. Elas alimentam automaticamente os formulários.")
    ws["A4"].font = font_normal

    start_row = 6
    for col, title, headers, rows in sections:
        cell = ws.cell(row=start_row, column=col, value=title)
        cell.font = font_section
        cell.fill = fill_brand
        for i, h in enumerate(headers):
            hcell = ws.cell(row=start_row + 1, column=col + i, value=h)
            hcell.fill = fill_gold
            hcell.font = font_header
            hcell.alignment = center
            hcell.border = border_thin
        for r_idx, row in enumerate(rows, start=start_row + 2):
            for c_idx, val in enumerate(row):
                c = ws.cell(row=r_idx, column=col + c_idx, value=val)
                c.border = border_thin
                c.font = font_normal
                if isinstance(val, float):
                    _money(c)

    # Bloco de parâmetros operacionais
    ws.cell(row=6, column=21, value="PARAMETROS").font = Font(name="Georgia", size=16, bold=True, color=GOLD)
    ws.cell(row=6, column=21).fill = fill_brand
    params = [
        ("Taxa de Matrícula", 50.00),
        ("Dia padrão vencimento", 10),
        ("Estoque mínimo padrão", 5),
        ("Dias alerta manutenção", 30),
        ("Saldo inicial caixa", 5000.00),
        ("Próximo código aluno", 9),
        ("Meta receita mês (R$)", 50000.00),
    ]
    ws.cell(row=7, column=21, value="Parâmetro").fill = fill_gold
    ws.cell(row=7, column=22, value="Valor").fill = fill_gold
    for i, (nome, valor) in enumerate(params, start=8):
        ws.cell(row=i, column=21, value=nome).border = border_thin
        c = ws.cell(row=i, column=22, value=valor)
        c.border = border_thin
        if isinstance(valor, float):
            _money(c)

    # Sprint 3.4 — duplicidade configurável (não sobrescreve meta V14)
    ws["U15"] = "Bloquear e-mail duplicado"
    ws["V15"] = "NÃO"
    ws["U16"] = "Bloquear telefone duplicado"
    ws["V16"] = "NÃO"
    for r in (15, 16):
        ws.cell(row=r, column=21).border = border_thin
        ws.cell(row=r, column=22).border = border_thin

    ws["U20"] = "Usuário sessão"
    ws["V20"] = ""
    ws["U21"] = "Perfil sessão"
    ws["V21"] = ""
    for r in (20, 21):
        ws.cell(row=r, column=21).border = border_thin
        ws.cell(row=r, column=22).border = border_thin

    ws.cell(row=17, column=21, value="INSTRUÇÕES DE USO").font = font_section
    instrucoes = [
        "1. Cadastre planos, categorias e professores nesta aba.",
        "2. Cadastre alunos pelo FORM_ALUNO (grava em BD_ALUNOS / tbAlunos).",
        "3. Código sugerido: =\"A\"&TEXT(COUNTA([Código])+1,\"000\")",
        "4. Ao marcar mensalidade como PAGO, o financeiro e o dashboard atualizam.",
        "5. Vendas de produtos: lance em 04_FINANCEIRO (Receita/Produtos) e baixe em 09_ESTOQUE.",
        "6. Células com fórmulas estão protegidas no conceito — não sobrescreva colunas calculadas.",
        "7. Use Filtros das tabelas Excel para pesquisar alunos rapidamente.",
        "8. Relatórios: filtre por mês nas abas Financeiro, Fluxo e Dashboards.",
    ]
    for i, texto in enumerate(instrucoes, start=18):
        ws.cell(row=i, column=21, value=texto).font = font_normal

    set_column_widths(ws, {i: 18 for i in range(1, 23)})
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["U"].width = 28
    ws.column_dimensions["V"].width = 14


# ---------------------------------------------------------------------------
# 02 ALUNOS (banco de dados — cadastro via FORM_ALUNO)
# ---------------------------------------------------------------------------
def build_alunos(wb) -> None:
    """Lista visível espelhando BD_ALUNOS (consulta / PDF). Cadastro pelo FORM_ALUNO."""
    from banco_sheets import _aluno_row

    ws = wb.create_sheet("02_ALUNOS")
    style_sheet_tab(ws)
    paint_title_bar(ws, "👤 BASE DE ALUNOS", "Consulta — cadastre pelo FORM_ALUNO (banco: BD_ALUNOS)", 21)
    ws["C3"] = "➕ Abrir formulário de cadastro"
    ws["C3"].font = Font(name="Calibri", size=11, bold=True, color=BLACK)
    ws["C3"].fill = fill_gold
    ws["C3"].hyperlink = Hyperlink(ref="C3", location="'FORM_ALUNO'!A1", tooltip="Formulário")
    ws.sheet_view.showGridLines = False

    headers = list(BD_ALUNOS_HEADERS)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, len(headers))

    for r_idx, aluno in enumerate(ALUNOS, start=6):
        values = _aluno_row(aluno, r_idx - 5)
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if isinstance(val, date):
                _date_fmt(cell)
            if c_idx == 18:
                _money(cell)

    last_data = 5 + len(ALUNOS)
    end_row = last_data + 49
    for r in range(last_data + 1, end_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border_thin

    _table(ws, "tblAlunos", f"A5:U{end_row}")

    ws.cell(
        row=4,
        column=1,
        value="Fonte canônica: BD_ALUNOS (tbAlunos). Esta aba é sincronizada ao salvar no formulário.",
    )
    ws["A4"].font = Font(name="Calibri", size=10, italic=True, color="555555")

    # Status = col U (21)
    _dv_list(ws, "'15_CONFIG'!$A$8:$A$13", f"P6:P{end_row}")
    _dv_list(ws, "'08_PROFESSORES'!$B$6:$B$100", f"Q6:Q{end_row}")
    _dv_list(ws, "'15_CONFIG'!$O$8:$O$12", f"G6:G{end_row}")
    _dv_list(ws, "'15_CONFIG'!$K$8:$K$12", f"U6:U{end_row}")

    ws.conditional_formatting.add(
        f"U6:U{end_row}",
        FormulaRule(formula=['U6="Ativo"'], fill=PatternFill("solid", fgColor=GREEN_BG)),
    )
    ws.conditional_formatting.add(
        f"U6:U{end_row}",
        FormulaRule(formula=['U6="Congelado"'], fill=PatternFill("solid", fgColor=YELLOW_BG)),
    )
    ws.conditional_formatting.add(
        f"U6:U{end_row}",
        FormulaRule(formula=['OR(U6="Cancelado",U6="Inadimplente")'], fill=PatternFill("solid", fgColor=RED_BG)),
    )

    set_column_widths(
        ws,
        {
            1: 6,
            2: 12,
            3: 22,
            4: 15,
            5: 12,
            6: 13,
            7: 10,
            8: 15,
            9: 15,
            10: 22,
            11: 11,
            12: 16,
            13: 8,
            14: 12,
            15: 12,
            16: 12,
            17: 16,
            18: 11,
            19: 14,
            20: 12,
            21: 12,
        },
    )
    ws.freeze_panes = "D6"


# ---------------------------------------------------------------------------
# 03 MENSALIDADES
# ---------------------------------------------------------------------------
def build_mensalidades(wb) -> None:
    ws = wb.create_sheet("03_MENSALIDADES")
    style_sheet_tab(ws)
    paint_title_bar(ws, "💳 MENSALIDADES", "Geração e controle automático de cobranças", 10)

    headers = [
        "Aluno",
        "Código",
        "Competência",
        "Valor",
        "Vencimento",
        "Pagamento",
        "Status",
        "Forma Pagamento",
        "Dias Atraso",
        "Situação Visual",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, len(headers))

    mensalidades = gerar_mensalidades()
    for r_idx, m in enumerate(mensalidades, start=6):
        ws.cell(row=r_idx, column=1, value=m["aluno"])
        ws.cell(row=r_idx, column=2, value=m["codigo"])
        ws.cell(row=r_idx, column=3, value=m["competencia"])
        ws.cell(row=r_idx, column=4, value=m["valor"])
        ws.cell(row=r_idx, column=5, value=m["vencimento"])
        ws.cell(row=r_idx, column=6, value=m["pagamento"] if m["pagamento"] else "")
        ws.cell(row=r_idx, column=7, value=m["status"])
        ws.cell(row=r_idx, column=8, value=m["forma"])
        # Dias atraso
        ws.cell(
            row=r_idx,
            column=9,
            value=f'=IF(OR(G{r_idx}="Pago",E{r_idx}=""),0,MAX(0,TODAY()-E{r_idx}))',
        )
        ws.cell(
            row=r_idx,
            column=10,
            value=f'=IF(G{r_idx}="Pago","🟢 Pago",IF(G{r_idx}="Vence hoje","🟡 Vence hoje",IF(OR(G{r_idx}="Atrasado",I{r_idx}>0),"🔴 Atrasado","⚪ Pendente")))',
        )
        for c in range(1, 11):
            ws.cell(row=r_idx, column=c).border = border_thin
            ws.cell(row=r_idx, column=c).font = font_normal
        _money(ws.cell(row=r_idx, column=4))
        _date_fmt(ws.cell(row=r_idx, column=3))
        _date_fmt(ws.cell(row=r_idx, column=5))
        if m["pagamento"]:
            _date_fmt(ws.cell(row=r_idx, column=6))

    last = 5 + len(mensalidades)
    # Linhas extras
    for r in range(last + 1, last + 40):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(row=r, column=9, value=f'=IF(OR(G{r}="Pago",E{r}=""),0,MAX(0,TODAY()-E{r}))')
        ws.cell(
            row=r,
            column=10,
            value=f'=IF(A{r}="","",IF(G{r}="Pago","🟢 Pago",IF(G{r}="Vence hoje","🟡 Vence hoje",IF(OR(G{r}="Atrasado",I{r}>0),"🔴 Atrasado","⚪ Pendente"))))',
        )

    end_row = last + 39
    _table(ws, "tblMensalidades", f"A5:J{end_row}")
    _dv_list(ws, "'BD_ALUNOS'!$C$2:$C$200", f"A6:A{end_row}")
    _dv_list(ws, "'15_CONFIG'!$M$8:$M$12", f"G6:G{end_row}")
    _dv_list(ws, "'15_CONFIG'!$C$8:$C$20", f"H6:H{end_row}")

    # Legend + totals
    ws.cell(row=4, column=1, value="Ao marcar Status = Pago e preencher Data Pagamento, Dashboards e Fluxo atualizam automaticamente.")
    ws["A4"].font = Font(name="Calibri", size=10, italic=True)

    ws.cell(row=3, column=8, value="Total Pendente/Atrasado:")
    ws.cell(row=3, column=8).font = font_header
    ws.cell(row=3, column=9, value=f'=SUMIFS(D6:D{end_row},G6:G{end_row},"<>Pago")')
    _money(ws.cell(row=3, column=9))
    ws.cell(row=3, column=9).fill = fill_yellow

    ws.conditional_formatting.add(
        f"G6:G{end_row}",
        FormulaRule(formula=['G6="Pago"'], fill=PatternFill("solid", fgColor=GREEN_BG)),
    )
    ws.conditional_formatting.add(
        f"G6:G{end_row}",
        FormulaRule(formula=['G6="Vence hoje"'], fill=PatternFill("solid", fgColor=YELLOW_BG)),
    )
    ws.conditional_formatting.add(
        f"G6:G{end_row}",
        FormulaRule(formula=['G6="Atrasado"'], fill=PatternFill("solid", fgColor=RED_BG)),
    )

    set_column_widths(ws, {1: 24, 2: 10, 3: 12, 4: 12, 5: 12, 6: 12, 7: 12, 8: 16, 9: 12, 10: 16})
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 04 FINANCEIRO
# ---------------------------------------------------------------------------
def build_financeiro(wb) -> None:
    """Hub financeiro — Contas a Receber + KPIs (Sprint 4.0)."""
    from banco_sheets import _competencia_texto, _encargos

    ws = wb.create_sheet("04_FINANCEIRO")
    style_sheet_tab(ws)
    paint_title_bar(ws, "💰 FINANCEIRO", "Motor financeiro — receber, pagar, fluxo e dashboard", 11)

    # KPIs topo (valores em N3:N16 — atualizados por AtualizarDashboardFinanceiro)
    # Legado dashboards: N6=Receita Mês, N7=Despesas, N8=Lucro
    for col, label in ((1, "Receita Hoje"), (3, "Receita Mês"), (5, "Despesas"), (7, "Lucro"), (9, "Saldo"), (11, "Inadimplência")):
        ws.cell(row=3, column=col, value=label).font = Font(name="Calibri", size=9, color="666666")
    ws["A4"] = "=N3"
    ws["C4"] = "=N6"
    ws["E4"] = "=N7"
    ws["G4"] = "=N8"
    ws["I4"] = "=N9"
    ws["K4"] = "=N10"
    for addr in ("A4", "C4", "E4", "G4", "I4"):
        _money(ws[addr])
        ws[addr].fill = fill_gold
        ws[addr].font = font_header
    ws["K4"].number_format = '0.0"%"'
    ws["K4"].fill = fill_gold
    ws["K4"].font = font_header

    ws["M2"] = "MOTOR KPIs"
    ws["M2"].font = font_section
    for r, label, val in (
        (3, "Receita Hoje", 0),
        (4, "(reservado)", 0),
        (5, "(reservado)", 0),
        (6, "Receita Mês", 0),
        (7, "Despesas Mês", 0),
        (8, "Lucro", 0),
        (9, "Saldo", 25000),
        (10, "Inadimplência %", 0),
        (11, "Receita Ano", 0),
        (12, "A Receber", 0),
        (13, "A Pagar", 0),
        (14, "Ticket Médio", 0),
        (15, "Churn %", 0),
        (16, "Prevista", 0),
        (17, "Recebida", 0),
        (18, "Em Atraso", 0),
    ):
        ws.cell(row=r, column=13, value=label)
        c = ws.cell(row=r, column=14, value=val)
        if r in (10, 15):
            c.number_format = "0.0"
        else:
            _money(c)

    ws.cell(row=5, column=1, value="Pesquisar aluno:")
    ws.cell(row=5, column=2, value="")
    ws.cell(row=5, column=2).fill = PatternFill("solid", fgColor="FFFFFF")
    ws.cell(row=5, column=2).border = border_thin

    headers = [
        "ID",
        "Aluno",
        "Competência",
        "Valor Final",
        "Situação",
        "Vencimento",
        "Pagamento",
        "Matrícula",
        "Forma",
        "Valor Original",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=7, column=i, value=h)
    apply_header_row(ws, 7, 1, len(headers))

    r_idx = 8
    rid = 0
    abertos_n = 0
    for m in gerar_mensalidades():
        rid += 1
        if m["status"] == "Pago":
            continue
        abertos_n += 1
        multa, juros, vfinal = _encargos(m["valor"], m["vencimento"], m["status"])
        sit = "Pendente" if m["status"] == "Vence hoje" else m["status"]
        vals = [
            rid,
            m["aluno"],
            _competencia_texto(m["competencia"]),
            vfinal,
            sit,
            m["vencimento"],
            m["pagamento"],
            m["codigo"],
            m["forma"] or "",
            m["valor"],
        ]
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx in (4, 10):
                _money(cell)
            if c_idx in (6, 7) and isinstance(val, date):
                _date_fmt(cell)
        r_idx += 1

    last = 7 + max(abertos_n, 1)
    end_row = last + 60
    for r in range(last + 1, end_row + 1):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = border_thin

    _table(ws, "tblFinanceiro", f"A7:J{end_row}")
    _dv_list(ws, '"Pendente,Pago,Atrasado,Cancelado"', f"E8:E{end_row}")

    set_column_widths(
        ws,
        {1: 6, 2: 22, 3: 14, 4: 12, 5: 12, 6: 12, 7: 12, 8: 14, 9: 10, 10: 12, 13: 14, 14: 12},
    )
    ws.freeze_panes = "A8"


# ---------------------------------------------------------------------------
# 05 FLUXO DE CAIXA
# ---------------------------------------------------------------------------
def build_fluxo(wb) -> None:
    ws = wb.create_sheet("05_FLUXO_CAIXA")
    style_sheet_tab(ws)
    paint_title_bar(ws, "💰 FLUXO DE CAIXA", "Saldo diário, mensal e anual — automático", 8)

    ws.cell(row=4, column=1, value="Saldo Inicial:")
    ws.cell(row=4, column=1).font = font_header
    ws.cell(row=4, column=2, value="='15_CONFIG'!V12")
    _money(ws.cell(row=4, column=2))
    ws.cell(row=4, column=2).fill = fill_gold

    headers = ["Data", "Entradas", "Saídas", "Saldo do Dia", "Saldo Acumulado", "Mês", "Ano", "Observação"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=6, column=i, value=h)
    apply_header_row(ws, 6, 1, 8)

    # Gera últimos 30 dias com fórmulas ligadas ao financeiro
    hoje = date.today()
    for i in range(30):
        r = 7 + i
        dia = hoje - __import__("datetime").timedelta(days=29 - i)
        ws.cell(row=r, column=1, value=dia)
        _date_fmt(ws.cell(row=r, column=1))
        # Entradas / Saídas via SOMARPRODUTO no financeiro
        ws.cell(
            row=r,
            column=2,
            value=f'=SUMIFS(\'04_FINANCEIRO\'!E:E,\'04_FINANCEIRO\'!A:A,A{r},\'04_FINANCEIRO\'!B:B,"Receita")',
        )
        ws.cell(
            row=r,
            column=3,
            value=f'=SUMIFS(\'04_FINANCEIRO\'!E:E,\'04_FINANCEIRO\'!A:A,A{r},\'04_FINANCEIRO\'!B:B,"Despesa")',
        )
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        if i == 0:
            ws.cell(row=r, column=5, value=f"=$B$4+D{r}")
        else:
            ws.cell(row=r, column=5, value=f"=E{r-1}+D{r}")
        ws.cell(row=r, column=6, value=f"=MONTH(A{r})")
        ws.cell(row=r, column=7, value=f"=YEAR(A{r})")
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border_thin
        for c in (2, 3, 4, 5):
            _money(ws.cell(row=r, column=c))

    end_row = 36
    _table(ws, "tblFluxo", f"A6:H{end_row}")

    # Totais
    ws.cell(row=38, column=1, value="INDICADORES").font = font_section
    ws.cell(row=38, column=1).fill = fill_brand
    indicadores = [
        (39, "Entradas (período)", f"=SUM(B7:B{end_row})"),
        (40, "Saídas (período)", f"=SUM(C7:C{end_row})"),
        (41, "Saldo do Mês", f"=SUMIFS(D7:D{end_row},F7:F{end_row},MONTH(TODAY()),G7:G{end_row},YEAR(TODAY()))"),
        (42, "Saldo Anual (aprox.)", f"=SUMIFS(D7:D{end_row},G7:G{end_row},YEAR(TODAY()))"),
        (43, "Saldo Acumulado Atual", f"=E{end_row}"),
    ]
    for row, label, formula in indicadores:
        ws.cell(row=row, column=1, value=label).fill = fill_gold
        ws.cell(row=row, column=1).border = border_thin
        c = ws.cell(row=row, column=2, value=formula)
        c.border = border_thin
        _money(c)

    set_column_widths(ws, {1: 12, 2: 14, 3: 14, 4: 14, 5: 16, 6: 8, 7: 8, 8: 20})
    ws.freeze_panes = "A7"


# ---------------------------------------------------------------------------
# 06 CONTAS A RECEBER
# ---------------------------------------------------------------------------
def build_contas_receber(wb) -> None:
    """Espelho operacional de BD_CONTAS_RECEBER (Sprint 4.0)."""
    from banco_sheets import _competencia_texto, _encargos

    ws = wb.create_sheet("06_CONTAS_RECEBER")
    style_sheet_tab(ws)
    paint_title_bar(ws, "💳 CONTAS A RECEBER", "Espelho do banco — motor em BD_CONTAS_RECEBER", 10)

    headers = [
        "ID",
        "Aluno",
        "Matrícula",
        "Competência",
        "Valor Final",
        "Vencimento",
        "Dias em Atraso",
        "Situação",
        "Forma",
        "Pagamento",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, len(headers))

    r_idx = 6
    rid = 0
    n_abertos = 0
    for m in gerar_mensalidades():
        rid += 1
        if m["status"] == "Pago":
            continue
        n_abertos += 1
        multa, juros, vfinal = _encargos(m["valor"], m["vencimento"], m["status"])
        sit = "Pendente" if m["status"] == "Vence hoje" else m["status"]
        vals = [
            rid,
            m["aluno"],
            m["codigo"],
            _competencia_texto(m["competencia"]),
            vfinal,
            m["vencimento"],
            f"=MAX(0,TODAY()-F{r_idx})",
            sit,
            m["forma"] or "",
            m["pagamento"],
        ]
        for c_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
        _money(ws.cell(row=r_idx, column=5))
        _date_fmt(ws.cell(row=r_idx, column=6))
        if m["pagamento"]:
            _date_fmt(ws.cell(row=r_idx, column=10))
        r_idx += 1

    last = 5 + max(n_abertos, 1)
    end_row = last + 40
    for r in range(last + 1, end_row + 1):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(row=r, column=7, value=f'=IF(F{r}="","",MAX(0,TODAY()-F{r}))')

    _table(ws, "tblContasReceber", f"A5:J{end_row}")

    ws.cell(row=4, column=1, value="Total a receber:")
    ws.cell(row=4, column=2, value="='04_FINANCEIRO'!N12")
    _money(ws.cell(row=4, column=2))
    ws.cell(row=4, column=2).fill = fill_red

    ws.conditional_formatting.add(
        f"G6:G{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor=RED_BG)),
    )

    set_column_widths(ws, {1: 6, 2: 22, 3: 14, 4: 14, 5: 12, 6: 12, 7: 12, 8: 12, 9: 10, 10: 12})
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 07 CONTAS A PAGAR
# ---------------------------------------------------------------------------
def build_contas_pagar(wb) -> None:
    ws = wb.create_sheet("07_CONTAS_PAGAR")
    style_sheet_tab(ws)
    paint_title_bar(ws, "💰 CONTAS A PAGAR", "Fornecedores, vencimentos e situações", 7)

    headers = ["Fornecedor", "Descrição", "Valor", "Vencimento", "Categoria", "Situação", "Dias p/ Vencer"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, 7)

    for r_idx, c in enumerate(CONTAS_PAGAR, start=6):
        ws.cell(row=r_idx, column=1, value=c["fornecedor"])
        ws.cell(row=r_idx, column=2, value=c["descricao"])
        ws.cell(row=r_idx, column=3, value=c["valor"])
        ws.cell(row=r_idx, column=4, value=c["vencimento"])
        ws.cell(row=r_idx, column=5, value=c["categoria"])
        ws.cell(row=r_idx, column=6, value=c["situacao"])
        ws.cell(row=r_idx, column=7, value=f"=D{r_idx}-TODAY()")
        for col in range(1, 8):
            ws.cell(row=r_idx, column=col).border = border_thin
        _money(ws.cell(row=r_idx, column=3))
        _date_fmt(ws.cell(row=r_idx, column=4))

    last = 5 + len(CONTAS_PAGAR)
    end_row = last + 40
    for r in range(last + 1, end_row + 1):
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(row=r, column=7, value=f'=IF(D{r}="","",D{r}-TODAY())')

    _table(ws, "tblContasPagar", f"A5:G{end_row}")
    cats = ",".join(CATEGORIAS_DESPESA)
    _dv_list(ws, f'"{cats}"', f"E6:E{end_row}")
    _dv_list(ws, '"Pendente,Pago,Atrasado,Agendado"', f"F6:F{end_row}")

    ws.cell(row=4, column=1, value="Total pendente:")
    ws.cell(row=4, column=2, value=f'=SUMIFS(C6:C{end_row},F6:F{end_row},"<>Pago")')
    _money(ws.cell(row=4, column=2))
    ws.cell(row=4, column=2).fill = fill_yellow

    ws.conditional_formatting.add(
        f"F6:F{end_row}",
        FormulaRule(formula=['F6="Atrasado"'], fill=PatternFill("solid", fgColor=RED_BG)),
    )
    ws.conditional_formatting.add(
        f"F6:F{end_row}",
        FormulaRule(formula=['F6="Pago"'], fill=PatternFill("solid", fgColor=GREEN_BG)),
    )

    set_column_widths(ws, {1: 22, 2: 28, 3: 12, 4: 12, 5: 18, 6: 12, 7: 14})
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 08 PROFESSORES
# ---------------------------------------------------------------------------
def build_professores(wb) -> None:
    ws = wb.create_sheet("08_PROFESSORES")
    style_sheet_tab(ws)
    paint_title_bar(ws, "🏋 PROFESSORES", "Equipe técnica e valores", 8)

    headers = ["ID", "Nome", "CREF", "Telefone", "Especialidade", "Salário", "Turmas", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, 8)

    for r_idx, p in enumerate(PROFESSORES, start=6):
        ws.cell(row=r_idx, column=1, value=f"P{r_idx-5:03d}")
        ws.cell(row=r_idx, column=2, value=p[0])
        ws.cell(row=r_idx, column=3, value=p[1])
        ws.cell(row=r_idx, column=4, value=p[2])
        ws.cell(row=r_idx, column=5, value=p[3])
        ws.cell(row=r_idx, column=6, value=p[4])
        # Turmas = contagem de alunos ativos do professor
        ws.cell(
            row=r_idx,
            column=7,
            value=f'=COUNTIFS(\'BD_ALUNOS\'!Q:Q,B{r_idx},\'BD_ALUNOS\'!U:U,"Ativo")',
        )
        ws.cell(row=r_idx, column=8, value=p[5])
        for c in range(1, 9):
            ws.cell(row=r_idx, column=c).border = border_thin
        _money(ws.cell(row=r_idx, column=6))

    end_row = 50
    for r in range(6 + len(PROFESSORES), end_row + 1):
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(row=r, column=1, value=f'=IF(B{r}="","","P"&TEXT(ROW()-5,"000"))')
        ws.cell(
            row=r,
            column=7,
            value=f'=IF(B{r}="","",COUNTIFS(\'BD_ALUNOS\'!Q:Q,B{r},\'BD_ALUNOS\'!U:U,"Ativo"))',
        )

    _table(ws, "tblProfessores", f"A5:H{end_row}")
    _dv_list(ws, "'15_CONFIG'!$Q$8:$Q$20", f"E6:E{end_row}")
    _dv_list(ws, "'15_CONFIG'!$S$8:$S$12", f"H6:H{end_row}")

    set_column_widths(ws, {1: 8, 2: 22, 3: 18, 4: 16, 5: 16, 6: 12, 7: 10, 8: 10})
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 09 ESTOQUE
# ---------------------------------------------------------------------------
def build_estoque(wb) -> None:
    ws = wb.create_sheet("09_ESTOQUE")
    style_sheet_tab(ws)
    paint_title_bar(ws, "📦 ESTOQUE", "Produtos · Entrada · Saída · Alerta de mínimo", 12)

    headers = [
        "Código",
        "Produto",
        "Categoria",
        "Qtd Atual",
        "Estoque Mínimo",
        "Custo Unit.",
        "Preço Venda",
        "Entradas",
        "Saídas",
        "Valor Estoque",
        "Margem Unit.",
        "Alerta",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, 12)

    for r_idx, p in enumerate(PRODUTOS, start=6):
        nome, cat, qtd, minimo, custo, venda = p
        ws.cell(row=r_idx, column=1, value=f"PR{r_idx-5:03d}")
        ws.cell(row=r_idx, column=2, value=nome)
        ws.cell(row=r_idx, column=3, value=cat)
        ws.cell(row=r_idx, column=4, value=qtd)
        ws.cell(row=r_idx, column=5, value=minimo)
        ws.cell(row=r_idx, column=6, value=custo)
        ws.cell(row=r_idx, column=7, value=venda)
        ws.cell(row=r_idx, column=8, value=0)  # entradas adicionais
        ws.cell(row=r_idx, column=9, value=0)  # saídas
        # Qtd efetiva = base + entradas - saídas — usamos D como editável "atual"
        ws.cell(row=r_idx, column=10, value=f"=D{r_idx}*F{r_idx}")
        ws.cell(row=r_idx, column=11, value=f"=G{r_idx}-F{r_idx}")
        ws.cell(
            row=r_idx,
            column=12,
            value=f'=IF(D{r_idx}="","",IF(D{r_idx}<=E{r_idx},"🔴 REPOR","🟢 OK"))',
        )
        for c in range(1, 13):
            ws.cell(row=r_idx, column=c).border = border_thin
        for c in (6, 7, 10, 11):
            _money(ws.cell(row=r_idx, column=c))

    end_row = 40
    for r in range(6 + len(PRODUTOS), end_row + 1):
        for c in range(1, 13):
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(row=r, column=10, value=f'=IF(D{r}="","",D{r}*F{r})')
        ws.cell(row=r, column=11, value=f'=IF(G{r}="","",G{r}-F{r})')
        ws.cell(row=r, column=12, value=f'=IF(D{r}="","",IF(D{r}<=E{r},"🔴 REPOR","🟢 OK"))')

    _table(ws, "tblEstoque", f"A5:L{end_row}")

    # Movimentações
    ws.cell(row=5, column=14, value="MOVIMENTAÇÕES (Entrada / Saída / Venda)").font = font_section
    ws.cell(row=5, column=14).fill = fill_brand
    ws.merge_cells("N5:S5")
    mov_headers = ["Data", "Produto", "Tipo", "Quantidade", "Valor Total", "Obs"]
    for i, h in enumerate(mov_headers, start=14):
        ws.cell(row=6, column=i, value=h)
        ws.cell(row=6, column=i).fill = fill_gold
        ws.cell(row=6, column=i).font = font_header

    # Exemplo de venda
    ws.cell(row=7, column=14, value=date.today() - __import__("datetime").timedelta(days=12))
    _date_fmt(ws.cell(row=7, column=14))
    ws.cell(row=7, column=15, value="Whey Protein 900g")
    ws.cell(row=7, column=16, value="Saída/Venda")
    ws.cell(row=7, column=17, value=1)
    ws.cell(row=7, column=18, value=149.90)
    _money(ws.cell(row=7, column=18))
    ws.cell(row=7, column=19, value="Venda — Pedro Henrique")

    for r in range(7, 40):
        for c in range(14, 20):
            ws.cell(row=r, column=c).border = border_thin

    _dv_list(ws, "'09_ESTOQUE'!$B$6:$B$40", "O7:O40")
    _dv_list(ws, '"Entrada,Saída/Venda,Ajuste"', "P7:P40")

    ws.cell(row=4, column=1, value="Fluxo de venda: registre Saída aqui → ajuste Qtd Atual → lance Receita/Produtos no Financeiro.")
    ws["A4"].font = Font(name="Calibri", size=10, italic=True)

    ws.conditional_formatting.add(
        f"L6:L{end_row}",
        FormulaRule(formula=['ISNUMBER(SEARCH("REPOR",L6))'], fill=PatternFill("solid", fgColor=RED_BG)),
    )

    set_column_widths(
        ws,
        {
            1: 8,
            2: 20,
            3: 14,
            4: 10,
            5: 12,
            6: 12,
            7: 12,
            8: 10,
            9: 10,
            10: 14,
            11: 12,
            12: 12,
            14: 12,
            15: 20,
            16: 14,
            17: 12,
            18: 12,
            19: 24,
        },
    )
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 10 EQUIPAMENTOS
# ---------------------------------------------------------------------------
def build_equipamentos(wb) -> None:
    ws = wb.create_sheet("10_EQUIPAMENTOS")
    style_sheet_tab(ws)
    paint_title_bar(ws, "📦 EQUIPAMENTOS", "Aparelhos · Garantia · Manutenção", 9)

    headers = [
        "Equipamento",
        "Quantidade",
        "Data Compra",
        "Valor",
        "Fim Garantia",
        "Próx. Manutenção",
        "Status Manutenção",
        "Dias p/ Manutenção",
        "Observações",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, 9)

    for r_idx, eq in enumerate(EQUIPAMENTOS, start=6):
        nome, qtd, compra, valor, garantia, manut = eq
        ws.cell(row=r_idx, column=1, value=nome)
        ws.cell(row=r_idx, column=2, value=qtd)
        ws.cell(row=r_idx, column=3, value=compra)
        ws.cell(row=r_idx, column=4, value=valor)
        ws.cell(row=r_idx, column=5, value=garantia)
        ws.cell(row=r_idx, column=6, value=manut)
        ws.cell(
            row=r_idx,
            column=7,
            value=f'=IF(F{r_idx}="","",IF(F{r_idx}<TODAY(),"🔴 VENCIDA",IF(F{r_idx}<=TODAY()+30,"🟡 PRÓXIMA","🟢 EM DIA")))',
        )
        ws.cell(row=r_idx, column=8, value=f"=IF(F{r_idx}=\"\",\"\",F{r_idx}-TODAY())")
        ws.cell(row=r_idx, column=9, value="")
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).border = border_thin
        for c in (3, 5, 6):
            _date_fmt(ws.cell(row=r_idx, column=c))
        _money(ws.cell(row=r_idx, column=4))

    end_row = 40
    for r in range(6 + len(EQUIPAMENTOS), end_row + 1):
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(
            row=r,
            column=7,
            value=f'=IF(F{r}="","",IF(F{r}<TODAY(),"🔴 VENCIDA",IF(F{r}<=TODAY()+30,"🟡 PRÓXIMA","🟢 EM DIA")))',
        )
        ws.cell(row=r, column=8, value=f'=IF(F{r}="","",F{r}-TODAY())')

    _table(ws, "tblEquipamentos", f"A5:I{end_row}")

    ws.conditional_formatting.add(
        f"G6:G{end_row}",
        FormulaRule(formula=['ISNUMBER(SEARCH("VENCIDA",G6))'], fill=PatternFill("solid", fgColor=RED_BG)),
    )
    ws.conditional_formatting.add(
        f"G6:G{end_row}",
        FormulaRule(formula=['ISNUMBER(SEARCH("PRÓXIMA",G6))'], fill=PatternFill("solid", fgColor=YELLOW_BG)),
    )
    ws.conditional_formatting.add(
        f"A6:I{end_row}",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("VENCIDA",$G6))'],
            fill=PatternFill("solid", fgColor="FFEBEE"),
        ),
    )

    set_column_widths(ws, {1: 24, 2: 12, 3: 12, 4: 12, 5: 12, 6: 16, 7: 16, 8: 14, 9: 20})
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 11 AVALIAÇÃO
# ---------------------------------------------------------------------------
def build_avaliacao(wb) -> None:
    ws = wb.create_sheet("11_AVALIACAO")
    style_sheet_tab(ws)
    paint_title_bar(ws, "📋 AVALIAÇÃO FÍSICA", "IMC · Composição · Evolução", 10)

    headers = [
        "Aluno",
        "Data",
        "Peso (kg)",
        "Altura (m)",
        "IMC",
        "% Gordura",
        "Massa Magra",
        "Professor",
        "Observações",
        "Classificação IMC",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, 10)

    for r_idx, av in enumerate(AVALIACOES, start=6):
        aluno, data_av, peso, altura, gordura, massa, prof, obs = av
        ws.cell(row=r_idx, column=1, value=aluno)
        ws.cell(row=r_idx, column=2, value=data_av)
        ws.cell(row=r_idx, column=3, value=peso)
        ws.cell(row=r_idx, column=4, value=altura)
        ws.cell(row=r_idx, column=5, value=f"=IF(OR(C{r_idx}=\"\",D{r_idx}=\"\"),\"\",C{r_idx}/(D{r_idx}^2))")
        ws.cell(row=r_idx, column=6, value=gordura)
        ws.cell(row=r_idx, column=7, value=massa)
        ws.cell(row=r_idx, column=8, value=prof)
        ws.cell(row=r_idx, column=9, value=obs)
        ws.cell(
            row=r_idx,
            column=10,
            value=(
                f'=IF(E{r_idx}="","",IF(E{r_idx}<18.5,"Abaixo",IF(E{r_idx}<25,"Normal",'
                f'IF(E{r_idx}<30,"Sobrepeso","Obesidade"))))'
            ),
        )
        for c in range(1, 11):
            ws.cell(row=r_idx, column=c).border = border_thin
        _date_fmt(ws.cell(row=r_idx, column=2))
        ws.cell(row=r_idx, column=5).number_format = "0.00"

    end_row = 50
    for r in range(6 + len(AVALIACOES), end_row + 1):
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(row=r, column=5, value=f'=IF(OR(C{r}="",D{r}=""),"",C{r}/(D{r}^2))')
        ws.cell(
            row=r,
            column=10,
            value=(
                f'=IF(E{r}="","",IF(E{r}<18.5,"Abaixo",IF(E{r}<25,"Normal",'
                f'IF(E{r}<30,"Sobrepeso","Obesidade"))))'
            ),
        )
        ws.cell(row=r, column=5).number_format = "0.00"

    _table(ws, "tblAvaliacao", f"A5:J{end_row}")
    _dv_list(ws, "'BD_ALUNOS'!$C$2:$C$200", f"A6:A{end_row}")
    _dv_list(ws, "'08_PROFESSORES'!$B$6:$B$50", f"H6:H{end_row}")

    # Dados para gráfico de evolução (Mariana exemplo)
    ws.cell(row=5, column=12, value="EVOLUÇÃO — selecione aluno no filtro da tabela").font = font_section
    ws.cell(row=5, column=12).fill = fill_brand
    ws.cell(row=6, column=12, value="Data")
    ws.cell(row=6, column=13, value="Peso")
    ws.cell(row=6, column=14, value="IMC")
    for c in range(12, 15):
        ws.cell(row=6, column=c).fill = fill_gold

    # Série Mariana (2 pontos) + espaço
    ws.cell(row=7, column=12, value=date(2026, 1, 10))
    ws.cell(row=7, column=13, value=68.5)
    ws.cell(row=7, column=14, value=24.2)
    ws.cell(row=8, column=12, value=date(2026, 4, 10))
    ws.cell(row=8, column=13, value=66.2)
    ws.cell(row=8, column=14, value=22.8)
    for r in (7, 8):
        _date_fmt(ws.cell(row=r, column=12))

    chart = LineChart()
    chart.title = "Evolução de Peso / IMC (exemplo Mariana)"
    chart.style = 10
    chart.y_axis.title = "Valor"
    chart.x_axis.title = "Data"
    chart.height = 10
    chart.width = 15
    data = Reference(ws, min_col=13, min_row=6, max_col=14, max_row=8)
    cats = Reference(ws, min_col=12, min_row=7, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "L10")

    set_column_widths(ws, {1: 24, 2: 12, 3: 10, 4: 10, 5: 8, 6: 10, 7: 12, 8: 18, 9: 22, 10: 14, 12: 12, 13: 10, 14: 10})
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 12 PRESENÇA
# ---------------------------------------------------------------------------
def build_presenca(wb) -> None:
    ws = wb.create_sheet("12_PRESENCA")
    style_sheet_tab(ws)
    paint_title_bar(ws, "📅 CONTROLE DE PRESENÇA", "Entrada · Saída · Frequência · Ranking", 9)

    headers = [
        "Aluno",
        "Data",
        "Entrada",
        "Saída",
        "Professor",
        "Tempo Treino (min)",
        "Mês",
        "Ano",
        "Check-in",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=5, column=i, value=h)
    apply_header_row(ws, 5, 1, 9)

    for r_idx, p in enumerate(PRESENCAS, start=6):
        aluno, data_p, entrada, saida, prof = p
        ws.cell(row=r_idx, column=1, value=aluno)
        ws.cell(row=r_idx, column=2, value=data_p)
        ws.cell(row=r_idx, column=3, value=entrada)
        ws.cell(row=r_idx, column=4, value=saida)
        ws.cell(row=r_idx, column=5, value=prof)
        # Tempo: usa HORADIF se possível — fallback fórmula com TIMEVALUE
        ws.cell(
            row=r_idx,
            column=6,
            value=f'=IF(OR(C{r_idx}="",D{r_idx}=""),"",ROUND((TIMEVALUE(D{r_idx})-TIMEVALUE(C{r_idx}))*1440,0))',
        )
        ws.cell(row=r_idx, column=7, value=f"=IF(B{r_idx}=\"\",\"\",MONTH(B{r_idx}))")
        ws.cell(row=r_idx, column=8, value=f"=IF(B{r_idx}=\"\",\"\",YEAR(B{r_idx}))")
        ws.cell(row=r_idx, column=9, value="✅")
        for c in range(1, 10):
            ws.cell(row=r_idx, column=c).border = border_thin
        _date_fmt(ws.cell(row=r_idx, column=2))

    end_row = 100
    for r in range(6 + len(PRESENCAS), end_row + 1):
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(
            row=r,
            column=6,
            value=f'=IF(OR(C{r}="",D{r}=""),"",ROUND((TIMEVALUE(D{r})-TIMEVALUE(C{r}))*1440,0))',
        )
        ws.cell(row=r, column=7, value=f'=IF(B{r}="","",MONTH(B{r}))')
        ws.cell(row=r, column=8, value=f'=IF(B{r}="","",YEAR(B{r}))')

    _table(ws, "tblPresenca", f"A5:I{end_row}")
    _dv_list(ws, "'BD_ALUNOS'!$C$2:$C$200", f"A6:A{end_row}")
    _dv_list(ws, "'08_PROFESSORES'!$B$6:$B$50", f"E6:E{end_row}")

    # Frequência / Ranking
    ws.cell(row=5, column=11, value="FREQUÊNCIA MENSAL / RANKING").font = font_section
    ws.cell(row=5, column=11).fill = fill_brand
    ws.merge_cells("K5:N5")
    ws.cell(row=6, column=11, value="Aluno")
    ws.cell(row=6, column=12, value="Check-ins (mês)")
    ws.cell(row=6, column=13, value="Status Frequência")
    for c in range(11, 14):
        ws.cell(row=6, column=c).fill = fill_gold
        ws.cell(row=6, column=c).font = font_header

    for i, aluno in enumerate(ALUNOS, start=7):
        ws.cell(row=i, column=11, value=aluno["nome"])
        ws.cell(
            row=i,
            column=12,
            value=(
                f'=COUNTIFS($A$6:$A${end_row},K{i},$G$6:$G${end_row},MONTH(TODAY()),'
                f'$H$6:$H${end_row},YEAR(TODAY()))'
            ),
        )
        ws.cell(
            row=i,
            column=13,
            value=f'=IF(L{i}=0,"🔴 Ausente",IF(L{i}<4,"🟡 Baixa","🟢 Frequente"))',
        )
        for c in range(11, 14):
            ws.cell(row=i, column=c).border = border_thin

    ws.cell(row=16, column=11, value="Top frequentes: filtre a coluna Check-ins do maior para o menor.")
    ws["K16"].font = Font(name="Calibri", size=10, italic=True)

    set_column_widths(ws, {1: 24, 2: 12, 3: 10, 4: 10, 5: 18, 6: 16, 7: 8, 8: 8, 9: 10, 11: 24, 12: 14, 13: 16})
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 13 DASH FINANCEIRO
# ---------------------------------------------------------------------------
def build_dash_financeiro(wb) -> None:
    ws = wb.create_sheet("13_DASH_FINANCEIRO")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=40, cols=13)
    add_sidebar(ws, active="04_FINANCEIRO", rows=40, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)

    set_column_widths(ws, {1: 24, 2: 2, 3: 13, 4: 13, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13, 10: 13, 11: 12})

    ws.merge_cells("C5:F5")
    ws["C5"] = f"{ICON['financeiro']}  DASHBOARD FINANCEIRO"
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color="A3001E")
    ws["C5"].fill = fill_light

    # Cartões — BI_BASE (Sprint 5.0)
    paint_kpi_card(ws, 7, 3, "Receita Prevista", "='BI_BASE'!E12", True, 2)
    paint_kpi_card(ws, 7, 5, "Receita Recebida", "='BI_BASE'!E13", True, 2)
    paint_kpi_card(ws, 7, 7, "Em Atraso", "='BI_BASE'!E14", True, 2)
    paint_kpi_card(ws, 7, 9, "Lucro", "='BI_BASE'!E4", True, 2)

    paint_kpi_card(ws, 11, 3, "Caixa", "='BI_BASE'!E11", True, 2)
    paint_kpi_card(ws, 11, 5, "A Receber", "='BI_BASE'!E15", True, 2)
    paint_kpi_card(ws, 11, 7, "A Pagar", "='BI_BASE'!E16", True, 2)
    paint_kpi_card(ws, 11, 9, "Inadimplência", "='BI_BASE'!E10", False, 2)
    ws.cell(row=12, column=9).number_format = '0.0"%"'

    ws["J3"] = '=COUNTIF(\'BD_ALUNOS\'!U:U,"Ativo")'
    ws["J3"].font = Font(name="Calibri", size=1, color="F3F3F3")
    ws["M5"] = "='04_FINANCEIRO'!N14"
    _money(ws["M5"])
    ws["A6"] = "='04_FINANCEIRO'!N6"
    _money(ws["A6"])
    ws["A12"] = '=COUNTIF(\'BD_ALUNOS\'!U:U,"Ativo")'

    # Gráfico — livro-razão BD_LANCAMENTOS (Crédito=H, Débito=G, Data=B)
    ws.cell(row=16, column=3, value="Mês").fill = fill_gold
    ws.cell(row=16, column=4, value="Receita").fill = fill_gold
    ws.cell(row=16, column=5, value="Despesa").fill = fill_gold
    ws.cell(row=16, column=6, value="Lucro").fill = fill_gold
    for c in range(3, 7):
        ws.cell(row=16, column=c).font = font_header

    mes_nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    ano = date.today().year
    mes_atual = date.today().month
    for i in range(6):
        m = mes_atual - 5 + i
        a = ano
        if m <= 0:
            m += 12
            a -= 1
        r = 17 + i
        ws.cell(row=r, column=3, value=f"{mes_nomes[m-1]}")
        ws.cell(
            row=r,
            column=4,
            value=f'=SUMIFS(\'BD_LANCAMENTOS\'!H:H,\'BD_LANCAMENTOS\'!B:B,">="&DATE({a},{m},1),\'BD_LANCAMENTOS\'!B:B,"<"&DATE({a if m < 12 else a + 1},{m + 1 if m < 12 else 1},1))',
        )
        ws.cell(
            row=r,
            column=5,
            value=f'=SUMIFS(\'BD_LANCAMENTOS\'!G:G,\'BD_LANCAMENTOS\'!B:B,">="&DATE({a},{m},1),\'BD_LANCAMENTOS\'!B:B,"<"&DATE({a if m < 12 else a + 1},{m + 1 if m < 12 else 1},1))',
        )
        ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin
        for c in (4, 5, 6):
            _money(ws.cell(row=r, column=c))

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Receita × Despesa"
    chart1.style = 10
    data = Reference(ws, min_col=4, min_row=16, max_col=5, max_row=22)
    cats = Reference(ws, min_col=3, min_row=17, max_row=22)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 10
    chart1.width = 14
    ws.add_chart(chart1, "C25")


# ---------------------------------------------------------------------------
# 14 DASH COMERCIAL
# ---------------------------------------------------------------------------
def build_dash_comercial(wb) -> None:
    ws = wb.create_sheet("14_DASH_COMERCIAL")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=40, cols=12)
    add_sidebar(ws, active="FORM_ALUNO", rows=40, labels=False)
    add_top_bar(ws, start_col=2, end_col=11)

    set_column_widths(ws, {1: 24, 2: 2, 3: 13, 4: 13, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13, 10: 12})

    ws.merge_cells("C5:F5")
    ws["C5"] = f"{ICON['alunos']}  DASHBOARD COMERCIAL"
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color="A3001E")
    ws["C5"].fill = fill_light

    # Meta e conversão
    paint_kpi_card(ws, 7, 3, "Matrículas", "='BI_BASE'!E19", False, 2)
    paint_kpi_card(ws, 7, 5, "Renovações", "='BI_BASE'!E18", False, 2)
    paint_kpi_card(ws, 7, 7, "Cancelamentos", "='BI_BASE'!E7", False, 2)
    paint_kpi_card(ws, 7, 9, "Congelamentos", "='BI_BASE'!E17", False, 2)

    paint_kpi_card(ws, 11, 3, "Alunos ativos", "='BI_BASE'!E5", False, 2)
    paint_kpi_card(ws, 11, 5, "Novos alunos", "='BI_BASE'!E6", False, 2)
    paint_kpi_card(ws, 11, 7, "Conversão", "='BI_BASE'!E20", False, 2)
    paint_kpi_card(ws, 11, 9, "Churn", "='BI_BASE'!E8", False, 2)
    ws.cell(row=12, column=7).number_format = '0.0"%"'
    ws.cell(row=12, column=9).number_format = '0.0"%"'

    paint_kpi_card(ws, 15, 3, "Receita/plano", "='BI_BASE'!E3", True, 2)
    paint_kpi_card(ws, 15, 5, "Inadimplentes", '=COUNTIF(\'BD_ALUNOS\'!U:U,"Inadimplente")', False, 2)
    paint_kpi_card(ws, 15, 7, "Total alunos", '=COUNTIF(\'BD_ALUNOS\'!A:A,">0")', False, 2)

    # Helper J3 para dashboard principal
    ws["J3"] = '=COUNTIF(\'BD_ALUNOS\'!U:U,"Ativo")'
    ws["J3"].font = Font(name="Calibri", size=1, color="F3F3F3")

    # Planos
    ws.merge_cells("C20:F20")
    ws["C20"] = "PLANOS MAIS VENDIDOS"
    ws["C20"].font = font_section
    ws["C20"].fill = fill_brand
    for i, h in enumerate(["Plano", "Qtd", "Receita est.", "%"], start=3):
        ws.cell(row=21, column=i, value=h)
        ws.cell(row=21, column=i).fill = fill_gold
        ws.cell(row=21, column=i).font = font_header

    for i, (plano, valor) in enumerate(PLANOS, start=22):
        ws.cell(row=i, column=3, value=plano)
        ws.cell(row=i, column=4, value=f'=COUNTIF(\'BD_ALUNOS\'!P:P,C{i})')
        ws.cell(row=i, column=5, value=f"=D{i}*{valor}")
        ws.cell(row=i, column=6, value=f"=IFERROR(D{i}/SUM($D$22:$D$27),0)")
        ws.cell(row=i, column=6).number_format = "0.0%"
        _money(ws.cell(row=i, column=5))
        for c in range(3, 7):
            ws.cell(row=i, column=c).fill = fill_panel
            ws.cell(row=i, column=c).border = border_thin

    chart = PieChart()
    chart.title = "Alunos por Plano"
    labels = Reference(ws, min_col=3, min_row=22, max_row=27)
    data = Reference(ws, min_col=4, min_row=21, max_row=27)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.height = 10
    chart.width = 12
    ws.add_chart(chart, "H20")


def build_all_sheets(wb) -> None:
    """Ordem de criação: Config primeiro (listas), depois módulos, login/dashboard por último."""
    build_config(wb)
    build_bd_usuarios(wb)
    build_bd_sessao(wb)
    build_auth_log(wb)
    build_bd_alunos(wb)
    build_bd_contas_receber(wb)
    build_bd_contas_pagar(wb)
    build_bd_lancamentos(wb)
    build_bd_fluxo_caixa(wb)
    build_bd_parametros(wb)
    build_bd_planos(wb)
    build_bd_formas_pagamento(wb)
    build_bd_status(wb)
    build_bd_permissoes(wb)
    build_bd_cores(wb)
    build_bd_metas(wb)
    build_bd_notificacoes(wb)
    build_bd_eventos(wb)
    build_bd_prioridades(wb)
    build_bd_leads(wb)
    build_bd_crm_historico(wb)
    build_bd_retencao(wb)
    build_bd_campanhas(wb)
    build_bd_indicacoes(wb)
    build_bd_avaliacoes(wb)
    build_bd_medidas(wb)
    build_bd_treinos(wb)
    build_bd_exercicios(wb)
    build_bd_treino_itens(wb)
    build_bd_fotos(wb)
    build_bd_acessos(wb)
    build_bd_presencas(wb)
    build_bd_unidades(wb)
    build_bd_parametros_unidade(wb)
    build_bd_professor_unidade(wb)
    build_bd_transferencias(wb)
    build_bd_usuario_unidade(wb)
    build_bd_produtos(wb)
    build_bd_fornecedores(wb)
    build_bd_compras(wb)
    build_bd_movimentacao_estoque(wb)
    build_bd_lotes(wb)
    build_bd_vendas(wb)
    build_bd_venda_itens(wb)
    build_bd_kits(wb)
    build_bd_kit_itens(wb)
    build_bd_indicadores(wb)
    build_bd_insights(wb)
    build_bd_previsoes(wb)
    build_bd_risco_retencao(wb)
    build_bd_chat(wb)
    build_bd_metas_aluno(wb)
    build_bd_portal_tokens(wb)
    build_bd_desafios(wb)
    build_bd_push(wb)
    build_bd_recomendacoes(wb)
    build_bd_athena_chat(wb)
    build_bd_empresas(wb)
    build_bd_licencas(wb)
    build_bd_config_empresa(wb)
    build_bd_franqueadoras(wb)
    build_bd_franqueados(wb)
    build_bd_contratos_franquia(wb)
    build_bd_royalties(wb)
    build_bi_base(wb)
    build_versao(wb)
    build_alunos(wb)
    build_mensalidades(wb)
    build_financeiro(wb)
    build_fluxo(wb)
    build_contas_receber(wb)
    build_contas_pagar(wb)
    build_professores(wb)
    build_estoque(wb)
    build_equipamentos(wb)
    build_avaliacao(wb)
    build_presenca(wb)
    build_dash_financeiro(wb)
    build_dash_comercial(wb)
    build_dash_professores(wb)
    build_dash_estoque_bi(wb)
    build_dash_equipamentos_bi(wb)
    build_agenda(wb)
    build_home(wb)
    build_crm(wb)
    build_dash_crm(wb)
    build_avaliacao_ui(wb)
    build_treinos_ui(wb)
    build_acesso_ui(wb)
    build_dash_frequencia(wb)
    build_pdv_ui(wb)
    build_inventario_ui(wb)
    build_dash_pdv(wb)
    build_bi_executivo(wb)
    build_insights_ui(wb)
    build_portal_aluno(wb)
    build_portal_professor(wb)
    build_portal_ops(wb)
    build_athena_ai_ui(wb)
    build_recomendacoes_ui(wb)
    build_master_ui(wb)
    build_nova_academia_ui(wb)
    build_unidades_ui(wb)
    build_franqueadora_ui(wb)
    build_form_aluno(wb)
    build_relatorios(wb)
    build_dashboard(wb)
    build_login(wb)

    order = [
        "00_LOGIN",
        "21_HOME",
        "01_DASHBOARD",
        "FORM_ALUNO",
        "02_ALUNOS",
        "03_MENSALIDADES",
        "04_FINANCEIRO",
        "05_FLUXO_CAIXA",
        "06_CONTAS_RECEBER",
        "07_CONTAS_PAGAR",
        "08_PROFESSORES",
        "09_ESTOQUE",
        "10_EQUIPAMENTOS",
        "11_AVALIACAO",
        "12_PRESENCA",
        "13_DASH_FINANCEIRO",
        "14_DASH_COMERCIAL",
        "17_DASH_PROFESSORES",
        "18_DASH_ESTOQUE",
        "19_DASH_EQUIPAMENTOS",
        "20_AGENDA",
        "22_CRM",
        "23_DASH_CRM",
        "24_AVALIACAO",
        "25_TREINOS",
        "26_ACESSO",
        "27_DASH_FREQUENCIA",
        "28_PDV",
        "29_INVENTARIO",
        "30_DASH_PDV",
        "31_BI_EXECUTIVO",
        "32_INSIGHTS",
        "33_PORTAL_ALUNO",
        "34_PORTAL_PROF",
        "35_PORTAL_OPS",
        "36_ATHENA_AI",
        "37_RECOMENDACOES",
        "38_MASTER",
        "39_NOVA_ACADEMIA",
        "40_UNIDADES",
        "41_FRANQUEADORA",
        "16_RELATORIOS",
        "15_CONFIG",
        "BD_USUARIOS",
        "BD_SESSAO",
        "BD_ALUNOS",
        "BD_CONTAS_RECEBER",
        "BD_CONTAS_PAGAR",
        "BD_LANCAMENTOS",
        "BD_FLUXO_CAIXA",
        "BD_PARAMETROS",
        "BD_PLANOS",
        "BD_FORMAS_PAGAMENTO",
        "BD_STATUS",
        "BD_PERMISSOES",
        "BD_CORES",
        "BD_METAS",
        "BD_NOTIFICACOES",
        "BD_EVENTOS",
        "BD_PRIORIDADES",
        "BD_LEADS",
        "BD_CRM_HISTORICO",
        "BD_RETENCAO",
        "BD_CAMPANHAS",
        "BD_INDICACOES",
        "BD_AVALIACOES",
        "BD_MEDIDAS",
        "BD_TREINOS",
        "BD_EXERCICIOS",
        "BD_TREINO_ITENS",
        "BD_FOTOS",
        "BD_ACESSOS",
        "BD_PRESENCAS",
        "BD_UNIDADES",
        "BD_PARAMETROS_UNIDADE",
        "BD_PROFESSOR_UNIDADE",
        "BD_TRANSFERENCIAS",
        "BD_USUARIO_UNIDADE",
        "BD_PRODUTOS",
        "BD_FORNECEDORES",
        "BD_COMPRAS",
        "BD_MOVIMENTACAO_ESTOQUE",
        "BD_LOTES",
        "BD_VENDAS",
        "BD_VENDA_ITENS",
        "BD_KITS",
        "BD_KIT_ITENS",
        "BD_INDICADORES",
        "BD_INSIGHTS",
        "BD_PREVISOES",
        "BD_RISCO_RETENCAO",
        "BD_CHAT",
        "BD_METAS_ALUNO",
        "BD_PORTAL_TOKENS",
        "BD_DESAFIOS",
        "BD_PUSH",
        "BD_RECOMENDACOES",
        "BD_ATHENA_CHAT",
        "BD_EMPRESAS",
        "BD_LICENCAS",
        "BD_CONFIG_EMPRESA",
        "BD_FRANQUEADORAS",
        "BD_FRANQUEADOS",
        "BD_CONTRATOS_FRANQUIA",
        "BD_ROYALTIES",
        "BI_BASE",
        "VERSAO",
        "LOG",
    ]
    for idx, name in enumerate(order):
        wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))
