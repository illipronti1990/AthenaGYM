"""Épico 2 Sprint C/D — Multiunidade: params, professor×unidade, transferências, UI."""

from __future__ import annotations

from datetime import date

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    border_thin,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()

PARAM_UNIDADE_HEADERS = ["ID", "UnidadeID", "Chave", "Valor"]
PARAM_UNIDADE_SEED = [
    (1, 1, "HorarioAbertura", "06:00"),
    (2, 1, "HorarioFechamento", "23:00"),
    (3, 1, "BloquearInadimplente", "SIM"),
    (4, 1, "CorDashboard", "Vermelho"),
    (5, 2, "HorarioAbertura", "06:00"),
    (6, 2, "HorarioFechamento", "22:00"),
    (7, 2, "BloquearInadimplente", "SIM"),
    (8, 2, "CorDashboard", "Azul"),
]

PROF_UNIDADE_HEADERS = ["ID", "ProfessorID", "Professor", "UnidadeID", "Unidade", "Status"]
PROF_UNIDADE_SEED = [
    (1, "P001", "Carlos Mendes", 1, "ATHENA GYM Matriz", "Ativo"),
    (2, "P001", "Carlos Mendes", 2, "ATHENA GYM Zona Sul", "Ativo"),
    (3, "P002", "Ana Paula Souza", 1, "ATHENA GYM Matriz", "Ativo"),
    (4, "P003", "Roberto Lima", 2, "ATHENA GYM Zona Sul", "Ativo"),
]

TRANSF_HEADERS = [
    "ID", "Data", "Código", "Produto", "Qtde", "OrigemID", "Origem", "DestinoID", "Destino", "Usuário", "Status", "Obs",
]
TRANSF_SEED = [
    (1, HOJE, "PRD-001", "Whey Protein 900g", 2, 1, "ATHENA GYM Matriz", 2, "ATHENA GYM Zona Sul", "admin", "Concluída", "Seed demo"),
]

USUARIO_UNIDADE_HEADERS = ["ID", "UsuarioID", "Usuário", "UnidadeID", "Unidade", "Status"]
USUARIO_UNIDADE_SEED = [
    (1, 2, "admin", 1, "ATHENA GYM Matriz", "Ativo"),
    (2, 2, "admin", 2, "ATHENA GYM Zona Sul", "Ativo"),
    (3, 3, "recepcao", 1, "ATHENA GYM Matriz", "Ativo"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def _bd_sheet(wb, name: str, headers: list, seed: list, table: str) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).fill = fill_gold
        ws.cell(row=1, column=i).font = font_header
        ws.cell(row=1, column=i).border = border_thin
    for r_idx, row in enumerate(seed, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val).border = border_thin
    last = max(1 + len(seed), 2)
    _make_table(ws, table, f"A1:{get_column_letter(len(headers))}{last}")
    widths = {i: 14 for i in range(1, len(headers) + 1)}
    widths[1] = 6
    set_column_widths(ws, widths)
    ws.sheet_state = "hidden"


def build_bd_parametros_unidade(wb) -> None:
    _bd_sheet(wb, "BD_PARAMETROS_UNIDADE", PARAM_UNIDADE_HEADERS, PARAM_UNIDADE_SEED, "tbParametrosUnidade")


def build_bd_professor_unidade(wb) -> None:
    _bd_sheet(wb, "BD_PROFESSOR_UNIDADE", PROF_UNIDADE_HEADERS, PROF_UNIDADE_SEED, "tbProfessorUnidade")


def build_bd_transferencias(wb) -> None:
    _bd_sheet(wb, "BD_TRANSFERENCIAS", TRANSF_HEADERS, TRANSF_SEED, "tbTransferencias")


def build_bd_usuario_unidade(wb) -> None:
    _bd_sheet(wb, "BD_USUARIO_UNIDADE", USUARIO_UNIDADE_HEADERS, USUARIO_UNIDADE_SEED, "tbUsuarioUnidade")


def build_unidades_ui(wb) -> None:
    """40_UNIDADES — cadastro, troca, transferência e KPIs comparativos."""
    ws = wb.create_sheet("40_UNIDADES")
    style_sheet_tab(ws, BRAND_RED)
    paint_canvas(ws, rows=48, cols=12)
    add_sidebar(ws, active="40_UNIDADES", rows=48, labels=False)
    add_top_bar(ws, start_col=2, end_col=11)
    set_column_widths(ws, {1: 24, 2: 2, 3: 16, 4: 18, 5: 14, 6: 14, 7: 14, 8: 14})

    ws.merge_cells("C5:H5")
    ws["C5"] = "CADASTRO DE UNIDADE"
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light

    ws.merge_cells("C6:H6")
    ws["C6"] = "Épico 2 C/D — filiais · transferências · professores multiunidade · KPIs"
    ws["C6"].font = Font(name="Calibri", size=10, color="666666")

    fields = [
        (9, "Nome", "D9"),
        (10, "Código", "D10"),
        (11, "Cidade", "D11"),
        (12, "Responsável", "D12"),
        (13, "Telefone", "D13"),
        (14, "Status", "D14"),
    ]
    for row, lab, _ in fields:
        ws.cell(row=row, column=3, value=lab).fill = fill_panel
        ws.cell(row=row, column=3).border = border_thin
        ws.cell(row=row, column=3).font = Font(name="Calibri", size=10, color="FFFFFF")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        ws.cell(row=row, column=4).fill = fill_light
        ws.cell(row=row, column=4).border = border_thin

    ws["D14"] = "Ativa"
    ws["D10"] = "MX"

    ws.merge_cells("C17:F17")
    ws["C17"] = "Unidade ativa na sessão (TrocarUnidade)"
    ws["C17"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    ws["C18"] = "UnidadeID"
    ws["D18"] = 1
    ws["C19"] = "Nome sessão"
    ws["D19"] = "ATHENA GYM Matriz"

    paint_kpi_card(ws, 22, 3, "Alunos unidade", "0", False, 2)
    paint_kpi_card(ws, 22, 5, "Receita mês", "R$ 0,00", False, 2)
    paint_kpi_card(ws, 22, 7, "Estoque SKUs", "0", False, 2)
    paint_kpi_card(ws, 25, 3, "Unidades ativas", "2", False, 2)

    ws.merge_cells("C29:H29")
    ws["C29"] = "Comparativo por unidade (AtualizarDashboardUnidades)"
    ws["C29"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    headers = ["UnidadeID", "Nome", "Alunos ativos", "Receita mês", "SKUs estoque"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=30, column=2 + i, value=h)
        cell.fill = fill_gold
        cell.font = font_header
        cell.border = border_thin
    for r in range(31, 36):
        for c in range(3, 8):
            ws.cell(row=r, column=c).border = border_thin
            ws.cell(row=r, column=c).fill = fill_light

    ws.merge_cells("C38:H38")
    ws["C38"] = "Ações: CadastrarUnidade · TrocarUnidade · TransferirEstoque · VincularProfessorUnidade · AtualizarDashboardUnidades"
    ws["C38"].font = Font(name="Calibri", size=9, color="666666")
