"""Identidade visual ATHENAS GYM — sistema premium (não planilha)."""

from __future__ import annotations

from pathlib import Path

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# Paleta oficial
BRAND_RED = "A3001E"
BRAND_RED_DARK = "7A0016"
BRAND_RED_DEEP = "4A000F"
BRAND_RED_STRIPE = "B3122A"
GOLD = "D4AF37"
GOLD_SOFT = "E8C96A"
BLACK = "1B1B1B"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F3F3"
GRAY = "E6E6E6"
DARK_GRAY = "2B2B2B"
MUTED = "6B6B6B"
GREEN_BG = "C8E6C9"
YELLOW_BG = "FFF9C4"
RED_BG = "FFCDD2"

LOGO_PATH = Path(__file__).resolve().parent / "assets" / "athenas_logo_a.png"

fill_gold = PatternFill("solid", fgColor=GOLD)
fill_black = PatternFill("solid", fgColor=BLACK)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_light = PatternFill("solid", fgColor=LIGHT_GRAY)
fill_gray = PatternFill("solid", fgColor=GRAY)
fill_green = PatternFill("solid", fgColor=GREEN_BG)
fill_yellow = PatternFill("solid", fgColor=YELLOW_BG)
fill_red = PatternFill("solid", fgColor=RED_BG)
fill_brand = PatternFill("solid", fgColor=BRAND_RED)
fill_brand_dark = PatternFill("solid", fgColor=BRAND_RED_DARK)
fill_brand_deep = PatternFill("solid", fgColor=BRAND_RED_DEEP)
fill_brand_stripe = PatternFill("solid", fgColor=BRAND_RED_STRIPE)
fill_panel = PatternFill("solid", fgColor=WHITE)

font_title = Font(name="Georgia", size=26, bold=True, color=GOLD)
font_subtitle = Font(name="Calibri", size=11, color=WHITE)
font_header = Font(name="Calibri", size=11, bold=True, color=BLACK)
font_header_white = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_normal = Font(name="Calibri", size=11, color=DARK_GRAY)
font_menu = Font(name="Calibri", size=12, bold=True, color=WHITE)
font_menu_active = Font(name="Calibri", size=12, bold=True, color=BLACK)
font_kpi = Font(name="Georgia", size=28, bold=True, color=BRAND_RED)
font_kpi_gold = Font(name="Georgia", size=28, bold=True, color=GOLD)
font_kpi_label = Font(name="Calibri", size=10, bold=True, color=MUTED)
font_section = Font(name="Georgia", size=14, bold=True, color=GOLD)
font_link = Font(name="Calibri", size=11, bold=True, color=BLACK)
font_brand = Font(name="Georgia", size=20, bold=True, color=GOLD)
font_logo_a = Font(name="Georgia", size=40, bold=True, color=GOLD)
font_top_meta = Font(name="Calibri", size=10, color=WHITE)
font_divider = Font(name="Calibri", size=10, color=GOLD)

thin = Side(style="thin", color=GRAY)
border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_gold = Side(style="medium", color=GOLD)
border_gold = Border(left=thick_gold, right=thick_gold, top=thick_gold, bottom=thick_gold)
side_panel = Side(style="thin", color="E8E8E8")
border_panel = Border(left=side_panel, right=side_panel, top=side_panel, bottom=side_panel)
# “cartão”: borda suave + topo dourado
card_border = Border(
    left=Side(style="thin", color="EDEDED"),
    right=Side(style="thin", color="EDEDED"),
    top=Side(style="medium", color=GOLD),
    bottom=Side(style="thin", color="EDEDED"),
)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

# Ícones oficiais Fase 2 — um alfabeto só (não misturar estilos)
ICON = {
    "dashboard": "🏠",
    "alunos": "👤",
    "mensalidades": "💳",
    "financeiro": "💰",
    "estoque": "📦",
    "professores": "🏋",
    "relatorios": "📊",
    "config": "⚙",
    "sair": "🚪",
    "fluxo": "💰",
    "ticket": "💰",
    "novos": "👤",
    "inadimplentes": "💳",
}

# Botões padrão (mesmo tamanho / estilo em todo o sistema)
BTN = {
    "novo": "+ Novo",
    "salvar": "Salvar",
    "editar": "Editar",
    "excluir": "Excluir",
    "cancelar": "Cancelar",
    "pesquisar": "Pesquisar",
}

SIDEBAR_ITEMS = [
    (f"{ICON['dashboard']}  Home", "21_HOME"),
    (f"{ICON['dashboard']}  Master", "38_MASTER"),
    (f"{ICON['dashboard']}  Athena AI", "36_ATHENA_AI"),
    (f"{ICON['dashboard']}  BI Exec", "31_BI_EXECUTIVO"),
    (f"{ICON['dashboard']}  Dashboard", "01_DASHBOARD"),
    (f"{ICON['alunos']}  Portal", "33_PORTAL_ALUNO"),
    (f"{ICON['dashboard']}  Sync Cloud", "35_PORTAL_OPS"),
    (f"{ICON['alunos']}  CRM", "22_CRM"),
    (f"{ICON['alunos']}  Alunos", "FORM_ALUNO"),
    (f"{ICON['professores']}  Avaliacao", "24_AVALIACAO"),
    (f"{ICON['professores']}  Treinos", "25_TREINOS"),
    (f"{ICON['dashboard']}  Acesso", "26_ACESSO"),
    (f"{ICON['estoque']}  PDV", "28_PDV"),
    (f"{ICON['mensalidades']}  Mensalidades", "03_MENSALIDADES"),
    (f"{ICON['financeiro']}  Financeiro", "04_FINANCEIRO"),
    (f"{ICON['estoque']}  Estoque", "09_ESTOQUE"),
    (f"{ICON['relatorios']}  Relatórios", "16_RELATORIOS"),
    (f"{ICON['config']}  Configurações", "15_CONFIG"),
]


def apply_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill_gold
        cell.font = font_header
        cell.alignment = center
        cell.border = border_thin


def set_column_widths(ws, widths: dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def style_sheet_tab(ws, color: str = BRAND_RED) -> None:
    ws.sheet_properties.tabColor = color


def paint_red_texture(ws, rows: int = 40, cols: int = 14) -> None:
    ws.sheet_view.showGridLines = False
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_brand_stripe if col % 2 == 0 else fill_brand
            cell.border = Border()


def paint_canvas(ws, rows: int = 45, cols: int = 14, bg: PatternFill | None = None) -> None:
    """Fundo limpo sem grade (área de app)."""
    ws.sheet_view.showGridLines = False
    fill = bg or fill_light
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = Border()


def paint_title_bar(ws, title: str, subtitle: str = "", last_col: int = 10) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font = font_title
    c1.fill = fill_brand
    c1.alignment = center
    c2 = ws.cell(row=2, column=1, value=subtitle or "ATHENAS GYM — ERP 2.0")
    c2.font = font_subtitle
    c2.fill = fill_brand_dark
    c2.alignment = center
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 18
    gold_bottom = Border(bottom=thick_gold)
    for col in range(1, last_col + 1):
        ws.cell(row=1, column=col).fill = fill_brand
        cell2 = ws.cell(row=2, column=col)
        cell2.fill = fill_brand_dark
        cell2.border = gold_bottom


def add_nav_home(ws, cell_ref: str = "A3") -> None:
    cell = ws[cell_ref]
    cell.value = f"{ICON['dashboard']}  Menu"
    cell.font = Font(name="Calibri", size=11, bold=True, color=BLACK)
    cell.fill = fill_gold
    cell.alignment = center
    cell.hyperlink = Hyperlink(ref=cell_ref, location="'01_DASHBOARD'!A1", tooltip="Dashboard")
    ws.row_dimensions[3].height = 22


def add_logo(ws, anchor: str = "A1", width: int = 56, height: int = 56) -> None:
    if not LOGO_PATH.exists():
        return
    img = XLImage(str(LOGO_PATH))
    img.width = width
    img.height = height
    ws.add_image(img, anchor)


def add_sidebar(
    ws,
    active: str = "01_DASHBOARD",
    rows: int = 42,
    *,
    labels: bool = False,
) -> None:
    """Fundo do menu lateral.

    labels=False (padrão): só cor de fundo — os botões shape no .xlsm fazem a navegação
    (evita texto + shape sobrepostos).
    """
    # Largura suficiente para o texto do menu caber DENTRO da coluna A
    ws.column_dimensions["A"].width = 24
    for row in range(1, rows + 1):
        cell = ws.cell(row=row, column=1)
        cell.fill = fill_brand_deep
        cell.border = Border()
        cell.value = None

    ws.row_dimensions[1].height = 36
    ws.cell(row=1, column=1).fill = fill_brand_deep
    # Logo pequeno só se não houver shapes (labels=True); com shapes evita overlap
    if labels:
        add_logo(ws, "A1", 36, 36)

    ws.cell(row=2, column=1, value="ATHENAS GYM").font = Font(
        name="Georgia", size=11, bold=True, color=GOLD
    )
    ws.cell(row=2, column=1).fill = fill_brand_deep
    ws.cell(row=2, column=1).alignment = center

    ws.cell(row=3, column=1, value="ERP 2.0").font = Font(name="Calibri", size=8, color=GOLD_SOFT)
    ws.cell(row=3, column=1).fill = fill_brand_deep
    ws.cell(row=3, column=1).alignment = center

    ws.cell(row=4, column=1).fill = fill_brand_deep
    ws.cell(row=4, column=1).border = Border(bottom=thick_gold)
    ws.row_dimensions[4].height = 6

    for i, (label, sheet) in enumerate(SIDEBAR_ITEMS):
        row = 6 + i
        is_active = sheet == active
        cell = ws.cell(row=row, column=1)
        cell.alignment = left
        ws.row_dimensions[row].height = 34  # espaço vertical para botões do menu
        if is_active:
            cell.fill = fill_gold
        else:
            cell.fill = fill_brand_dark
        if labels:
            cell.value = label
            cell.hyperlink = Hyperlink(ref=f"A{row}", location=f"'{sheet}'!A1", tooltip=label)
            cell.font = font_menu_active if is_active else font_menu

    for row in range(6 + len(SIDEBAR_ITEMS), rows - 1):
        ws.cell(row=row, column=1).fill = fill_brand_deep

    ws.cell(row=rows - 1, column=1).fill = fill_brand_deep
    if labels:
        ws.cell(row=rows - 1, column=1, value=f"{ICON['sair']}  Sair").font = Font(
            name="Calibri", size=11, bold=True, color=GOLD
        )
        ws.cell(row=rows - 1, column=1).hyperlink = Hyperlink(
            ref=f"A{rows - 1}", location="'00_LOGIN'!A1", tooltip="Sair"
        )


def add_top_bar(ws, start_col: int = 2, end_col: int = 12, perfil_cell: str = "'BD_SESSAO'!B3") -> None:
    """Barra superior limpa — sem merges sobrepostos.

    Linha 1: ATHENAS GYM | ERP 2.0
    Linha 2: Bem-vindo, Nome · Perfil | data/hora
    """
    for col in range(start_col, end_col + 1):
        ws.cell(row=1, column=col).fill = fill_brand
        ws.cell(row=1, column=col).value = None
        ws.cell(row=2, column=col).fill = fill_brand_dark
        ws.cell(row=2, column=col).value = None
        ws.cell(row=3, column=col).fill = fill_gold
        ws.cell(row=3, column=col).value = None

    brand_start = start_col + 1
    # Deixa 1 coluna à direita para ERP 2.0 e 1 para data na linha 2
    brand_end = max(brand_start, end_col - 1)

    # Evita merge inválido em telas estreitas
    if brand_end > brand_start:
        ws.merge_cells(
            start_row=1, start_column=brand_start, end_row=1, end_column=brand_end - 1
        )
    ws.cell(row=1, column=brand_start, value="ATHENAS GYM").font = font_brand
    ws.cell(row=1, column=brand_start).fill = fill_brand
    ws.cell(row=1, column=brand_start).alignment = left

    ws.cell(row=1, column=end_col, value="ERP 2.0").font = Font(
        name="Calibri", size=11, bold=True, color=GOLD
    )
    ws.cell(row=1, column=end_col).fill = fill_brand
    ws.cell(row=1, column=end_col).alignment = center

    # Uma única frase na linha 2 (nome + perfil) — sem segunda coluna "Administrador"
    if brand_end > brand_start:
        ws.merge_cells(
            start_row=2, start_column=brand_start, end_row=2, end_column=brand_end - 1
        )
    ws.cell(
        row=2,
        column=brand_start,
        value=(
            '="Bem-vindo, "&IF(\'BD_SESSAO\'!B2="","—",\'BD_SESSAO\'!B2)'
            f'&"  ·  "&IF({perfil_cell}="","—",{perfil_cell})'
        ),
    ).font = font_top_meta
    ws.cell(row=2, column=brand_start).fill = fill_brand_dark
    ws.cell(row=2, column=brand_start).alignment = left

    # Data/hora portátil (pt-BR: YYYY vira literal; YEAR() evita isso)
    ws.cell(
        row=2,
        column=end_col,
        value=(
            '=TEXT(DAY(NOW()),"00")&"/"&TEXT(MONTH(NOW()),"00")&"/"&YEAR(NOW())'
            '&" "&TEXT(HOUR(NOW()),"00")&":"&TEXT(MINUTE(NOW()),"00")'
        ),
    ).font = Font(name="Calibri", size=9, bold=True, color=GOLD)
    ws.cell(row=2, column=end_col).fill = fill_brand_dark
    ws.cell(row=2, column=end_col).alignment = center

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 4


def paint_dash_card(
    ws,
    row: int,
    col: int,
    label: str,
    formula: str,
    *,
    money: bool = False,
    wide: int = 2,
    trend_formula: str | None = None,
) -> None:
    """Cartão KPI moderno (label + valor + tendência) — grade do dashboard."""
    end_col = col + wide - 1
    for r in range(row, row + 4):
        for c in range(col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_panel
            cell.border = card_border

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    lab = ws.cell(row=row, column=col, value=label)
    lab.font = font_kpi_label
    lab.fill = fill_panel
    lab.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=end_col)
    val = ws.cell(row=row + 1, column=col, value=formula)
    val.font = Font(name="Georgia", size=16, bold=True, color=BRAND_RED)
    val.fill = fill_panel
    val.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    if money:
        val.number_format = 'R$ #,##0.00'

    ws.merge_cells(start_row=row + 3, start_column=col, end_row=row + 3, end_column=end_col)
    trend = ws.cell(row=row + 3, column=col, value=trend_formula or "—")
    trend.font = Font(name="Calibri", size=11, bold=True, color=GOLD)
    trend.fill = fill_panel
    trend.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 18
    ws.row_dimensions[row + 2].height = 18
    ws.row_dimensions[row + 3].height = 18


def paint_stack_card(
    ws,
    row: int,
    col: int,
    label: str,
    formula: str,
    *,
    money: bool = False,
    wide: int = 4,
    height: int = 4,
    trend_formula: str | None = None,
) -> int:
    """Cartão empilhado (label + valor + tendência). Retorna próxima linha livre."""
    end_col = col + wide - 1
    end_row = row + height - 1

    for r in range(row, end_row + 1):
        for c in range(col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_panel
            cell.border = card_border

    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    label_cell = ws.cell(row=row, column=col, value=label.upper())
    label_cell.font = font_kpi_label
    label_cell.fill = fill_panel
    label_cell.alignment = Alignment(horizontal="left", vertical="bottom")

    value_end = row + 2 if trend_formula else end_row
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=value_end, end_column=end_col)
    value_cell = ws.cell(row=row + 1, column=col, value=formula)
    value_cell.font = font_kpi
    value_cell.fill = fill_panel
    value_cell.alignment = Alignment(horizontal="left", vertical="center")
    if money:
        value_cell.number_format = 'R$ #,##0.00'

    if trend_formula:
        ws.merge_cells(start_row=end_row, start_column=col, end_row=end_row, end_column=end_col)
        trend_cell = ws.cell(row=end_row, column=col, value=trend_formula)
        trend_cell.font = Font(name="Calibri", size=11, bold=True, color=GOLD)
        trend_cell.fill = fill_panel
        trend_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[row].height = 18
    for r in range(row + 1, end_row + 1):
        ws.row_dimensions[r].height = 20

    return end_row + 2


def paint_kpi_card(ws, row: int, col: int, label: str, formula: str, money: bool = False, wide: int = 2) -> None:
    """Cartão compacto em grade (dashboards)."""
    end_col = col + wide - 1
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=end_col)

    for r in range(row, row + 3):
        for c in range(col, end_col + 1):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = card_border

    ws.cell(row=row, column=col).value = label.upper()
    ws.cell(row=row, column=col).font = font_kpi_label
    ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="bottom")

    ws.cell(row=row + 1, column=col).value = formula
    ws.cell(row=row + 1, column=col).font = font_kpi
    ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal="left", vertical="center")
    if money:
        ws.cell(row=row + 1, column=col).number_format = 'R$ #,##0.00'

    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row + 1].height = 20
    ws.row_dimensions[row + 2].height = 16
