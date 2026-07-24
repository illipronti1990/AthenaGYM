"""Abas de autenticação / sessão / auditoria — Sprint 3.1+ / Épico 2."""

from __future__ import annotations

from datetime import date

from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    apply_header_row,
    fill_gold,
    font_header,
    set_column_widths,
    style_sheet_tab,
)

_ANO = date.today().year

# ID | EmpresaID | UnidadeID | Nome | Usuário | Senha | Perfil | Status | Token | Matrícula
# UnidadeID 0 = multi (todas); 1 = Matriz
USUARIOS_SEED = [
    (1, 0, 0, "Super Admin Plataforma", "super", "123456", "SuperAdmin", "Ativo", "", ""),
    (2, 1, 0, "Administrador", "admin", "123456", "Administrador", "Ativo", "", ""),
    (3, 1, 1, "Financeiro", "financeiro", "123456", "Financeiro", "Ativo", "", ""),
    (4, 1, 1, "Recepção", "recepcao", "123456", "Recepção", "Ativo", "", ""),
    (5, 1, 1, "Professor", "professor", "123456", "Professor", "Ativo", "", ""),
    (6, 1, 1, "Aluno Demo", "aluno", "123456", "Aluno", "Ativo", "demo-token-aluno-001", f"ATH-{_ANO}-000001"),
    (7, 0, 0, "Franqueadora Rede", "franqueadora", "123456", "Franqueadora", "Ativo", "", ""),
    (8, 1, 0, "Franqueado SP", "franqueado", "123456", "Franqueado", "Ativo", "", ""),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def build_bd_usuarios(wb) -> None:
    """BD_USUARIOS com Tabela Excel tbUsuarios (Ctrl+T)."""
    ws = wb.create_sheet("BD_USUARIOS")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False

    headers = ["ID", "EmpresaID", "UnidadeID", "Nome", "Usuário", "Senha", "Perfil", "Status", "Token", "Matrícula"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    for r, row in enumerate(USUARIOS_SEED, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    last = 1 + len(USUARIOS_SEED)
    _make_table(ws, "tbUsuarios", f"A1:J{last}")

    ws.cell(
        row=last + 2,
        column=1,
        value="Épico 2: UnidadeID (0=todas/multi). EmpresaID 0=plataforma SuperAdmin.",
    )
    ws.cell(row=last + 2, column=1).font = Font(name="Calibri", size=9, color=BRAND_RED)

    set_column_widths(ws, {1: 6, 2: 10, 3: 10, 4: 22, 5: 14, 6: 12, 7: 16, 8: 10, 9: 22, 10: 16})
    ws.freeze_panes = "A2"


def build_bd_sessao(wb) -> None:
    """Espelho da sessão para fórmulas da UI (variáveis VBA em modSessao)."""
    ws = wb.create_sheet("BD_SESSAO")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    labels = [
        "UsuarioLogado",
        "NomeUsuario",
        "PerfilUsuario",
        "DataLogin",
        "EmpresaID",
        "NomeEmpresa",
        "PlanoEmpresa",
        "UnidadeID",
        "NomeUnidade",
        "FranqueadoraID",
        "FranqueadoID",
    ]
    for r, lab in enumerate(labels, start=1):
        ws.cell(row=r, column=1, value=lab).font = font_header
        ws.cell(row=r, column=1).fill = fill_gold
        ws.cell(row=r, column=2).value = ""
    ws["A13"] = "Espelho da sessão (modSessao) — empresa + unidade + franquia — não editar"
    ws["A13"].font = Font(name="Calibri", size=9, color=BRAND_RED)
    set_column_widths(ws, {1: 18, 2: 28})


def build_log(wb) -> None:
    """LOG expandido — Sprint 3.5 (tbLog)."""
    ws = wb.create_sheet("LOG")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False

    headers = [
        "Data",
        "Hora",
        "Usuário",
        "Perfil",
        "Módulo",
        "Ação",
        "Registro",
        "Computador",
        "Versão",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    for c in range(1, len(headers) + 1):
        ws.cell(row=2, column=c).value = None

    _make_table(ws, "tbLog", f"A1:I2")

    set_column_widths(
        ws,
        {1: 12, 2: 10, 3: 14, 4: 14, 5: 14, 6: 28, 7: 20, 8: 16, 9: 10},
    )
    ws.freeze_panes = "A2"
