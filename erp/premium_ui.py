"""Telas premium — aparência de software ATHENAS GYM."""

from __future__ import annotations

from datetime import date

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

from styles import (
    BLACK,
    BRAND_RED,
    BTN,
    GOLD,
    ICON,
    WHITE,
    add_logo,
    add_sidebar,
    add_top_bar,
    border_gold,
    border_panel,
    card_border,
    center,
    fill_brand,
    fill_brand_dark,
    fill_gold,
    fill_light,
    fill_panel,
    fill_white,
    font_brand,
    font_header,
    font_kpi,
    font_kpi_label,
    font_normal,
    left,
    paint_canvas,
    paint_dash_card,
    paint_kpi_card,
    paint_red_texture,
    set_column_widths,
    style_sheet_tab,
    thick_gold,
)


def _money(cell) -> None:
    cell.number_format = 'R$ #,##0.00'


def _soft_panel(ws, r1: int, c1: int, r2: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill_panel
            cell.border = card_border


def build_login(wb) -> None:
    ws = wb.create_sheet("00_LOGIN", 0)
    style_sheet_tab(ws, BRAND_RED)
    paint_red_texture(ws, rows=36, cols=12)

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 11
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["L"].width = 5

    for row in range(3, 30):
        for col in range(3, 11):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="5C0010")
            ws.cell(row=row, column=col).border = Border()

    ws.merge_cells("C4:J4")
    ws["C4"] = "ATHENAS GYM"
    ws["C4"].font = Font(name="Georgia", size=28, bold=True, color=GOLD)
    ws["C4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws["C4"].fill = PatternFill("solid", fgColor="5C0010")
    ws.row_dimensions[4].height = 48

    for col in range(3, 11):
        ws.cell(row=5, column=col).fill = fill_gold
    ws.row_dimensions[5].height = 4

    # Logo imagem central
    ws.row_dimensions[7].height = 10
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 20
    ws.row_dimensions[10].height = 20
    add_logo(ws, "E7", 140, 140)

    ws.merge_cells("C14:J14")
    ws["C14"] = "Sistema de Gestão Premium"
    ws["C14"].font = Font(name="Calibri", size=12, color=WHITE)
    ws["C14"].alignment = center
    ws["C14"].fill = PatternFill("solid", fgColor="5C0010")

    ws.merge_cells("D16:I16")
    ws["D16"] = "Clique em ENTRAR para autenticar"
    ws["D16"].font = Font(name="Calibri", size=11, bold=True, color=GOLD)
    ws["D16"].fill = PatternFill("solid", fgColor="5C0010")
    ws["D16"].alignment = center

    ws.merge_cells("D18:I18")
    ws["D18"] = "Login via formulário seguro (senha mascarada)"
    ws["D18"].font = Font(name="Calibri", size=10, color=WHITE)
    ws["D18"].alignment = center
    ws["D18"].fill = PatternFill("solid", fgColor="5C0010")

    ws.merge_cells("D20:I20")
    ws["D20"] = "Demo: admin / financeiro / recepcao / professor"
    ws["D20"].font = Font(name="Calibri", size=9, color=WHITE)
    ws["D20"].alignment = center
    ws["D20"].fill = PatternFill("solid", fgColor="5C0010")

    ws.merge_cells("D21:I21")
    ws["D21"] = "Senha demo: 123456"
    ws["D21"].font = Font(name="Calibri", size=9, bold=True, color=GOLD)
    ws["D21"].alignment = center
    ws["D21"].fill = PatternFill("solid", fgColor="5C0010")

    # Espaço reservado para o único botão ENTRAR (shape COM em gerar_erp.py)
    ws.row_dimensions[24].height = 42
    for col in range(4, 10):
        cell = ws.cell(row=24, column=col)
        cell.fill = PatternFill("solid", fgColor="5C0010")
        cell.border = Border()
        cell.value = None

    ws.merge_cells("D27:I27")
    ws["D27"] = f"© {date.today().year} Athenas Gym — Fase 3 Login Real"
    ws["D27"].font = Font(name="Calibri", size=9, color=GOLD)
    ws["D27"].alignment = center
    ws["D27"].fill = PatternFill("solid", fgColor="5C0010")


def build_dashboard(wb) -> None:
    """Dashboard Executivo 360° — Sprint 5.0 (alimentado por BI_BASE / modBI)."""
    ws = wb.create_sheet("01_DASHBOARD", 0)
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=56, cols=14)
    add_sidebar(ws, active="01_DASHBOARD", rows=56, labels=False)
    add_top_bar(ws, start_col=2, end_col=13)

    set_column_widths(
        ws,
        {
            1: 24, 2: 3, 3: 14, 4: 14, 5: 14, 6: 14, 7: 2,
            8: 12, 9: 12, 10: 12, 11: 12, 12: 10, 13: 8, 14: 2,
        },
    )

    ws.merge_cells("C5:H5")
    ws["C5"] = "DASHBOARD EXECUTIVO"
    ws["C5"].font = Font(name="Georgia", size=20, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light
    ws["C5"].alignment = left
    ws.row_dimensions[5].height = 28

    ws.merge_cells("I5:L5")
    ws["I5"] = "Atualize com ⚙ Atualizar BI · Filtros em BI_BASE"
    ws["I5"].font = Font(name="Calibri", size=9, color="666666")

    for col in range(3, 13):
        ws.cell(row=6, column=col).border = Border(bottom=thick_gold)

    # KPIs linha 1 — BI_BASE
    paint_dash_card(ws, 7, 3, f"{ICON['financeiro']}  Receita Hoje", "='BI_BASE'!E2", money=True, wide=2, trend_formula='="● hoje"')
    paint_dash_card(ws, 7, 5, f"{ICON['financeiro']}  Receita Mês", "='BI_BASE'!E3", money=True, wide=2, trend_formula='="● mês"')
    paint_dash_card(ws, 7, 8, f"{ICON['financeiro']}  Lucro", "='BI_BASE'!E4", money=True, wide=2, trend_formula='="● líquido"')
    paint_dash_card(ws, 7, 10, f"{ICON['alunos']}  Alunos Ativos", "='BI_BASE'!E5", money=False, wide=2, trend_formula='="● base"')

    paint_dash_card(ws, 12, 3, f"{ICON['novos']}  Novos", "='BI_BASE'!E6", money=False, wide=2, trend_formula='="● mês"')
    paint_dash_card(ws, 12, 5, "Cancelamentos", "='BI_BASE'!E7", money=False, wide=2, trend_formula='="● total"')
    paint_dash_card(ws, 12, 8, "Churn", "='BI_BASE'!E8", money=False, wide=2, trend_formula="='BI_BASE'!H3")
    paint_dash_card(ws, 12, 10, f"{ICON['ticket']}  Ticket", "='BI_BASE'!E9", money=True, wide=2, trend_formula='="● médio"')
    ws.cell(row=13, column=8).number_format = '0.0"%"'

    paint_dash_card(ws, 17, 3, f"{ICON['inadimplentes']}  Inadimplência", "='BI_BASE'!E10", money=False, wide=2, trend_formula="='BI_BASE'!H2")
    paint_dash_card(ws, 17, 5, f"{ICON['fluxo']}  Caixa", "='BI_BASE'!E11", money=True, wide=2, trend_formula='="● atual"')
    paint_dash_card(
        ws, 17, 8, "% Meta Receita",
        "=IFERROR('BI_BASE'!E3/INDEX('BD_METAS'!B:B,MATCH(\"Receita\",'BD_METAS'!A:A,0)),0)",
        money=False, wide=2, trend_formula="='BI_BASE'!H4",
    )
    paint_dash_card(ws, 17, 10, "Estoque baixo", "='BI_BASE'!E24", money=False, wide=2, trend_formula="='BI_BASE'!H5")
    ws.cell(row=18, column=3).number_format = '0.0"%"'
    ws.cell(row=18, column=8).number_format = "0%"

    # Semáforos (escritos por modBI)
    ws["L7"] = "SEMÁFOROS"
    ws["L7"].font = Font(name="Calibri", size=9, bold=True, color=BRAND_RED)
    ws["L8"] = "Inad."
    ws["M8"] = "🟡"
    ws["L9"] = "Churn"
    ws["M9"] = "🟡"
    ws["L10"] = "Meta"
    ws["M10"] = "🟡"
    ws["L11"] = "Estoque"
    ws["M11"] = "🟡"

    # Metas progresso
    _soft_panel(ws, 22, 3, 28, 6)
    ws.merge_cells("C22:F22")
    ws["C22"] = "METAS DO MÊS"
    ws["C22"].font = Font(name="Georgia", size=11, bold=True, color=BRAND_RED)
    for c, h in ((3, "Indicador"), (4, "Meta"), (5, "Atual"), (6, "Progresso")):
        ws.cell(row=23, column=c, value=h).fill = fill_gold
        ws.cell(row=23, column=c).font = font_header
    for i in range(5):
        r = 24 + i
        ws.cell(row=r, column=3, value=f"='BD_METAS'!A{2+i}")
        ws.cell(row=r, column=4, value=f"='BD_METAS'!B{2+i}")
        ws.cell(row=r, column=5, value=f"='BD_METAS'!C{2+i}")
        ws.cell(row=r, column=6, value=f"='BD_METAS'!F{2+i}")
        ws.cell(row=r, column=6).number_format = "0%"
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_panel

    # Ranking planos
    _soft_panel(ws, 22, 8, 28, 12)
    ws.merge_cells("H22:L22")
    ws["H22"] = "TOP PLANOS"
    ws["H22"].font = Font(name="Georgia", size=11, bold=True, color=BRAND_RED)
    for c, h in ((8, "#"), (9, "Plano"), (10, "Qtd"), (11, "Barra")):
        ws.cell(row=23, column=c, value=h).fill = fill_gold
        ws.cell(row=23, column=c).font = font_header
    for i in range(5):
        r = 24 + i
        ws.cell(row=r, column=8, value=f"='BI_BASE'!A{12+i}")
        ws.cell(row=r, column=9, value=f"='BI_BASE'!B{12+i}")
        ws.cell(row=r, column=10, value=f"='BI_BASE'!C{12+i}")
        ws.cell(row=r, column=11, value=f"='BI_BASE'!D{12+i}")
        for c in range(8, 12):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_panel

    # Gráfico receita x despesa a partir da série BI
    _soft_panel(ws, 30, 3, 37, 7)
    ws.merge_cells("C30:G30")
    ws["C30"] = f"{ICON['financeiro']}  RECEITA × DESPESAS"
    ws["C30"].font = Font(name="Georgia", size=11, bold=True, color=BRAND_RED)
    for c, h in ((3, "Mês"), (4, "Receita"), (5, "Despesa"), (6, "Lucro")):
        ws.cell(row=31, column=c, value=h).fill = fill_gold
        ws.cell(row=31, column=c).font = font_header
    for i in range(6):
        r = 32 + i
        ws.cell(row=r, column=3, value=f"='BI_BASE'!A{26+i}")
        ws.cell(row=r, column=4, value=f"='BI_BASE'!B{26+i}")
        ws.cell(row=r, column=5, value=f"='BI_BASE'!C{26+i}")
        ws.cell(row=r, column=6, value=f"='BI_BASE'!D{26+i}")
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_panel
        for c in (4, 5, 6):
            _money(ws.cell(row=r, column=c))

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.style = 10
    chart.title = None
    data = Reference(ws, min_col=4, min_row=31, max_col=5, max_row=37)
    cats = Reference(ws, min_col=3, min_row=32, max_row=37)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 10
    chart.legend.position = "b"
    ws.add_chart(chart, "H30")

    # Alertas + Agenda do dia
    _soft_panel(ws, 39, 3, 45, 7)
    ws.merge_cells("C39:G39")
    ws["C39"] = "ALERTAS INTELIGENTES"
    ws["C39"].font = Font(name="Georgia", size=11, bold=True, color=BRAND_RED)
    for c, h in ((3, "Pri"), (4, "Mensagem"), (6, "Destino")):
        ws.cell(row=40, column=c, value=h).fill = fill_gold
        ws.cell(row=40, column=c).font = font_header
    for i in range(5):
        r = 41 + i
        ws.cell(row=r, column=3, value="")
        ws.cell(row=r, column=4, value="")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        ws.cell(row=r, column=6, value="")
        for c in (3, 4, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_panel

    _soft_panel(ws, 39, 8, 45, 12)
    ws.merge_cells("H39:L39")
    ws["H39"] = "📅 EVENTOS DE HOJE"
    ws["H39"].font = Font(name="Georgia", size=11, bold=True, color=BRAND_RED)
    for c, h in ((8, "Pri"), (9, "Hora"), (10, "Evento")):
        ws.cell(row=40, column=c, value=h).fill = fill_gold
        ws.cell(row=40, column=c).font = font_header
    for i in range(5):
        r = 41 + i
        for c in (8, 9, 10):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_panel
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)

    ws.merge_cells("C47:K47")
    ws["C47"] = "Hubs: Executivo · Comercial · Financeiro · Professores · Estoque · Equipamentos"
    ws["C47"].font = Font(name="Calibri", size=9, color="666666")

    # Faixa de ações (shapes em gerar_erp)
    ws.merge_cells("C49:K49")
    ws["C49"] = "Ações rápidas / navegação BI"
    ws["C49"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    ws["C49"].fill = fill_light
    for r in range(50, 53):
        ws.row_dimensions[r].height = 20


def build_form_aluno(wb) -> None:
    ws = wb.create_sheet("FORM_ALUNO")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=48, cols=10)
    add_sidebar(ws, active="FORM_ALUNO", rows=48, labels=False)
    add_top_bar(ws, start_col=2, end_col=9)
    set_column_widths(ws, {1: 24, 2: 3, 3: 22, 4: 28, 5: 14, 6: 14, 7: 16, 8: 14})

    ws.merge_cells("C5:E5")
    ws["C5"] = "Cadastro de Aluno"
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light

    # Painel até dia vencimento (24) — botões ficam FORA do painel
    _soft_panel(ws, 7, 3, 24, 5)

    labels = [
        (8, "Nome *"),
        (9, "CPF *"),
        (10, "Plano *"),
        (11, "Professor"),
        (12, "Valor (R$) *"),
        (13, "Forma de pagamento *"),
        (14, "RG"),
        (15, "Sexo"),
        (16, "Data nascimento"),
        (17, "Telefone *"),
        (18, "Email"),
        (19, "CEP"),
        (20, "Endereço"),
        (21, "Número"),
        (22, "Bairro"),
        (23, "Cidade"),
        (24, "Dia vencimento *"),
    ]
    for row, label in labels:
        ws.row_dimensions[row].height = 18
        ws.cell(row=row, column=3, value=label).font = Font(name="Calibri", size=11, bold=True, color=BLACK)
        ws.cell(row=row, column=3).fill = fill_panel
        default = ""
        if row == 23:
            default = "São Paulo"
        elif row == 13:
            default = "PIX"
        elif row == 24:
            default = 10
        cell = ws.cell(row=row, column=4, value=default)
        cell.fill = fill_white
        cell.border = border_gold
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        ws.cell(row=row, column=5).fill = fill_white
        ws.cell(row=row, column=5).border = border_gold

    # Texto (@) — evita Excel comer zero à esquerda (CEP/RG/Número)
    for addr in ("D9", "D14", "D16", "D17", "D18", "D19", "D21"):
        ws[addr].number_format = "@"

    ws["D12"] = '=IF(D10="","",IFERROR(VLOOKUP(D10,\'15_CONFIG\'!$A$8:$B$20,2,FALSE),""))'
    _money(ws["D12"])

    for formula1, cell in [
        ("'15_CONFIG'!$A$8:$A$13", "D10"),
        ("'08_PROFESSORES'!$B$6:$B$50", "D11"),
        ("'15_CONFIG'!$C$8:$C$20", "D13"),
        ("'15_CONFIG'!$O$8:$O$12", "D15"),
    ]:
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
        dv.add(cell)
        ws.add_data_validation(dv)

    # Metadados / painel de sucesso (Sprint 3.2)
    ws["F7"] = "ID edição →"
    ws["G7"] = 0
    ws["F8"] = "ID"
    ws["H8"] = "(ID auto)"
    ws["F9"] = "Matrícula"
    ws["H9"] = "(matrícula auto)"
    ws["G11"] = ""
    ws["G12"] = ""
    ws["G11"].font = Font(name="Calibri", size=11, bold=True, color="1B7A3D")
    ws["G12"].font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)
    ws.merge_cells("G11:I11")
    ws.merge_cells("G12:I12")
    # Âncora C27 para shapes (abaixo do Dia vencimento na linha 24)
    for r in (25, 26, 27, 28, 29, 30):
        ws.row_dimensions[r].height = 24


def build_relatorios(wb) -> None:
    ws = wb.create_sheet("16_RELATORIOS")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=36, cols=8)
    add_sidebar(ws, active="16_RELATORIOS", rows=36, labels=False)
    add_top_bar(ws, start_col=2, end_col=8)
    set_column_widths(ws, {1: 24, 2: 3, 3: 36, 4: 18, 5: 12})

    ws.merge_cells("C5:E5")
    ws["C5"] = "Central de Relatórios"
    ws["C5"].font = Font(name="Georgia", size=16, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light

    # Sem texto na área dos botões — âncora C7 para shapes
    ws["C7"] = ""
    for r in range(7, 22):
        ws.row_dimensions[r].height = 24
