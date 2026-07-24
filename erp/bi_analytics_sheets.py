"""Sprint 10.0 — Business Intelligence + Inteligencia Analitica."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()

INDICADORES_HEADERS = [
    "Indicador", "Valor Atual", "Meta", "Tendência", "Variação", "Última Atualização", "Área",
]
INSIGHTS_HEADERS = ["ID", "Data", "Tipo", "Mensagem", "Prioridade", "Módulo"]
PREVISOES_HEADERS = ["Tipo", "Horizonte", "Valor", "Base", "Atualizado"]
RISCO_HEADERS = [
    "Matrícula", "Aluno", "Score", "Classificação", "Frequência", "Mensalidade",
    "Dias sem acesso", "Avaliação", "Plano vence", "Recomendação",
]

INDICADORES_SEED = [
    ("Receita", 0, 50000, "▲", "0%", HOJE, "Financeiro"),
    ("Churn", 0, 5, "▼", "0%", HOJE, "Comercial"),
    ("Ticket Médio", 0, 180, "▲", "0%", HOJE, "Financeiro"),
    ("Inadimplência", 0, 5, "▲", "0%", HOJE, "Financeiro"),
    ("Alunos Ativos", 0, 120, "▲", "0", HOJE, "Comercial"),
    ("LTV", 0, 2000, "▲", "0%", HOJE, "Estratégico"),
    ("CAC", 0, 150, "▼", "0%", HOJE, "Estratégico"),
    ("Frequência", 0, 80, "▲", "0%", HOJE, "Operacional"),
    ("Estoque Saúde", 0, 90, "▲", "0%", HOJE, "Estoque"),
    ("Lucro", 0, 20000, "▲", "0%", HOJE, "Financeiro"),
]

INSIGHTS_SEED = [
    (1, HOJE, "Receita", "Acompanhe a evolução da receita no painel executivo.", "📊", "BI"),
    (2, HOJE, "Operação", "Horário de pico e frequência alimentam planejamento de equipe.", "📊", "Acesso"),
    (3, HOJE, "Estoque", "Produtos abaixo do mínimo geram alerta de reposição automática.", "📦", "Estoque"),
]

PREVISOES_SEED = [
    ("Receita", "Mês atual", 0, "Série 6 meses", HOJE),
    ("Receita", "Próximo mês", 0, "Tendência linear", HOJE),
    ("Caixa", "Hoje", 0, "Saldo fluxo", HOJE),
    ("Caixa", "30 dias", 0, "Receber - Pagar", HOJE),
    ("Caixa", "60 dias", 0, "Receber - Pagar", HOJE),
]

RISCO_SEED = [
    (f"ATH-{HOJE.year}-000002", "Pedro Henrique Alves", 72, "🟡 Médio risco", "Baixa", "Em dia",
     12, "OK", "60 dias", "Oferecer avaliação / treino"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def _bd_sheet(wb, name, headers, rows, table, widths, date_cols=()) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).fill = fill_gold
        ws.cell(row=1, column=i).font = font_header
        ws.cell(row=1, column=i).border = border_thin
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx in date_cols:
                _date_fmt(cell)
    last = max(1 + len(rows), 2)
    _make_table(ws, table, f"A1:{get_column_letter(len(headers))}{last}")
    set_column_widths(ws, widths)
    ws.sheet_state = "hidden"


def build_bd_indicadores(wb) -> None:
    _bd_sheet(
        wb, "BD_INDICADORES", INDICADORES_HEADERS, INDICADORES_SEED, "tbIndicadores",
        {1: 16, 2: 12, 3: 10, 4: 10, 5: 10, 6: 14, 7: 12},
        date_cols={6},
    )


def build_bd_insights(wb) -> None:
    _bd_sheet(
        wb, "BD_INSIGHTS", INSIGHTS_HEADERS, INSIGHTS_SEED, "tbInsights",
        {1: 6, 2: 12, 3: 12, 4: 55, 5: 10, 6: 12},
        date_cols={2},
    )


def build_bd_previsoes(wb) -> None:
    _bd_sheet(
        wb, "BD_PREVISOES", PREVISOES_HEADERS, PREVISOES_SEED, "tbPrevisoes",
        {1: 12, 2: 14, 3: 12, 4: 18, 5: 12},
        date_cols={5},
    )


def build_bd_risco_retencao(wb) -> None:
    _bd_sheet(
        wb, "BD_RISCO_RETENCAO", RISCO_HEADERS, RISCO_SEED, "tbRiscoRetencao",
        {1: 16, 2: 22, 3: 8, 4: 14, 5: 10, 6: 12, 7: 12, 8: 10, 9: 12, 10: 28},
    )


def build_bi_executivo(wb) -> None:
    """31_BI_EXECUTIVO — resumo executivo + metas + ranking + inteligencia comercial."""
    ws = wb.create_sheet("31_BI_EXECUTIVO")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=56, cols=13)
    add_sidebar(ws, active="31_BI_EXECUTIVO", rows=56, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 12, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 12, 11: 12},
    )
    _title(ws, "ATHENAS GYM — RESUMO EXECUTIVO")

    paint_kpi_card(ws, 7, 3, "Receita", "R$ 0", False, 2)
    paint_kpi_card(ws, 7, 5, "Alunos ativos", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Inadimplência", "0%", False, 2)
    paint_kpi_card(ws, 7, 9, "Lucro", "R$ 0", False, 2)

    paint_kpi_card(ws, 11, 3, "Churn", "0%", False, 2)
    paint_kpi_card(ws, 11, 5, "Ticket médio", "R$ 0", False, 2)
    paint_kpi_card(ws, 11, 7, "LTV", "R$ 0", False, 2)
    paint_kpi_card(ws, 11, 9, "CAC", "R$ 0", False, 2)

    paint_kpi_card(ws, 15, 3, "Frequência", "0%", False, 2)
    paint_kpi_card(ws, 15, 5, "Estoque saúde", "0%", False, 2)
    paint_kpi_card(ws, 15, 7, "Previsão mês+", "R$ 0", False, 2)
    paint_kpi_card(ws, 15, 9, "Caixa 30d", "R$ 0", False, 2)

    ws.merge_cells("C20:F20")
    ws["C20"] = "INTELIGÊNCIA COMERCIAL"
    ws["C20"].font = font_section
    ws["C20"].fill = fill_brand
    for i, lab in enumerate(
        ["Plano mais vendido", "Horário pico", "Professor retenção", "Produto top", "Canal aquisição"]
    ):
        r = 21 + i
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(row=r, column=4).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
        ws.cell(row=r, column=4).font = Font(name="Georgia", size=11, bold=True, color=BRAND_RED)

    ws.merge_cells("H20:K20")
    ws["H20"] = "PREVISÃO DE CAIXA"
    ws["H20"].font = font_section
    ws["H20"].fill = fill_brand
    for i, lab in enumerate(["Hoje", "30 dias", "60 dias"]):
        r = 21 + i
        ws.cell(row=r, column=8, value=lab).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
        ws.cell(row=r, column=9).fill = fill_panel
        ws.cell(row=r, column=9).border = border_thin
        ws.cell(row=r, column=9).font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)

    ws.merge_cells("C27:F27")
    ws["C27"] = "PREVISÃO FINANCEIRA (SÉRIE)"
    ws["C27"].font = font_section
    ws["C27"].fill = fill_brand
    for i, h in enumerate(["Mês", "Receita", "Tendência"], start=3):
        ws.cell(row=28, column=i, value=h).fill = fill_gold
        ws.cell(row=28, column=i).font = font_header
    for i in range(7):
        r = 29 + i
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("H27:K27")
    ws["H27"] = "DASHBOARD METAS"
    ws["H27"].font = font_section
    ws["H27"].fill = fill_brand
    for i, h in enumerate(["Indicador", "Progresso", "Barra", "Semáforo"], start=8):
        ws.cell(row=28, column=i, value=h).fill = fill_gold
        ws.cell(row=28, column=i).font = font_header
    for i in range(8):
        r = 29 + i
        for c in range(8, 12):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C38:K38")
    ws["C38"] = "RANKING GERAL"
    ws["C38"].font = font_section
    ws["C38"].fill = fill_brand
    for i, h in enumerate(["Categoria", "Destaque", "Score", "Detalhe"], start=3):
        ws.cell(row=39, column=i, value=h).fill = fill_gold
        ws.cell(row=39, column=i).font = font_header
    for i in range(6):
        r = 40 + i
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C48:K48")
    ws["C48"] = "KPIs estratégicos: Receita · Lucro · Churn · LTV · CAC · Ticket · por plano/professor/unidade"
    ws["C48"].font = Font(name="Calibri", size=9, color="666666")


def build_insights_ui(wb) -> None:
    """32_INSIGHTS — central de insights, retenção, estoque inteligente, simulador."""
    ws = wb.create_sheet("32_INSIGHTS")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=58, cols=13)
    add_sidebar(ws, active="31_BI_EXECUTIVO", rows=58, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 14, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 14, 11: 12},
    )
    _title(ws, "CENTRAL DE INSIGHTS + SIMULADOR")

    ws.merge_cells("C7:K7")
    ws["C7"] = "INSIGHTS AUTOMÁTICOS"
    ws["C7"].font = font_section
    ws["C7"].fill = fill_brand
    for i in range(8):
        r = 8 + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
        ws.cell(row=r, column=3).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=11)

    ws.merge_cells("C17:F17")
    ws["C17"] = "INTELIGÊNCIA DE RETENÇÃO (ALTO RISCO)"
    ws["C17"].font = font_section
    ws["C17"].fill = fill_brand
    for i, h in enumerate(["Aluno", "Score", "Classe", "Recomendação"], start=3):
        ws.cell(row=18, column=i, value=h).fill = fill_gold
        ws.cell(row=18, column=i).font = font_header
    for i in range(8):
        r = 19 + i
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("H17:K17")
    ws["H17"] = "COMPRAS INTELIGENTES"
    ws["H17"].font = font_section
    ws["H17"].fill = fill_brand
    for i, h in enumerate(["Produto", "Dias resto", "Comprar", "Qtde sug."], start=8):
        ws.cell(row=18, column=i, value=h).fill = fill_gold
        ws.cell(row=18, column=i).font = font_header
    for i in range(8):
        r = 19 + i
        for c in range(8, 12):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C28:K28")
    ws["C28"] = "SIMULADOR FINANCEIRO"
    ws["C28"].font = font_section
    ws["C28"].fill = fill_brand

    ws["C29"] = "Cenário"
    ws["C29"].fill = fill_panel
    ws["C29"].border = border_thin
    ws.merge_cells("D29:F29")
    ws["D29"] = "Aumentar plano Premium %"
    ws["D29"].fill = fill_panel
    ws["D29"].border = border_thin

    ws["C30"] = "Parâmetro %"
    ws["C30"].fill = fill_panel
    ws["C30"].border = border_thin
    ws.merge_cells("D30:F30")
    ws["D30"] = 10
    ws["D30"].fill = fill_panel
    ws["D30"].border = border_thin

    ws["C31"] = "Resultado"
    ws["C31"].fill = fill_panel
    ws["C31"].border = border_thin
    ws.merge_cells("D31:F31")
    ws["D31"] = "—"
    ws["D31"].fill = fill_panel
    ws["D31"].border = border_thin
    ws["D31"].font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)

    ws["H29"] = "Cenário 2"
    ws["H29"].fill = fill_panel
    ws["H29"].border = border_thin
    ws.merge_cells("I29:K29")
    ws["I29"] = "Contratar professor (custo R$)"
    ws["I29"].fill = fill_panel
    ws["I29"].border = border_thin

    ws["H30"] = "Custo mensal"
    ws["H30"].fill = fill_panel
    ws["H30"].border = border_thin
    ws.merge_cells("I30:K30")
    ws["I30"] = 3500
    ws["I30"].fill = fill_panel
    ws["I30"].border = border_thin

    ws["H31"] = "Capacidade +"
    ws["H31"].fill = fill_panel
    ws["H31"].border = border_thin
    ws.merge_cells("I31:K31")
    ws["I31"] = 90
    ws["I31"].fill = fill_panel
    ws["I31"].border = border_thin

    ws["H32"] = "Resultado"
    ws["H32"].fill = fill_panel
    ws["H32"].border = border_thin
    ws.merge_cells("I32:K32")
    ws["I32"] = "—"
    ws["I32"].fill = fill_panel
    ws["I32"].border = border_thin
    ws["I32"].font = Font(name="Georgia", size=11, bold=True, color=BRAND_RED)

    ws.row_dimensions[34].height = 36

    ws.merge_cells("C36:K36")
    ws["C36"] = "INDICADORES (BD_INDICADORES)"
    ws["C36"].font = font_section
    ws["C36"].fill = fill_brand
    for i, h in enumerate(["Indicador", "Atual", "Meta", "Tendência", "Variação"], start=3):
        ws.cell(row=37, column=i, value=h).fill = fill_gold
        ws.cell(row=37, column=i).font = font_header
    for i in range(10):
        r = 38 + i
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C50:K50")
    ws["C50"] = "Critérios de risco: frequência ↓ · atraso · sem avaliação · sem compra · plano vencendo · sem acesso"
    ws["C50"].font = Font(name="Calibri", size=9, color="666666")
