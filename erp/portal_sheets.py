"""Sprint 11.0 — Portal do Aluno / Professor / Recepcao / Gestor (Excel + base cloud)."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()
ANO = HOJE.year
MAT1 = f"ATH-{ANO}-000001"

CHAT_HEADERS = ["ID", "De", "Para", "Matrícula", "Mensagem", "Data", "Hora", "Lida"]
METAS_ALUNO_HEADERS = ["ID", "Matrícula", "Aluno", "Objetivo", "Meta", "Atual", "Progresso", "Unidade"]
PORTAL_TOKENS_HEADERS = ["Usuário", "Token", "Perfil", "Matrícula", "Expira", "Dispositivo"]
DESAFIOS_HEADERS = ["ID", "Matrícula", "Desafio", "Meta", "Atual", "Progresso", "Conquista"]
PUSH_HEADERS = ["ID", "Usuário", "Matrícula", "Mensagem", "Data", "Hora", "Lida", "Tipo"]

CHAT_SEED = [
    (1, "professor", "aluno", MAT1, "Aumente a carga do supino nesta semana.", HOJE - timedelta(days=1), "10:15", "SIM"),
    (2, "aluno", "professor", MAT1, "Ok, combinado!", HOJE - timedelta(days=1), "10:22", "SIM"),
    (3, "professor", "aluno", MAT1, "Nova ficha ABCD disponível no app.", HOJE, "08:00", "NÃO"),
]

METAS_ALUNO_SEED = [
    (1, MAT1, "Mariana Oliveira", "Hipertrofia", 30, 23, 0.77, "dias treino/mês"),
    (2, f"ATH-{ANO}-000002", "Pedro Henrique Alves", "Emagrecimento", 90, 88, 0.98, "kg"),
]

TOKENS_SEED = [
    ("aluno", "demo-token-aluno-001", "Aluno", MAT1, HOJE + timedelta(days=30), "App"),
    ("professor", "demo-token-prof-001", "Professor", "", HOJE + timedelta(days=30), "Web"),
]

DESAFIOS_SEED = [
    (1, MAT1, "Treinar 20 dias", 20, 9, 0.45, ""),
    (2, MAT1, "30 dias seguidos", 30, 0, 0, ""),
    (3, MAT1, "100 treinos", 100, 26, 0.26, "🏅 Primeiro mês"),
]

PUSH_SEED = [
    (1, "aluno", MAT1, "Sua mensalidade vence amanhã.", HOJE, "08:00", "NÃO", "Financeiro"),
    (2, "aluno", MAT1, "Você tem treino hoje — Treino A.", HOJE, "07:00", "NÃO", "Treino"),
    (3, "aluno", MAT1, "Nova avaliação disponível.", HOJE - timedelta(days=2), "09:00", "SIM", "Avaliação"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def _bd_sheet(wb, name, headers, rows, table, widths, date_cols=()) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).fill = fill_gold
        ws.cell(row=1, column=i).font = font_header
        ws.cell(row=1, column=i).border = border_thin
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx in date_cols:
                _date_fmt(cell)
    last = max(1 + len(rows), 2)
    _make_table(ws, table, f"A1:{get_column_letter(len(headers))}{last}")
    set_column_widths(ws, widths)
    ws.sheet_state = "hidden"


def build_bd_chat(wb) -> None:
    _bd_sheet(
        wb, "BD_CHAT", CHAT_HEADERS, CHAT_SEED, "tbChat",
        {1: 6, 2: 12, 3: 12, 4: 16, 5: 40, 6: 12, 7: 8, 8: 8},
        date_cols={6},
    )


def build_bd_metas_aluno(wb) -> None:
    _bd_sheet(
        wb, "BD_METAS_ALUNO", METAS_ALUNO_HEADERS, METAS_ALUNO_SEED, "tbMetasAluno",
        {1: 6, 2: 16, 3: 20, 4: 14, 5: 10, 6: 10, 7: 10, 8: 16},
    )


def build_bd_portal_tokens(wb) -> None:
    _bd_sheet(
        wb, "BD_PORTAL_TOKENS", PORTAL_TOKENS_HEADERS, TOKENS_SEED, "tbPortalTokens",
        {1: 12, 2: 24, 3: 12, 4: 16, 5: 12, 6: 12},
        date_cols={5},
    )


def build_bd_desafios(wb) -> None:
    _bd_sheet(
        wb, "BD_DESAFIOS", DESAFIOS_HEADERS, DESAFIOS_SEED, "tbDesafios",
        {1: 6, 2: 16, 3: 18, 4: 8, 5: 8, 6: 10, 7: 16},
    )


def build_bd_push(wb) -> None:
    _bd_sheet(
        wb, "BD_PUSH", PUSH_HEADERS, PUSH_SEED, "tbPush",
        {1: 6, 2: 12, 3: 16, 4: 40, 5: 12, 6: 8, 7: 8, 8: 12},
        date_cols={5},
    )


def build_portal_aluno(wb) -> None:
    """33_PORTAL_ALUNO — experiencia do aluno no Excel (espelho do app)."""
    ws = wb.create_sheet("33_PORTAL_ALUNO")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=58, cols=13)
    add_sidebar(ws, active="33_PORTAL_ALUNO", rows=58, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 14, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 12, 11: 12},
    )
    _title(ws, "PORTAL DO ALUNO")

    ws.merge_cells("C7:F7")
    ws["C7"] = "LOGIN ALUNO (demo)"
    ws["C7"].font = font_section
    ws["C7"].fill = fill_brand
    for r, lab, val in ((8, "Matrícula", MAT1), (9, "Nome", ""), (10, "Usuário portal", "aluno")):
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(row=r, column=4, value=val).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin

    paint_kpi_card(ws, 12, 3, "Plano", "—", False, 2)
    paint_kpi_card(ws, 12, 5, "Professor", "—", False, 2)
    paint_kpi_card(ws, 12, 7, "Mensalidade", "—", False, 2)
    paint_kpi_card(ws, 12, 9, "Último treino", "—", False, 2)

    ws.merge_cells("C17:F17")
    ws["C17"] = "MEU TREINO"
    ws["C17"].font = font_section
    ws["C17"].fill = fill_brand
    for i, h in enumerate(["Dia", "Exercício", "Séries", "Reps"], start=3):
        ws.cell(row=18, column=i, value=h).fill = fill_gold
        ws.cell(row=18, column=i).font = font_header
    for i in range(8):
        r = 19 + i
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("H17:K17")
    ws["H17"] = "MINHA EVOLUÇÃO"
    ws["H17"].font = font_section
    ws["H17"].fill = fill_brand
    for i, lab in enumerate(["Peso", "IMC", "Gordura %", "Próx. avaliação"]):
        r = 18 + i
        ws.cell(row=r, column=8, value=lab).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
        ws.cell(row=r, column=9).fill = fill_panel
        ws.cell(row=r, column=9).border = border_thin
        ws.cell(row=r, column=9).font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)

    ws.merge_cells("C28:F28")
    ws["C28"] = "FREQUÊNCIA / ACESSOS"
    ws["C28"].font = font_section
    ws["C28"].fill = fill_brand
    for i, h in enumerate(["Data", "Entrada", "Saída"], start=3):
        ws.cell(row=29, column=i, value=h).fill = fill_gold
        ws.cell(row=29, column=i).font = font_header
    for i in range(6):
        r = 30 + i
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("H28:K28")
    ws["H28"] = "FINANCEIRO"
    ws["H28"].font = font_section
    ws["H28"].fill = fill_brand
    for i, lab in enumerate(["Competência", "Valor", "Vencimento", "Status", "PIX / Pagar"]):
        r = 29 + i
        ws.cell(row=r, column=8, value=lab).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=11)
        ws.cell(row=r, column=9).fill = fill_panel
        ws.cell(row=r, column=9).border = border_thin

    ws.merge_cells("C37:F37")
    ws["C37"] = "OBJETIVOS / DESAFIOS"
    ws["C37"].font = font_section
    ws["C37"].fill = fill_brand
    for i, h in enumerate(["Item", "Meta", "Atual", "Barra"], start=3):
        ws.cell(row=38, column=i, value=h).fill = fill_gold
        ws.cell(row=38, column=i).font = font_header
    for i in range(5):
        r = 39 + i
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("H37:K37")
    ws["H37"] = "NOTIFICAÇÕES"
    ws["H37"].font = font_section
    ws["H37"].fill = fill_brand
    for i in range(6):
        r = 38 + i
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
        ws.cell(row=r, column=8).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin

    ws.merge_cells("C46:K46")
    ws["C46"] = "CHAT COM PROFESSOR"
    ws["C46"].font = font_section
    ws["C46"].fill = fill_brand
    for i in range(5):
        r = 47 + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
        ws.cell(row=r, column=3).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin

    ws.merge_cells("C53:K53")
    ws["C53"] = "Espelho do App/Portal · API cloud em /cloud/api · Flutter em /cloud/mobile"
    ws["C53"].font = Font(name="Calibri", size=9, color="666666")


def build_portal_professor(wb) -> None:
    """34_PORTAL_PROF — painel do professor."""
    ws = wb.create_sheet("34_PORTAL_PROF")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=48, cols=13)
    add_sidebar(ws, active="34_PORTAL_PROF", rows=48, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 14, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 12},
    )
    _title(ws, "PORTAL DO PROFESSOR")

    paint_kpi_card(ws, 7, 3, "Avaliações hoje", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Treinos ativos", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Alunos", 0, False, 2)
    paint_kpi_card(ws, 7, 9, "Reavaliações", 0, False, 2)

    ws.merge_cells("C12:F12")
    ws["C12"] = "ALUNO SELECIONADO"
    ws["C12"].font = font_section
    ws["C12"].fill = fill_brand
    for r, lab in ((13, "Matrícula"), (14, "Nome"), (15, "Peso"), (16, "Treino"), (17, "Última avaliação")):
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(row=r, column=4).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
    ws["D13"] = MAT1

    ws.merge_cells("H12:K12")
    ws["H12"] = "AGENDA DO DIA"
    ws["H12"].font = font_section
    ws["H12"].fill = fill_brand
    for i, h in enumerate(["Hora", "Aluno", "Tipo"], start=8):
        ws.cell(row=13, column=i, value=h).fill = fill_gold
        ws.cell(row=13, column=i).font = font_header
    for i in range(8):
        r = 14 + i
        for c in range(8, 11):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C24:K24")
    ws["C24"] = "CHAT — MENSAGENS COM O ALUNO"
    ws["C24"].font = font_section
    ws["C24"].fill = fill_brand
    for i in range(8):
        r = 25 + i
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
        ws.cell(row=r, column=3).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin

    ws["C34"] = "Nova mensagem"
    ws["C34"].fill = fill_panel
    ws["C34"].border = border_thin
    ws.merge_cells("D34:K34")
    ws["D34"].fill = fill_panel
    ws["D34"].border = border_thin

    ws.row_dimensions[36].height = 36


def build_portal_ops(wb) -> None:
    """35_PORTAL_OPS — recepção + gestor (KPIs operacionais/web)."""
    ws = wb.create_sheet("35_PORTAL_OPS")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=44, cols=13)
    add_sidebar(ws, active="35_PORTAL_OPS", rows=44, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 12, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 12},
    )
    _title(ws, "PORTAL RECEPÇÃO / GESTOR")

    ws.merge_cells("C7:F7")
    ws["C7"] = "RECEPÇÃO — HOJE"
    ws["C7"].font = font_section
    ws["C7"].fill = fill_brand
    paint_kpi_card(ws, 8, 3, "Matrículas", 0, False, 2)
    paint_kpi_card(ws, 8, 5, "Pagamentos", 0, False, 2)
    paint_kpi_card(ws, 8, 7, "Visitas", 0, False, 2)
    paint_kpi_card(ws, 8, 9, "Renovações", 0, False, 2)

    ws.merge_cells("C13:F13")
    ws["C13"] = "GESTOR — RESUMO"
    ws["C13"].font = font_section
    ws["C13"].fill = fill_brand
    paint_kpi_card(ws, 14, 3, "Receita", "R$ 0", False, 2)
    paint_kpi_card(ws, 14, 5, "Lucro", "R$ 0", False, 2)
    paint_kpi_card(ws, 14, 7, "Churn", "0%", False, 2)
    paint_kpi_card(ws, 14, 9, "A receber", "R$ 0", False, 2)

    ws.merge_cells("C20:K20")
    ws["C20"] = "SYNC CLOUD / API"
    ws["C20"].font = font_section
    ws["C20"].fill = fill_brand
    for r, lab in (
        (21, "URL API"),
        (22, "Última sync"),
        (23, "Status"),
        (24, "Arquivo export"),
    ):
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        ws.cell(row=r, column=4).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
    ws["D21"] = "http://127.0.0.1:8002"
    ws["D23"] = "Pronto para sincronizar"

    ws.merge_cells("C27:K27")
    ws["C27"] = "Ações: matrícula · pagamento · acesso · experimental — via ERP Excel; web consome a mesma API."
    ws["C27"].font = Font(name="Calibri", size=9, color="666666")
    ws.row_dimensions[29].height = 36
