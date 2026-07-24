"""Sprint 5.0 — Metas, alertas e base BI (camada tipo Power Query)."""

from __future__ import annotations

from datetime import date

from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    apply_header_row,
    border_thin,
    font_normal,
    set_column_widths,
    style_sheet_tab,
)

METAS_SEED = [
    ("Receita", 50000, "R$", "Financeiro"),
    ("Novos alunos", 40, "N", "Comercial"),
    ("Churn", 5, "%", "Comercial"),
    ("Inadimplência", 5, "%", "Financeiro"),
    ("Ticket Médio", 180, "R$", "Financeiro"),
    ("Alunos Ativos", 120, "N", "Comercial"),
    ("Avaliações", 30, "N", "Operacional"),
    ("LTV", 2000, "R$", "Estratégico"),
    ("CAC", 150, "R$", "Estratégico"),
    ("Lucro", 20000, "R$", "Financeiro"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def _money(cell) -> None:
    cell.number_format = 'R$ #,##0.00'


def build_bd_metas(wb) -> None:
    ws = wb.create_sheet("BD_METAS")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    headers = ["Indicador", "Meta", "Atual", "Unidade", "Área", "Progresso", "Semáforo"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    for r_idx, (ind, meta, und, area) in enumerate(METAS_SEED, start=2):
        ws.cell(row=r_idx, column=1, value=ind).border = border_thin
        ws.cell(row=r_idx, column=2, value=meta).border = border_thin
        ws.cell(row=r_idx, column=3, value=0).border = border_thin
        ws.cell(row=r_idx, column=4, value=und).border = border_thin
        ws.cell(row=r_idx, column=5, value=area).border = border_thin
        ws.cell(row=r_idx, column=6, value=0).border = border_thin
        ws.cell(row=r_idx, column=6).number_format = "0%"
        ws.cell(row=r_idx, column=7, value="🟡").border = border_thin
        if und == "R$":
            _money(ws.cell(row=r_idx, column=2))
            _money(ws.cell(row=r_idx, column=3))

    last = 1 + len(METAS_SEED)
    _make_table(ws, "tbMetas", f"A1:G{last}")
    set_column_widths(ws, {1: 16, 2: 12, 3: 12, 4: 10, 5: 12, 6: 12, 7: 12})
    ws.freeze_panes = "A2"


def build_bd_notificacoes(wb) -> None:
    """Centro de notificações — alimentado por modBI."""
    ws = wb.create_sheet("BD_NOTIFICACOES")
    style_sheet_tab(ws, BRAND_RED)
    ws.sheet_view.showGridLines = False
    headers = ["Data", "Hora", "Tipo", "Prioridade", "Mensagem", "Destino", "Lida"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    seed = [
        (date.today(), "08:00", "Financeiro", "🔴", "Mensalidades vencidas — verifique Contas a Receber", "04_FINANCEIRO", "NÃO"),
        (date.today(), "08:00", "Estoque", "🟡", "Produtos abaixo do mínimo — verifique Estoque", "09_ESTOQUE", "NÃO"),
        (date.today(), "08:00", "Meta", "🟢", "Acompanhe o progresso das metas no Dashboard", "01_DASHBOARD", "NÃO"),
    ]
    for r_idx, row in enumerate(seed, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if c_idx == 1:
                cell.number_format = "DD/MM/YYYY"

    last = 1 + len(seed)
    for c in range(1, 8):
        ws.cell(row=last + 1, column=c).value = None
    last = last + 1
    _make_table(ws, "tbNotificacoes", f"A1:G{last}")
    set_column_widths(ws, {1: 12, 2: 8, 3: 12, 4: 10, 5: 48, 6: 16, 7: 8})
    ws.freeze_panes = "A2"


def build_bi_base(wb) -> None:
    """
    BI_BASE — camada consolidada (substitui Power Query nesta fase).
    VBA (AtualizarBIBase) reescreve os dados a partir dos BD_*.
    Dashboards e gráficos leem desta aba.
    """
    ws = wb.create_sheet("BI_BASE")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False

    # Filtros (segmentações manuais — VBA respeita)
    ws["A1"] = "FILTROS BI"
    ws["A1"].font = Font(name="Calibri", size=11, bold=True, color=BRAND_RED)
    ws["A2"] = "Data Início"
    ws["B2"] = date(date.today().year, date.today().month, 1)
    ws["B2"].number_format = "DD/MM/YYYY"
    ws["A3"] = "Data Fim"
    ws["B3"] = date.today()
    ws["B3"].number_format = "DD/MM/YYYY"
    ws["A4"] = "Professor"
    ws["B4"] = "(Todos)"
    ws["A5"] = "Plano"
    ws["B5"] = "(Todos)"
    ws["A6"] = "Aluno"
    ws["B6"] = "(Todos)"
    ws["A7"] = "Unidade"
    ws["B7"] = "ATHENAS GYM"
    for r in range(2, 8):
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=2).border = border_thin
        ws.cell(row=r, column=2).fill = __import__("openpyxl.styles", fromlist=["PatternFill"]).PatternFill(
            "solid", fgColor="FFF8E7"
        )

    # KPIs consolidados (escritos por modBI)
    ws["D1"] = "KPIs EXECUTIVOS"
    ws["D1"].font = Font(name="Calibri", size=11, bold=True, color=BRAND_RED)
    kpis = [
        ("ReceitaHoje", 0),
        ("ReceitaMes", 0),
        ("Lucro", 0),
        ("AlunosAtivos", 0),
        ("NovosAlunos", 0),
        ("Cancelamentos", 0),
        ("Churn", 0),
        ("TicketMedio", 0),
        ("Inadimplencia", 0),
        ("CaixaAtual", 0),
        ("ReceitaPrevista", 0),
        ("ReceitaRecebida", 0),
        ("ReceitaAtraso", 0),
        ("ContasReceber", 0),
        ("ContasPagar", 0),
        ("Congelamentos", 0),
        ("Renovacoes", 0),
        ("Matriculas", 0),
        ("Conversao", 0),
        ("QtdProfessores", 0),
        ("AvaliacoesMes", 0),
        ("Produtos", 0),
        ("EstoqueBaixo", 0),
        ("ValorEstoque", 0),
        ("Equipamentos", 0),
        ("EmManutencao", 0),
        ("EventosHoje", 0),
        ("EventosAlta", 0),
        ("EventosMensalidade", 0),
        ("EventosAvaliacao", 0),
        ("CRM_Leads", 0),
        ("CRM_Conversao", 0),
        ("CRM_Risco", 0),
        ("Acesso_Hoje", 0),
        ("Acesso_Ausentes", 0),
        ("LTV", 0),
        ("CAC", 0),
        ("FrequenciaPct", 0),
        ("EstoqueSaudePct", 0),
        ("PrevisaoReceita", 0),
        ("PrevisaoCaixa30", 0),
    ]
    for i, (nome, val) in enumerate(kpis, start=2):
        ws.cell(row=i, column=4, value=nome).border = border_thin
        cell = ws.cell(row=i, column=5, value=val)
        cell.border = border_thin
        if "Receita" in nome or "Lucro" in nome or "Caixa" in nome or "Ticket" in nome or "Valor" in nome or "Contas" in nome:
            _money(cell)
        if nome in ("Churn", "Inadimplencia", "Conversao"):
            cell.number_format = "0.0"

    # Semáforos
    ws["G1"] = "SEMÁFOROS"
    ws["G1"].font = Font(name="Calibri", size=11, bold=True, color=BRAND_RED)
    for i, nome in enumerate(("Inadimplencia", "Churn", "MetaReceita", "Estoque"), start=2):
        ws.cell(row=i, column=7, value=nome).border = border_thin
        ws.cell(row=i, column=8, value="🟡").border = border_thin

    # Rankings (preenchidos por VBA)
    ws["A10"] = "RANKING PLANOS"
    ws["A10"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    for i, h in enumerate(["#", "Plano", "Qtd", "Barra"], start=1):
        ws.cell(row=11, column=i, value=h)
    apply_header_row(ws, 11, 1, 4)
    for r in range(12, 22):
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = border_thin

    ws["F10"] = "RANKING PROFESSORES"
    ws["F10"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    for i, h in enumerate(["#", "Professor", "Alunos"], start=6):
        ws.cell(row=11, column=i, value=h)
    apply_header_row(ws, 11, 6, 3)
    for r in range(12, 22):
        for c in range(6, 9):
            ws.cell(row=r, column=c).border = border_thin

    ws["J10"] = "ALERTAS ATIVOS"
    ws["J10"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    for i, h in enumerate(["Pri", "Mensagem", "Destino"], start=10):
        ws.cell(row=11, column=i, value=h)
    apply_header_row(ws, 11, 10, 3)
    for r in range(12, 22):
        for c in range(10, 13):
            ws.cell(row=r, column=c).border = border_thin

    # Série mensal para gráficos (6 meses)
    ws["A24"] = "SÉRIE MENSAL"
    ws["A24"].font = Font(name="Calibri", size=10, bold=True, color=BRAND_RED)
    for i, h in enumerate(["Mês", "Receita", "Despesa", "Lucro", "Alunos"], start=1):
        ws.cell(row=25, column=i, value=h)
    apply_header_row(ws, 25, 1, 5)
    mes_nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    ano = date.today().year
    mes_atual = date.today().month
    for i in range(6):
        m = mes_atual - 5 + i
        a = ano
        if m <= 0:
            m += 12
            a -= 1
        r = 26 + i
        ws.cell(row=r, column=1, value=mes_nomes[m - 1]).border = border_thin
        for c in range(2, 6):
            cell = ws.cell(row=r, column=c, value=0)
            cell.border = border_thin
            if c < 5:
                _money(cell)

    _make_table(ws, "tbSerieMensal", "A25:E31")

    set_column_widths(
        ws,
        {1: 14, 2: 14, 3: 12, 4: 16, 5: 14, 6: 14, 7: 14, 8: 8, 9: 4, 10: 6, 11: 40, 12: 14},
    )
    ws.freeze_panes = "A10"
