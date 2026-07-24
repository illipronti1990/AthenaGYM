"""Sprint 8.0 — Controle de Acesso, Presenca e Frequencia."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    apply_header_row,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_normal,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()
ANO = HOJE.year

ACESSO_HEADERS = [
    "ID", "Matrícula", "Nome", "Data", "Entrada", "Saída",
    "Tempo Permanência", "Forma Acesso", "Responsável", "Status",
]
PRESENCA_BD_HEADERS = ["Matrícula", "Nome", "Data", "Presente", "Professor", "Aula", "Unidade"]

ACESSO_SEED = [
    (1, f"ATH-{ANO}-000001", "Mariana Oliveira", HOJE, "07:58", "09:24", "01:26", "Manual", "Recepção", "Liberado"),
    (2, f"ATH-{ANO}-000002", "Pedro Henrique Alves", HOJE, "18:05", "", "", "Manual", "Recepção", "Liberado"),
    (3, f"ATH-{ANO}-000003", "Fernanda Ribeiro", HOJE - timedelta(days=1), "08:10", "09:05", "00:55", "QR Code", "Recepção", "Liberado"),
    (4, f"ATH-{ANO}-000004", "Thiago Ferreira", HOJE - timedelta(days=1), "06:30", "07:45", "01:15", "Manual", "Recepção", "Liberado"),
    (5, f"ATH-{ANO}-000005", "Beatriz Nogueira", HOJE - timedelta(days=2), "19:00", "20:10", "01:10", "Manual", "Recepção", "Liberado"),
    (6, f"ATH-{ANO}-000001", "Mariana Oliveira", HOJE - timedelta(days=1), "07:05", "08:20", "01:15", "Manual", "Recepção", "Liberado"),
    (7, f"ATH-{ANO}-000002", "Pedro Henrique Alves", HOJE - timedelta(days=3), "18:00", "19:10", "01:10", "Manual", "Recepção", "Liberado"),
]

PRESENCA_BD_SEED = [
    (f"ATH-{ANO}-000001", "Mariana Oliveira", HOJE, "Sim", "Carlos Mendes", "Musculação", "Unidade Centro"),
    (f"ATH-{ANO}-000002", "Pedro Henrique Alves", HOJE, "Sim", "Ana Paula Souza", "Musculação", "Unidade Centro"),
    (f"ATH-{ANO}-000003", "Fernanda Ribeiro", HOJE - timedelta(days=1), "Sim", "Rafael Lima", "Funcional", "Unidade Centro"),
    (f"ATH-{ANO}-000004", "Thiago Ferreira", HOJE - timedelta(days=1), "Sim", "Rafael Lima", "Personal", "Unidade Centro"),
    (f"ATH-{ANO}-000005", "Beatriz Nogueira", HOJE - timedelta(days=2), "Sim", "Juliana Costa", "Musculação", "Unidade Centro"),
    (f"ATH-{ANO}-000001", "Mariana Oliveira", HOJE - timedelta(days=1), "Sim", "Carlos Mendes", "Musculação", "Unidade Centro"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def _bd_sheet(wb, name, headers, rows, table, widths, date_cols=()) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val if val is not None else "")
            cell.border = border_thin
            cell.font = font_normal
            if c_idx in date_cols and val:
                _date_fmt(cell)
    last = 1 + len(rows)
    for c in range(1, len(headers) + 1):
        ws.cell(row=last + 1, column=c).border = border_thin
    end = get_column_letter(len(headers))
    _make_table(ws, table, f"A1:{end}{last + 1}")
    set_column_widths(ws, widths)
    ws.freeze_panes = "A2"


def build_bd_acessos(wb) -> None:
    _bd_sheet(
        wb, "BD_ACESSOS", ACESSO_HEADERS, ACESSO_SEED, "tbAcessos",
        {1: 6, 2: 16, 3: 22, 4: 12, 5: 10, 6: 10, 7: 14, 8: 12, 9: 14, 10: 12},
        date_cols={4},
    )


def build_bd_presencas(wb) -> None:
    _bd_sheet(
        wb, "BD_PRESENCAS", PRESENCA_BD_HEADERS, PRESENCA_BD_SEED, "tbPresencasBD",
        {1: 16, 2: 22, 3: 12, 4: 10, 5: 16, 6: 14, 7: 16},
        date_cols={3},
    )


def build_acesso_ui(wb) -> None:
    """26_ACESSO — tela da recepcao."""
    ws = wb.create_sheet("26_ACESSO")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=58, cols=13)
    add_sidebar(ws, active="26_ACESSO", rows=58, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 18, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 14, 11: 12},
    )
    _title(ws, "CONTROLE DE ACESSO")

    paint_kpi_card(ws, 7, 3, "Entradas hoje", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Dentro agora", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Bloqueios", 0, False, 2)
    paint_kpi_card(ws, 7, 9, "Ausentes +15d", 0, False, 2)

    ws.merge_cells("C11:F11")
    ws["C11"] = "IDENTIFICACAO"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    for r, lab in ((12, "Matricula"), (13, "CPF"), (14, "Forma Acesso")):
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).font = font_header
        ws.cell(row=r, column=4).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
    ws["D14"] = "Manual"

    ws.merge_cells("C16:F16")
    ws["C16"] = "DADOS DO ALUNO"
    ws["C16"].font = font_section
    ws["C16"].fill = fill_brand
    for r, lab in (
        (17, "Aluno"), (18, "Plano"), (19, "Situacao"),
        (20, "Mensalidade"), (21, "Professor"), (22, "Resultado"),
    ):
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).font = font_header
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(row=r, column=4).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
    ws["D22"].font = Font(name="Georgia", size=14, bold=True, color=BRAND_RED)

    ws.row_dimensions[24].height = 36
    ws.row_dimensions[25].height = 36

    ws.merge_cells("H11:K11")
    ws["H11"] = "ULTIMOS ACESSOS (ALUNO)"
    ws["H11"].font = font_section
    ws["H11"].fill = fill_brand
    for i, h in enumerate(["Data", "Entrada", "Saida", "Status"], start=8):
        ws.cell(row=12, column=i, value=h).fill = fill_gold
        ws.cell(row=12, column=i).font = font_header
    for i in range(8):
        r = 13 + i
        for c in range(8, 12):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C27:K27")
    ws["C27"] = "ACESSOS DE HOJE"
    ws["C27"].font = font_section
    ws["C27"].fill = fill_brand
    for i, h in enumerate(["ID", "Hora", "Aluno", "Forma", "Status", "Saida"], start=3):
        ws.cell(row=28, column=i, value=h).fill = fill_gold
        ws.cell(row=28, column=i).font = font_header
    for i in range(12):
        r = 29 + i
        for c in range(3, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C42:K42")
    ws["C42"] = "ALUNOS AUSENTES / RISCO"
    ws["C42"].font = font_section
    ws["C42"].fill = fill_brand
    for i, h in enumerate(["Matricula", "Aluno", "Dias sem acesso", "Motivo"], start=3):
        ws.cell(row=43, column=i, value=h).fill = fill_gold
        ws.cell(row=43, column=i).font = font_header
    for i in range(8):
        r = 44 + i
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C53:K53")
    ws["C53"] = "Regras via BD_PARAMETROS (grupo Acesso) · Futuro: QR / Biometria / RFID via modIntegracoes"
    ws["C53"].font = Font(name="Calibri", size=9, color="666666")


def build_dash_frequencia(wb) -> None:
    """27_DASH_FREQUENCIA — KPIs, pico, ranking."""
    ws = wb.create_sheet("27_DASH_FREQUENCIA")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=48, cols=13)
    add_sidebar(ws, active="26_ACESSO", rows=48, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 12, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 12, 11: 12},
    )
    _title(ws, "DASHBOARD FREQUENCIA")

    paint_kpi_card(ws, 7, 3, "Entradas hoje", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Media diaria", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Pico", "—", False, 2)
    paint_kpi_card(ws, 7, 9, "Bloqueios", 0, False, 2)

    ws.merge_cells("C11:E11")
    ws["C11"] = "HORARIOS (HOJE)"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    for i, h in enumerate(["Hora", "Qtd", "Barra"], start=3):
        ws.cell(row=12, column=i, value=h).fill = fill_gold
        ws.cell(row=12, column=i).font = font_header
    for i in range(16):
        r = 13 + i
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("G11:K11")
    ws["G11"] = "RANKING FREQUENCIA (MES)"
    ws["G11"].font = font_section
    ws["G11"].fill = fill_brand
    for i, h in enumerate(["#", "Aluno", "Presencas", "Barra"], start=7):
        ws.cell(row=12, column=i, value=h).fill = fill_gold
        ws.cell(row=12, column=i).font = font_header
    for i in range(10):
        r = 13 + i
        for c in range(7, 11):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C30:E30")
    ws["C30"] = "RESUMO"
    ws["C30"].font = font_section
    ws["C30"].fill = fill_brand
    for i, lab in enumerate(["Ativos hoje", "Ausentes +15d", "Ausentes +30d", "Tempo medio"]):
        r = 31 + i
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=4, value=0 if i < 3 else "—").fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
        ws.cell(row=r, column=4).font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)

    ws.merge_cells("G30:K30")
    ws["G30"] = "FAIXAS MAIS UTILIZADAS"
    ws["G30"].font = font_section
    ws["G30"].fill = fill_brand
    for i, h in enumerate(["Faixa", "Entradas"], start=7):
        ws.cell(row=31, column=i, value=h).fill = fill_gold
        ws.cell(row=31, column=i).font = font_header
    for i in range(5):
        r = 32 + i
        for c in range(7, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin
