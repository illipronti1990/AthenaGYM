"""Sprint 9.0 — PDV Inteligente + Gestao de Estoque."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()
ANO = HOJE.year

UNIDADES_HEADERS = [
    "ID", "EmpresaID", "Nome", "Código", "CNPJ", "Telefone", "WhatsApp", "Email",
    "CEP", "Endereço", "Cidade", "Estado", "Responsável", "Status", "DataCadastro",
]
PRODUTOS_HEADERS = [
    "Código", "Código de Barras", "Produto", "Categoria", "Marca", "Unidade Medida",
    "Custo", "Preço Venda", "Estoque Atual", "Estoque Mínimo", "Estoque Máximo",
    "Localização", "Status", "Unidade", "UnidadeID", "Classe ABC",
]
FORNECEDORES_HEADERS = [
    "Código", "Empresa", "CNPJ", "Contato", "Telefone", "E-mail", "Cidade", "Condição de Pagamento",
]
COMPRAS_HEADERS = ["ID", "Compra", "Produto", "Código", "Quantidade", "Valor", "Fornecedor", "Data", "Unidade", "UnidadeID"]
MOV_HEADERS = ["ID", "Data", "Produto", "Código", "Tipo", "Quantidade", "Usuário", "Obs", "Unidade", "UnidadeID"]
LOTES_HEADERS = ["ID", "Código", "Produto", "Lote", "Fabricação", "Validade", "Quantidade", "Status"]
VENDAS_HEADERS = [
    "ID", "Data", "Cliente", "Matrícula", "Itens", "Total", "Custo", "Lucro",
    "Forma", "Status", "Usuário", "Na Mensalidade", "Unidade", "UnidadeID",
]
VENDA_ITENS_HEADERS = ["ID", "Venda ID", "Código", "Produto", "Qtde", "Preço", "Subtotal", "Custo"]
KITS_HEADERS = ["Código", "Kit", "Preço", "Status"]
KIT_ITENS_HEADERS = ["Kit Código", "Produto Código", "Produto", "Qtde"]

UNIDADES_SEED = [
    (
        1, 1, "ATHENAS GYM Matriz", "MX", "12.345.678/0001-90",
        "(11) 3000-0001", "(11) 99000-0001", "matriz@athenas.gym",
        "01310-100", "Av. Paulista, 1000", "São Paulo", "SP", "Carlos Mendes", "Ativa", HOJE,
    ),
    (
        2, 1, "ATHENAS GYM Zona Sul", "ZS", "12.345.678/0002-71",
        "(11) 3000-0002", "(11) 99000-0002", "zonasul@athenas.gym",
        "04038-001", "Av. Ibirapuera, 500", "São Paulo", "SP", "Ana Paula Souza", "Ativa", HOJE,
    ),
    (
        3, 2, "ATHENAS Campinas Centro", "CP", "23.456.789/0001-01",
        "(19) 3000-2001", "(19) 99000-2001", "campinas@athenas.gym",
        "13010-100", "Av. Norte-Sul, 500", "Campinas", "SP", "Marcos Silva", "Ativa", HOJE,
    ),
    (
        4, 3, "ATHENAS Santos Gonzaga", "ST", "34.567.890/0001-12",
        "(13) 3000-3001", "(13) 99000-3001", "santos@athenas.gym",
        "11010-100", "Av. Ana Costa, 200", "Santos", "SP", "Patricia Reis", "Ativa", HOJE,
    ),
]

PRODUTOS_SEED = [
    ("PRD-001", "7891000100011", "Whey Protein 900g", "Suplementos", "AthMax", "UN",
     89.90, 149.90, 20, 5, 50, "Prateleira A1", "Ativo", "ATHENAS GYM Matriz", 1, "A"),
    ("PRD-002", "7891000100028", "Creatina 300g", "Suplementos", "AthMax", "UN",
     45.00, 79.90, 3, 10, 40, "Prateleira A2", "Ativo", "ATHENAS GYM Matriz", 1, "A"),
    ("PRD-003", "7891000100035", "Barra Proteica", "Barras de Proteína", "FitBar", "UN",
     6.50, 12.90, 40, 10, 100, "Balcão", "Ativo", "ATHENAS GYM Matriz", 1, "B"),
    ("PRD-004", "7891000100042", "Camiseta Athenas", "Roupas", "Athenas", "UN",
     25.00, 69.90, 30, 5, 60, "Vitrine", "Ativo", "ATHENAS GYM Matriz", 1, "B"),
    ("PRD-005", "7891000100059", "Garrafa Squeeze", "Garrafas", "Hydra", "UN",
     12.00, 39.90, 20, 5, 40, "Balcão", "Ativo", "ATHENAS GYM Matriz", 1, "C"),
    ("PRD-006", "7891000100066", "Luvas Treino", "Luvas", "GripPro", "PAR",
     18.00, 49.90, 15, 3, 30, "Vitrine", "Ativo", "ATHENAS GYM Matriz", 1, "C"),
    ("PRD-007", "7891000100073", "Pré-Treino 300g", "Suplementos", "AthMax", "UN",
     55.00, 99.90, 12, 4, 30, "Prateleira A3", "Ativo", "ATHENAS GYM Matriz", 1, "A"),
    ("PRD-008", "7891000100080", "Coqueteleira", "Coqueteleiras", "ShakeIt", "UN",
     8.00, 29.90, 25, 5, 50, "Balcão", "Ativo", "ATHENAS GYM Matriz", 1, "C"),
    ("PRD-009", "7891000100097", "Água 500ml", "Bebidas", "Crystal", "UN",
     1.20, 4.00, 80, 20, 200, "Geladeira", "Ativo", "ATHENAS GYM Matriz", 1, "B"),
    ("PRD-010", "7891000100103", "Vitamina D", "Vitaminas", "VitaPlus", "UN",
     22.00, 49.90, 8, 5, 25, "Prateleira B1", "Ativo", "ATHENAS GYM Matriz", 1, "B"),
    ("PRD-001", "7891000100011", "Whey Protein 900g", "Suplementos", "AthMax", "UN",
     89.90, 149.90, 8, 3, 30, "Prateleira A1", "Ativo", "ATHENAS GYM Zona Sul", 2, "A"),
    ("PRD-003", "7891000100035", "Barra Proteica", "Barras de Proteína", "FitBar", "UN",
     6.50, 12.90, 15, 5, 50, "Balcão", "Ativo", "ATHENAS GYM Zona Sul", 2, "B"),
]

FORNECEDORES_SEED = [
    ("FOR-001", "Nutri Distribuidora LTDA", "12.345.678/0001-90", "Marcos", "(11) 3333-1001",
     "compras@nutri.com", "São Paulo", "30 dias"),
    ("FOR-002", "Fit Wear Confecções", "98.765.432/0001-10", "Lucia", "(11) 3333-2002",
     "vendas@fitwear.com", "Guarulhos", "À vista"),
]

COMPRAS_SEED = [
    (1, "CMP-001", "Whey Protein 900g", "PRD-001", 10, 899.00, "Nutri Distribuidora LTDA", HOJE - timedelta(days=10), "ATHENAS GYM Matriz", 1),
    (2, "CMP-001", "Creatina 300g", "PRD-002", 15, 675.00, "Nutri Distribuidora LTDA", HOJE - timedelta(days=10), "ATHENAS GYM Matriz", 1),
]

MOV_SEED = [
    (1, HOJE - timedelta(days=10), "Whey Protein 900g", "PRD-001", "Entrada", 10, "Admin", "Compra CMP-001", "ATHENAS GYM Matriz", 1),
    (2, HOJE - timedelta(days=10), "Creatina 300g", "PRD-002", "Entrada", 15, "Admin", "Compra CMP-001", "ATHENAS GYM Matriz", 1),
    (3, HOJE - timedelta(days=1), "Creatina 300g", "PRD-002", "Venda", 2, "Recepção", "PDV", "ATHENAS GYM Matriz", 1),
    (4, HOJE, "Barra Proteica", "PRD-003", "Venda", 3, "Recepção", "PDV", "ATHENAS GYM Matriz", 1),
]

LOTES_SEED = [
    (1, "PRD-001", "Whey Protein 900g", "Lote-W25", HOJE - timedelta(days=60), HOJE + timedelta(days=300), 20, "OK"),
    (2, "PRD-002", "Creatina 300g", "Lote-C25", HOJE - timedelta(days=30), HOJE + timedelta(days=20), 3, "Vencendo"),
    (3, "PRD-010", "Vitamina D", "Lote-VD", HOJE - timedelta(days=90), HOJE + timedelta(days=45), 8, "OK"),
]

VENDAS_SEED = [
    (1, HOJE - timedelta(days=1), "Balcão", "", 1, 159.80, 90.00, 69.80, "PIX", "Finalizada", "Recepção", "NÃO", "ATHENAS GYM Matriz", 1),
    (2, HOJE, "Balcão", "", 1, 38.70, 19.50, 19.20, "Dinheiro", "Finalizada", "Recepção", "NÃO", "ATHENAS GYM Matriz", 1),
]

VENDA_ITENS_SEED = [
    (1, 1, "PRD-002", "Creatina 300g", 2, 79.90, 159.80, 90.00),
    (2, 2, "PRD-003", "Barra Proteica", 3, 12.90, 38.70, 19.50),
]

KITS_SEED = [
    ("KIT-001", "Kit Hipertrofia", 239.90, "Ativo"),
]

KIT_ITENS_SEED = [
    ("KIT-001", "PRD-001", "Whey Protein 900g", 1),
    ("KIT-001", "PRD-002", "Creatina 300g", 1),
    ("KIT-001", "PRD-008", "Coqueteleira", 1),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def _bd_sheet(wb, name, headers, rows, table, widths, date_cols=(), money_cols=()) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).fill = fill_gold
        ws.cell(row=1, column=i).font = font_header
        ws.cell(row=1, column=i).border = border_thin
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx in date_cols:
                _date_fmt(cell)
            if c_idx in money_cols:
                cell.number_format = "R$ #,##0.00"
    last = max(1 + len(rows), 2)
    _make_table(ws, table, f"A1:{get_column_letter(len(headers))}{last}")
    set_column_widths(ws, widths)
    ws.sheet_state = "hidden"


def build_bd_unidades(wb) -> None:
    widths = {i: 14 for i in range(1, 16)}
    widths.update({1: 6, 2: 10, 3: 26, 4: 8, 5: 18, 10: 22, 13: 16})
    _bd_sheet(
        wb, "BD_UNIDADES", UNIDADES_HEADERS, UNIDADES_SEED, "tbUnidades",
        widths, date_cols={15},
    )


def build_bd_produtos(wb) -> None:
    _bd_sheet(
        wb, "BD_PRODUTOS", PRODUTOS_HEADERS, PRODUTOS_SEED, "tbProdutos",
        {1: 10, 2: 14, 3: 20, 4: 14, 5: 10, 6: 10, 7: 10, 8: 10, 9: 10, 10: 10, 11: 10, 12: 14, 13: 10, 14: 20, 15: 10, 16: 10},
        money_cols={7, 8},
    )


def build_bd_fornecedores(wb) -> None:
    _bd_sheet(
        wb, "BD_FORNECEDORES", FORNECEDORES_HEADERS, FORNECEDORES_SEED, "tbFornecedores",
        {1: 10, 2: 28, 3: 18, 4: 12, 5: 14, 6: 20, 7: 12, 8: 14},
    )


def build_bd_compras(wb) -> None:
    _bd_sheet(
        wb, "BD_COMPRAS", COMPRAS_HEADERS, COMPRAS_SEED, "tbCompras",
        {1: 6, 2: 10, 3: 20, 4: 10, 5: 10, 6: 10, 7: 24, 8: 12, 9: 20, 10: 10},
        date_cols={8}, money_cols={6},
    )


def build_bd_movimentacao_estoque(wb) -> None:
    _bd_sheet(
        wb, "BD_MOVIMENTACAO_ESTOQUE", MOV_HEADERS, MOV_SEED, "tbMovEstoque",
        {1: 6, 2: 12, 3: 20, 4: 10, 5: 12, 6: 10, 7: 12, 8: 16, 9: 20, 10: 10},
        date_cols={2},
    )


def build_bd_lotes(wb) -> None:
    _bd_sheet(
        wb, "BD_LOTES", LOTES_HEADERS, LOTES_SEED, "tbLotes",
        {1: 6, 2: 10, 3: 20, 4: 12, 5: 12, 6: 12, 7: 10, 8: 10},
        date_cols={5, 6},
    )


def build_bd_vendas(wb) -> None:
    _bd_sheet(
        wb, "BD_VENDAS", VENDAS_HEADERS, VENDAS_SEED, "tbVendas",
        {1: 6, 2: 12, 3: 14, 4: 14, 5: 8, 6: 10, 7: 10, 8: 10, 9: 10, 10: 12, 11: 12, 12: 12, 13: 20, 14: 10},
        date_cols={2}, money_cols={6, 7, 8},
    )


def build_bd_venda_itens(wb) -> None:
    _bd_sheet(
        wb, "BD_VENDA_ITENS", VENDA_ITENS_HEADERS, VENDA_ITENS_SEED, "tbVendaItens",
        {1: 6, 2: 10, 3: 10, 4: 20, 5: 8, 6: 10, 7: 10, 8: 10},
        money_cols={6, 7, 8},
    )


def build_bd_kits(wb) -> None:
    _bd_sheet(
        wb, "BD_KITS", KITS_HEADERS, KITS_SEED, "tbKits",
        {1: 10, 2: 18, 3: 10, 4: 10},
        money_cols={3},
    )


def build_bd_kit_itens(wb) -> None:
    _bd_sheet(
        wb, "BD_KIT_ITENS", KIT_ITENS_HEADERS, KIT_ITENS_SEED, "tbKitItens",
        {1: 10, 2: 12, 3: 20, 4: 8},
    )


def build_pdv_ui(wb) -> None:
    """28_PDV — ponto de venda."""
    ws = wb.create_sheet("28_PDV")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=52, cols=13)
    add_sidebar(ws, active="28_PDV", rows=52, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 16, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 12, 11: 12},
    )
    _title(ws, "PDV — PONTO DE VENDA")

    paint_kpi_card(ws, 7, 3, "Vendas hoje", "R$ 0", False, 2)
    paint_kpi_card(ws, 7, 5, "Itens", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Ticket médio", "R$ 0", False, 2)
    paint_kpi_card(ws, 7, 9, "Lucro hoje", "R$ 0", False, 2)

    ws.merge_cells("C11:F11")
    ws["C11"] = "VENDA ATUAL"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    for r, lab in (
        (12, "Cliente / Aluno"),
        (13, "Matrícula"),
        (14, "Código / Barras"),
        (15, "Produto"),
        (16, "Preço"),
        (17, "Quantidade"),
        (18, "Forma Pagamento"),
        (19, "Kit (opcional)"),
    ):
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=3).font = font_header
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(row=r, column=4).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
    ws["D12"] = "Balcão"
    ws["D17"] = 1
    ws["D18"] = "PIX"

    ws.row_dimensions[21].height = 36
    ws.row_dimensions[22].height = 36

    ws.merge_cells("H11:K11")
    ws["H11"] = "TOTAL"
    ws["H11"].font = font_section
    ws["H11"].fill = fill_brand
    ws.merge_cells("H12:K14")
    ws["H12"] = 0
    ws["H12"].font = Font(name="Georgia", size=28, bold=True, color=BRAND_RED)
    ws["H12"].number_format = '"R$" #,##0.00'
    ws["H15"] = "Itens no carrinho"
    ws["H16"] = 0
    ws["H16"].font = Font(name="Georgia", size=14, bold=True, color=BRAND_RED)

    ws.merge_cells("C24:K24")
    ws["C24"] = "CARRINHO"
    ws["C24"].font = font_section
    ws["C24"].fill = fill_brand
    for i, h in enumerate(["Código", "Produto", "Qtde", "Preço", "Subtotal", ""], start=3):
        ws.cell(row=25, column=i, value=h).fill = fill_gold
        ws.cell(row=25, column=i).font = font_header
    for i in range(12):
        r = 26 + i
        for c in range(3, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C40:K40")
    ws["C40"] = "ÚLTIMAS VENDAS"
    ws["C40"].font = font_section
    ws["C40"].fill = fill_brand
    for i, h in enumerate(["ID", "Hora", "Cliente", "Total", "Forma", "Status"], start=3):
        ws.cell(row=41, column=i, value=h).fill = fill_gold
        ws.cell(row=41, column=i).font = font_header
    for i in range(6):
        r = 42 + i
        for c in range(3, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C49:K49")
    ws["C49"] = "Fluxo: estoque → venda → financeiro → caixa → dashboard · Aluno pode lançar na mensalidade"
    ws["C49"].font = Font(name="Calibri", size=9, color="666666")


def build_inventario_ui(wb) -> None:
    """29_INVENTARIO — contagem física × sistema."""
    ws = wb.create_sheet("29_INVENTARIO")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=42, cols=13)
    add_sidebar(ws, active="09_ESTOQUE", rows=42, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 12, 4: 22, 5: 10, 6: 10, 7: 10, 8: 18, 9: 12, 10: 12},
    )
    _title(ws, "INVENTÁRIO")

    paint_kpi_card(ws, 7, 3, "Itens", 0, False, 2)
    paint_kpi_card(ws, 7, 5, "Diferenças", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Ajustes", 0, False, 2)

    ws.merge_cells("C11:I11")
    ws["C11"] = "CONTAGEM"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    for i, h in enumerate(["Código", "Produto", "Sistema", "Físico", "Diferença", "Observação"], start=3):
        ws.cell(row=12, column=i, value=h).fill = fill_gold
        ws.cell(row=12, column=i).font = font_header
    for i in range(15):
        r = 13 + i
        for c in range(3, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.row_dimensions[30].height = 36
    ws.merge_cells("C32:I32")
    ws["C32"] = "Ao finalizar, o sistema gera Ajuste em BD_MOVIMENTACAO_ESTOQUE sem apagar histórico."
    ws["C32"].font = Font(name="Calibri", size=9, color="666666")


def build_dash_pdv(wb) -> None:
    """30_DASH_PDV — vendas, ranking, curva ABC."""
    ws = wb.create_sheet("30_DASH_PDV")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=44, cols=13)
    add_sidebar(ws, active="28_PDV", rows=44, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 14, 4: 12, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 12},
    )
    _title(ws, "DASHBOARD PDV / ESTOQUE")

    paint_kpi_card(ws, 7, 3, "Vendas hoje", "R$ 0", False, 2)
    paint_kpi_card(ws, 7, 5, "Produtos vendidos", 0, False, 2)
    paint_kpi_card(ws, 7, 7, "Ticket médio", "R$ 0", False, 2)
    paint_kpi_card(ws, 7, 9, "Lucro", "R$ 0", False, 2)

    paint_kpi_card(ws, 11, 3, "Valor estoque", "R$ 0", False, 2)
    paint_kpi_card(ws, 11, 5, "Críticos", 0, False, 2)
    paint_kpi_card(ws, 11, 7, "Vencendo", 0, False, 2)
    paint_kpi_card(ws, 11, 9, "Margem média", "0%", False, 2)

    ws.merge_cells("C16:F16")
    ws["C16"] = "MAIS VENDIDOS (MÊS)"
    ws["C16"].font = font_section
    ws["C16"].fill = fill_brand
    for i, h in enumerate(["Produto", "Qtde", "Faturamento", "Barra"], start=3):
        ws.cell(row=17, column=i, value=h).fill = fill_gold
        ws.cell(row=17, column=i).font = font_header
    for i in range(8):
        r = 18 + i
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("H16:K16")
    ws["H16"] = "CURVA ABC"
    ws["H16"].font = font_section
    ws["H16"].fill = fill_brand
    for i, h in enumerate(["Classe", "Produtos", "% Fat.", "Lista"], start=8):
        ws.cell(row=17, column=i, value=h).fill = fill_gold
        ws.cell(row=17, column=i).font = font_header
    for i, cls in enumerate(["A", "B", "C"]):
        r = 18 + i
        ws.cell(row=r, column=8, value=cls).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin
        for c in range(9, 12):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C28:K28")
    ws["C28"] = "ALERTAS — ESTOQUE BAIXO / LOTES VENCENDO"
    ws["C28"].font = font_section
    ws["C28"].fill = fill_brand
    for i, h in enumerate(["Tipo", "Código", "Produto", "Detalhe", "Ação"], start=3):
        ws.cell(row=29, column=i, value=h).fill = fill_gold
        ws.cell(row=29, column=i).font = font_header
    for i in range(8):
        r = 30 + i
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin
