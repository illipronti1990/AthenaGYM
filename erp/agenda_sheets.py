"""Sprint 5.1 — BD_EVENTOS + aba 20_AGENDA (Central Operacional)."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    ICON,
    add_sidebar,
    add_top_bar,
    apply_header_row,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_normal,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

TIPOS_EVENTO = [
    "Cadastro",
    "Mensalidade",
    "Avaliação Física",
    "Aula Experimental",
    "Renovação",
    "Manutenção",
    "Aniversário",
    "Estoque",
    "Professor",
    "Marketing",
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def build_bd_eventos(wb) -> None:
    """BD_EVENTOS — motor de eventos da Central Operacional."""
    ws = wb.create_sheet("BD_EVENTOS")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False

    headers = [
        "ID",
        "Tipo",
        "Título",
        "Referência",
        "Data",
        "Hora",
        "Responsável",
        "Status",
        "Prioridade",
        "Módulo",
        "Observação",
        "Origem",
        "UnidadeID",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    hoje = date.today()
    seed = [
        (1, "Mensalidade", "Mensalidade vence hoje", "Aluno demo", hoje, "08:00", "Recepção", "Pendente", "Alta", "Financeiro", "", "Auto", 1),
        (2, "Avaliação Física", "Avaliação física agendada", "Maria Silva", hoje, "10:00", "Professor", "Pendente", "Média", "Operacional", "", "Manual", 1),
        (3, "Aula Experimental", "Aula experimental", "Carlos Lima", hoje, "19:00", "Recepção", "Pendente", "Baixa", "Comercial", "", "Manual", 2),
        (4, "Manutenção", "Revisão de equipamento", "Esteira 03", hoje, "16:00", "Manutenção", "Pendente", "Alta", "Equipamentos", "", "Auto", 1),
        (5, "Estoque", "Estoque mínimo atingido", "Creatina", hoje, "09:00", "Estoque", "Pendente", "Média", "Estoque", "", "Auto", 1),
        (6, "Aniversário", "Aniversário do aluno", "João Pedro", hoje, "00:00", "Marketing", "Pendente", "Baixa", "Comercial", "", "Auto", 2),
        (7, "Mensalidade", "Mensalidade vence em 5 dias", "Ana Costa", hoje + timedelta(days=5), "08:00", "Financeiro", "Pendente", "Média", "Financeiro", "", "Auto", 1),
        (8, "Renovação", "Plano próximo do vencimento", "Premium — Pedro", hoje + timedelta(days=30), "09:00", "Recepção", "Pendente", "Média", "Comercial", "", "Auto", 2),
    ]

    for r_idx, row in enumerate(seed, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal
            if c_idx == 5 and isinstance(val, date):
                _date_fmt(cell)

    last = 1 + len(seed)
    for c in range(1, len(headers) + 1):
        ws.cell(row=last + 1, column=c).value = None
    last = last + 1
    _make_table(ws, "tbEventos", f"A1:M{last}")

    set_column_widths(
        ws,
        {1: 6, 2: 16, 3: 28, 4: 18, 5: 12, 6: 8, 7: 12, 8: 12, 9: 10, 10: 14, 11: 20, 12: 10, 13: 10},
    )
    ws.freeze_panes = "A2"


def build_agenda(wb) -> None:
    """20_AGENDA — Central Operacional (hoje / semana / mês / alertas)."""
    ws = wb.create_sheet("20_AGENDA")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=55, cols=14)
    add_sidebar(ws, active="01_DASHBOARD", rows=55, labels=False)
    add_top_bar(ws, start_col=2, end_col=13)

    set_column_widths(
        ws,
        {1: 24, 2: 2, 3: 12, 4: 22, 5: 18, 6: 10, 7: 12, 8: 12, 9: 12, 10: 12, 11: 12, 12: 10},
    )

    ws.merge_cells("C5:H5")
    ws["C5"] = f"{ICON.get('dashboard', '📅')}  AGENDA INTELIGENTE"
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light

    ws.merge_cells("I5:L5")
    ws["I5"] = "Motor de eventos · atualizado por modAgenda"
    ws["I5"].font = Font(name="Calibri", size=9, color="666666")

    # KPIs do dia
    paint_kpi_card(ws, 7, 3, "Eventos hoje", "='BI_BASE'!E28", False, 2)
    paint_kpi_card(ws, 7, 5, "Alta prioridade", "='BI_BASE'!E29", False, 2)
    paint_kpi_card(ws, 7, 7, "Mensalidades", "='BI_BASE'!E30", False, 2)
    paint_kpi_card(ws, 7, 9, "Avaliações", "='BI_BASE'!E31", False, 2)

    # Agenda do dia
    ws.merge_cells("C11:H11")
    ws["C11"] = "📅 HOJE — PRIORIDADES"
    ws["C11"].font = font_section
    ws["C11"].fill = fill_brand
    for i, h in enumerate(["Pri", "Hora", "Tipo", "Título", "Referência", "Responsável"], start=3):
        ws.cell(row=12, column=i, value=h).fill = fill_gold
        ws.cell(row=12, column=i).font = font_header

    for i in range(12):
        r = 13 + i
        for c in range(3, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin
        # Preenchido por VBA AtualizarAgendaUI (linhas 13–24)

    # Semana
    ws.merge_cells("C26:I26")
    ws["C26"] = "📅 SEMANA — QUANTIDADE DE EVENTOS"
    ws["C26"].font = font_section
    ws["C26"].fill = fill_brand
    dias = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    for i, d in enumerate(dias):
        ws.cell(row=27, column=3 + i, value=d).fill = fill_gold
        ws.cell(row=27, column=3 + i).font = font_header
        ws.cell(row=27, column=3 + i).alignment = Alignment(horizontal="center")
        cell = ws.cell(row=28, column=3 + i, value=0)
        cell.fill = fill_panel
        cell.border = border_thin
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Georgia", size=14, bold=True, color=BRAND_RED)

    # Alertas da agenda
    ws.merge_cells("C30:H30")
    ws["C30"] = "🔔 ALERTAS OPERACIONAIS"
    ws["C30"].font = font_section
    ws["C30"].fill = fill_brand
    for i, h in enumerate(["Pri", "Alerta", "Destino"], start=3):
        ws.cell(row=31, column=i, value=h).fill = fill_gold
        ws.cell(row=31, column=i).font = font_header
    for i in range(6):
        r = 32 + i
        for c in (3, 4, 5):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

    # Aniversariantes
    ws.merge_cells("I30:L30")
    ws["I30"] = "🎂 ANIVERSARIANTES HOJE"
    ws["I30"].font = font_section
    ws["I30"].fill = fill_brand
    for i, h in enumerate(["Nome", "Idade"], start=9):
        ws.cell(row=31, column=i, value=h).fill = fill_gold
        ws.cell(row=31, column=i).font = font_header
    for i in range(6):
        r = 32 + i
        for c in (9, 10):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    # Distribuição por tipo (gráfico)
    ws.merge_cells("C39:F39")
    ws["C39"] = "EVENTOS DE HOJE POR TIPO"
    ws["C39"].font = font_section
    ws["C39"].fill = fill_brand
    for i, h in enumerate(["Tipo", "Qtd"], start=3):
        ws.cell(row=40, column=i, value=h).fill = fill_gold
        ws.cell(row=40, column=i).font = font_header
    for i, tipo in enumerate(["Mensalidade", "Avaliação Física", "Aula Experimental", "Renovação", "Manutenção", "Aniversário", "Estoque"]):
        r = 41 + i
        ws.cell(row=r, column=3, value=tipo).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.cell(row=r, column=4, value=0).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin
        ws.cell(row=r, column=5, value='=IF(D{0}=0,"",REPT("█",MIN(20,D{0})))'.format(r))
        ws.cell(row=r, column=5).fill = fill_panel

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Hoje por tipo"
    data = Reference(ws, min_col=4, min_row=40, max_row=47)
    cats = Reference(ws, min_col=3, min_row=41, max_row=47)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 10
    ws.add_chart(chart, "G39")

    ws.merge_cells("C50:L50")
    ws["C50"] = "Visões: Dia · Semana · Professor · Recepção · Financeiro · Manutenção  |  Botões abaixo atualizam o motor"
    ws["C50"].font = Font(name="Calibri", size=9, color="666666")
