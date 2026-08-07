"""Épico 1 — Multi-Tenant: empresas, licenças, config, Master, Nova Academia."""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)

HOJE = date.today()

EMPRESAS_HEADERS = [
    "EmpresaID", "Razão Social", "Nome Fantasia", "CNPJ", "Inscrição Estadual",
    "Telefone", "WhatsApp", "Email", "Site", "CEP", "Endereço", "Número", "Complemento",
    "Bairro", "Cidade", "Estado", "País", "Logo", "Cor Primária", "Cor Secundária",
    "Plano", "Status", "Data Cadastro", "Data Expiração", "FranqueadoraID", "FranqueadoID",
]

EMPRESAS_SEED = [
    (
        0, "ATHENA PLATFORM LTDA", "ATHENA PLATFORM", "00.000.000/0001-00", "",
        "", "", "platform@athena.gym", "https://athena.gym", "", "", "", "",
        "", "São Paulo", "SP", "Brasil", "", "#A3001B", "#D4AF37",
        "Enterprise", "Ativo", HOJE, HOJE + timedelta(days=3650), 0, 0,
    ),
    (
        1, "ATHENA GYM ACADEMIA LTDA", "ATHENA GYM", "12.345.678/0001-90", "ISENTO",
        "(11) 3000-0000", "(11) 99000-0000", "contato@athena.gym", "https://athena.gym",
        "01310-100", "Av. Paulista", "1000", "Sala 10", "Bela Vista", "São Paulo", "SP", "Brasil",
        "", "#A3001B", "#D4AF37", "Enterprise", "Ativo", HOJE - timedelta(days=365), HOJE + timedelta(days=365),
        1, 1,
    ),
    (
        2, "ATHENA GYM CAMPINAS LTDA", "ATHENA Campinas", "23.456.789/0001-01", "ISENTO",
        "(19) 3000-2000", "(19) 99000-2000", "campinas@athena.gym", "https://athena.gym",
        "13010-100", "Av. Norte-Sul", "500", "", "Centro", "Campinas", "SP", "Brasil",
        "", "#A3001B", "#D4AF37", "Enterprise", "Ativo", HOJE - timedelta(days=200), HOJE + timedelta(days=530),
        1, 2,
    ),
    (
        3, "ATHENA GYM SANTOS LTDA", "ATHENA Santos", "34.567.890/0001-12", "ISENTO",
        "(13) 3000-3000", "(13) 99000-3000", "santos@athena.gym", "https://athena.gym",
        "11010-100", "Av. Ana Costa", "200", "", "Gonzaga", "Santos", "SP", "Brasil",
        "", "#A3001B", "#D4AF37", "Enterprise", "Ativo", HOJE - timedelta(days=120), HOJE + timedelta(days=610),
        1, 3,
    ),
]

LICENCAS_HEADERS = ["ID", "EmpresaID", "Chave", "Plano", "Ativação", "Expiração", "Status"]
LICENCAS_SEED = [
    (1, 1, "ATH-ENT-1-DEMO-2026", "Enterprise", HOJE - timedelta(days=30), HOJE + timedelta(days=335), "Ativa"),
    (2, 2, "ATH-ENT-2-CAMP-2026", "Enterprise", HOJE - timedelta(days=30), HOJE + timedelta(days=335), "Ativa"),
    (3, 3, "ATH-ENT-3-SANT-2026", "Enterprise", HOJE - timedelta(days=30), HOJE + timedelta(days=335), "Ativa"),
]

CONFIG_EMP_HEADERS = ["ID", "EmpresaID", "Chave", "Valor"]
CONFIG_EMP_SEED = [
    (1, 1, "CorSistema", "#A3001B"),
    (2, 1, "BloquearInadimplente", "SIM"),
    (3, 1, "DiasTolerancia", "5"),
    (4, 1, "Moeda", "BRL"),
    (5, 1, "TimeZone", "America/Sao_Paulo"),
    (6, 1, "PrefixoMatricula", "ATH"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def _date_fmt(cell) -> None:
    cell.number_format = "DD/MM/YYYY"


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def _bd_sheet(wb, name, headers, rows, table, widths, date_cols=()) -> None:
    ws = wb.create_sheet(name)
    style_sheet_tab(ws, GOLD)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).fill = fill_gold
        ws.cell(row=1, column=i).font = font_header
        ws.cell(row=1, column=i).border = border_thin
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            if c_idx in date_cols:
                _date_fmt(cell)
    last = max(1 + len(rows), 2)
    _make_table(ws, table, f"A1:{get_column_letter(len(headers))}{last}")
    set_column_widths(ws, widths)
    ws.sheet_state = "hidden"


def build_bd_empresas(wb) -> None:
    widths = {i: 14 for i in range(1, 27)}
    widths.update({1: 10, 2: 28, 3: 18, 4: 18, 8: 22, 11: 20, 25: 14, 26: 12})
    _bd_sheet(
        wb, "BD_EMPRESAS", EMPRESAS_HEADERS, EMPRESAS_SEED, "tbEmpresas", widths, date_cols={23, 24}
    )


def build_bd_licencas(wb) -> None:
    _bd_sheet(
        wb, "BD_LICENCAS", LICENCAS_HEADERS, LICENCAS_SEED, "tbLicencas",
        {1: 6, 2: 10, 3: 22, 4: 12, 5: 12, 6: 12, 7: 12},
        date_cols={5, 6},
    )


def build_bd_config_empresa(wb) -> None:
    _bd_sheet(
        wb, "BD_CONFIG_EMPRESA", CONFIG_EMP_HEADERS, CONFIG_EMP_SEED, "tbConfigEmpresa",
        {1: 6, 2: 10, 3: 22, 4: 24},
    )


def build_master_ui(wb) -> None:
    """38_MASTER — painel SuperAdmin da plataforma."""
    ws = wb.create_sheet("38_MASTER")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=48, cols=13)
    add_sidebar(ws, active="38_MASTER", rows=48, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(ws, {1: 24, 2: 2, 3: 14, 4: 14, 5: 12, 6: 12, 7: 14, 8: 12, 9: 12, 10: 12})
    _title(ws, "ATHENA PLATFORM — MASTER")

    paint_kpi_card(ws, 8, 3, "Academias", 0, False, 2)
    paint_kpi_card(ws, 8, 5, "Alunos", 0, False, 2)
    paint_kpi_card(ws, 8, 7, "Faturamento", "R$ 0", False, 2)
    paint_kpi_card(ws, 8, 9, "Usuários online", 0, False, 2)

    ws.merge_cells("C14:K14")
    ws["C14"] = "ACADEMIAS CADASTRADAS"
    ws["C14"].font = font_section
    ws["C14"].fill = fill_brand
    for i, h in enumerate(["ID", "Fantasia", "CNPJ", "Plano", "Status", "Expiração"], start=3):
        ws.cell(row=15, column=i, value=h).fill = fill_gold
        ws.cell(row=15, column=i).font = font_header
    for i in range(10):
        r = 16 + i
        for c in range(3, 9):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C28:K28")
    ws["C28"] = "LICENÇAS"
    ws["C28"].font = font_section
    ws["C28"].fill = fill_brand
    for i, h in enumerate(["Empresa", "Chave", "Plano", "Dias restantes", "Status"], start=3):
        ws.cell(row=29, column=i, value=h).fill = fill_gold
        ws.cell(row=29, column=i).font = font_header
    for i in range(6):
        r = 30 + i
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws.merge_cells("C38:K38")
    ws["C38"] = "Somente SuperAdmin · TrocarEmpresa carrega tenant · Nova Academia faz bootstrap completo."
    ws["C38"].font = Font(name="Calibri", size=9, color="666666")
    ws.row_dimensions[40].height = 36


def build_nova_academia_ui(wb) -> None:
    """39_NOVA_ACADEMIA — cadastro SaaS de tenant."""
    ws = wb.create_sheet("39_NOVA_ACADEMIA")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=42, cols=13)
    add_sidebar(ws, active="39_NOVA_ACADEMIA", rows=42, labels=False)
    add_top_bar(ws, start_col=2, end_col=12)
    set_column_widths(ws, {1: 24, 2: 2, 3: 18, 4: 16, 5: 14, 6: 14, 7: 14, 8: 12, 9: 12})
    _title(ws, "NOVA ACADEMIA")

    fields = [
        (8, "Razão Social", ""),
        (9, "Nome Fantasia", ""),
        (10, "CNPJ", ""),
        (11, "Email", ""),
        (12, "Telefone", ""),
        (13, "Cidade", "São Paulo"),
        (14, "Estado", "SP"),
        (15, "Plano", "Pro"),
        (16, "Admin Login", "admin"),
        (17, "Admin Senha", "123456"),
        (18, "Admin Nome", "Administrador"),
    ]
    for r, lab, val in fields:
        ws.cell(row=r, column=3, value=lab).fill = fill_panel
        ws.cell(row=r, column=3).border = border_thin
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        ws.cell(row=r, column=4, value=val).fill = fill_panel
        ws.cell(row=r, column=4).border = border_thin

    ws.merge_cells("C20:H20")
    ws["C20"] = "Planos: Basic · Pro · Enterprise — ao criar: empresa + licença + admin + configs padrão."
    ws["C20"].font = Font(name="Calibri", size=9, color="666666")

    ws.merge_cells("C22:H22")
    ws["C22"] = "STATUS"
    ws["C22"].font = font_section
    ws["C22"].fill = fill_brand
    ws.merge_cells("C23:H23")
    ws["C23"] = "Preencha os campos e clique em Criar Academia."
    ws["C23"].fill = fill_panel
    ws["C23"].border = border_thin

    ws.row_dimensions[25].height = 36
