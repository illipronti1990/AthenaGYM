"""Sprint 5.1.1 — Operation Center (HOME) + BD_PRIORIDADES."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

from styles import (
    BRAND_RED,
    GOLD,
    add_sidebar,
    add_top_bar,
    apply_header_row,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_normal,
    font_section,
    paint_canvas,
    set_column_widths,
    style_sheet_tab,
)

PRIORIDADES_SEED = [
    ("Mensalidade vencida", 100, "Crítica", "🔴", "04_FINANCEIRO", "Atrasado"),
    ("Equipamento parado", 95, "Crítica", "🔴", "10_EQUIPAMENTOS", "Manutenção"),
    ("Estoque crítico", 90, "Alta", "🟠", "09_ESTOQUE", "Mínimo"),
    ("Plano vencendo", 70, "Média", "🟡", "02_ALUNOS", "Renovação"),
    ("Avaliação física", 60, "Informativa", "🟢", "20_AGENDA", "Avaliação"),
    ("Aula experimental", 50, "Informativa", "🟢", "20_AGENDA", "Experimental"),
    ("Pagamento previsto", 55, "Baixa", "🔵", "04_FINANCEIRO", "Pendente"),
    ("Aniversário", 40, "Informativa", "🟢", "20_AGENDA", "Aniversário"),
]


def _make_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    ws.add_table(table)


def build_bd_prioridades(wb) -> None:
    ws = wb.create_sheet("BD_PRIORIDADES")
    style_sheet_tab(ws, GOLD)
    ws.sheet_view.showGridLines = False
    headers = ["Tipo", "Peso", "Nível", "Ícone", "Destino", "Filtro"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    apply_header_row(ws, 1, 1, len(headers))

    for r_idx, row in enumerate(PRIORIDADES_SEED, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border_thin
            cell.font = font_normal

    last = 1 + len(PRIORIDADES_SEED)
    _make_table(ws, "tbPrioridades", f"A1:F{last}")
    set_column_widths(ws, {1: 22, 2: 8, 3: 12, 4: 8, 5: 18, 6: 14})
    ws.freeze_panes = "A2"


def build_home(wb) -> None:
    """21_HOME — Operation Center (primeira tela pós-login)."""
    ws = wb.create_sheet("21_HOME")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=48, cols=14)
    add_sidebar(ws, active="21_HOME", rows=48, labels=False)
    add_top_bar(ws, start_col=2, end_col=13)

    set_column_widths(
        ws,
        {
            1: 24,
            2: 2,
            3: 6,
            4: 40,
            5: 16,
            6: 8,
            7: 8,
            8: 16,
            9: 12,
            10: 14,
            11: 12,
            12: 12,
            13: 12,
        },
    )

    ws.merge_cells("C5:G5")
    ws["C5"] = "ATHENA GYM ERP — OPERATION CENTER"
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light

    ws.merge_cells("H5:I5")
    ws["H5"] = "='BD_SESSAO'!B2"
    ws["H5"].font = Font(name="Calibri", size=11, bold=True)
    ws["J5"] = '=TEXT(DAY(TODAY()),"00")&"/"&TEXT(MONTH(TODAY()),"00")&"/"&YEAR(TODAY())'
    ws["K5"] = "='BD_SESSAO'!B3"
    ws["K5"].font = Font(name="Calibri", size=10, color="666666")

    ws["C6"] = "🔔 Notificações"
    ws["C6"].font = Font(name="Calibri", size=11, bold=True)
    ws["D6"] = 0
    ws["D6"].font = Font(name="Georgia", size=16, bold=True, color=BRAND_RED)

    # Ações do dia
    ws.merge_cells("C8:G8")
    ws["C8"] = "📌 AÇÕES DO DIA"
    ws["C8"].font = font_section
    ws["C8"].fill = fill_brand
    for i, h in enumerate(["Pri", "Ação", "Abrir", "Peso", "Qtd"], start=3):
        ws.cell(row=9, column=i, value=h).fill = fill_gold
        ws.cell(row=9, column=i).font = font_header

    for i in range(10):
        r = 10 + i
        for c in range(3, 8):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=12).value = ""  # sheet destino
        ws.cell(row=r, column=13).value = ""  # filtro

    ws.column_dimensions["L"].hidden = True
    ws.column_dimensions["M"].hidden = True

    # Próximos eventos
    ws.merge_cells("C21:G21")
    ws["C21"] = "PRÓXIMOS EVENTOS"
    ws["C21"].font = font_section
    ws["C21"].fill = fill_brand
    for i, h in enumerate(["Hora", "Evento", "Referência"], start=3):
        ws.cell(row=22, column=i, value=h).fill = fill_gold
        ws.cell(row=22, column=i).font = font_header
    for i in range(6):
        r = 23 + i
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    # Centro de notificações
    ws.merge_cells("H8:K8")
    ws["H8"] = "🔔 POR MÓDULO"
    ws["H8"].font = font_section
    ws["H8"].fill = fill_brand
    for i, h in enumerate(["Módulo", "Pendências"], start=8):
        ws.cell(row=9, column=i, value=h).fill = fill_gold
        ws.cell(row=9, column=i).font = font_header
    for i, m in enumerate(["Financeiro", "Estoque", "Equipamentos", "Agenda", "Alunos"]):
        r = 10 + i
        ws.cell(row=r, column=8, value=m).fill = fill_panel
        ws.cell(row=r, column=8).border = border_thin
        cell = ws.cell(row=r, column=9, value=0)
        cell.fill = fill_panel
        cell.border = border_thin
        cell.font = Font(name="Georgia", size=12, bold=True, color=BRAND_RED)

    # Painel do perfil
    ws.merge_cells("H16:K16")
    ws["H16"] = "PAINEL DO PERFIL"
    ws["H16"].font = font_section
    ws["H16"].fill = fill_brand

    labels = [("H18", "Label1"), ("J18", "Label2"), ("H21", "Label3"), ("J21", "Label4")]
    for addr, _ in labels:
        ws[addr] = ""
        ws[addr].font = Font(name="Calibri", size=9, color="666666")
        ws[addr].fill = fill_panel

    vals = [("H19", "V1"), ("J19", "V2"), ("H22", "V3"), ("J22", "V4")]
    for addr, _ in vals:
        ws[addr] = 0
        ws[addr].font = Font(name="Georgia", size=14, bold=True, color=BRAND_RED)
        ws[addr].fill = fill_panel
        ws[addr].border = border_thin

    ws.merge_cells("C30:K30")
    ws["C30"] = "Selecione a linha da ação e clique em ABRIR AÇÃO · Tudo é gerado automaticamente pelo motor de prioridades"
    ws["C30"].font = Font(name="Calibri", size=9, color="666666")
