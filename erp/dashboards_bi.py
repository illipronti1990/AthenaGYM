"""Sprint 5.0 — Dashboards especializados (Professores, Estoque, Equipamentos)."""

from __future__ import annotations

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font

from styles import (
    BRAND_RED,
    GOLD,
    ICON,
    add_sidebar,
    add_top_bar,
    border_thin,
    fill_brand,
    fill_gold,
    fill_light,
    fill_panel,
    font_header,
    font_section,
    paint_canvas,
    paint_kpi_card,
    set_column_widths,
    style_sheet_tab,
)


def _money(cell) -> None:
    cell.number_format = 'R$ #,##0.00'


def _title(ws, text: str) -> None:
    ws.merge_cells("C5:H5")
    ws["C5"] = text
    ws["C5"].font = Font(name="Georgia", size=18, bold=True, color=BRAND_RED)
    ws["C5"].fill = fill_light


def build_dash_professores(wb) -> None:
    ws = wb.create_sheet("17_DASH_PROFESSORES")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=40, cols=12)
    add_sidebar(ws, active="08_PROFESSORES", rows=40, labels=False)
    add_top_bar(ws, start_col=2, end_col=11)
    set_column_widths(ws, {1: 24, 2: 2, 3: 13, 4: 13, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13, 10: 12})
    _title(ws, f"{ICON['professores']}  DASHBOARD PROFESSORES")

    paint_kpi_card(ws, 7, 3, "Professores", "='BI_BASE'!E21", False, 2)
    paint_kpi_card(ws, 7, 5, "Alunos ativos", "='BI_BASE'!E5", False, 2)
    paint_kpi_card(ws, 7, 7, "Avaliações/mês", "='BI_BASE'!E22", False, 2)
    paint_kpi_card(ws, 7, 9, "Professor do mês", "='BI_BASE'!G12", False, 2)

    ws.merge_cells("C12:F12")
    ws["C12"] = "RANKING — ALUNOS POR PROFESSOR"
    ws["C12"].font = font_section
    ws["C12"].fill = fill_brand
    for i, h in enumerate(["#", "Professor", "Alunos", "Barra"], start=3):
        ws.cell(row=13, column=i, value=h).fill = fill_gold
        ws.cell(row=13, column=i).font = font_header
    for i in range(10):
        r = 14 + i
        ws.cell(row=r, column=3, value=f"='BI_BASE'!F{12+i}")
        ws.cell(row=r, column=4, value=f"='BI_BASE'!G{12+i}")
        ws.cell(row=r, column=5, value=f"='BI_BASE'!H{12+i}")
        ws.cell(row=r, column=6, value=f'=REPT("█",MIN(20,IFERROR(E{r}/MAX($E$14:$E$23,1)*20,0)))')
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    chart = BarChart()
    chart.type = "bar"
    chart.title = "Top professores"
    data = Reference(ws, min_col=5, min_row=13, max_row=23)
    cats = Reference(ws, min_col=4, min_row=14, max_row=23)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 12
    ws.add_chart(chart, "H12")


def build_dash_estoque_bi(wb) -> None:
    ws = wb.create_sheet("18_DASH_ESTOQUE")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=40, cols=12)
    add_sidebar(ws, active="09_ESTOQUE", rows=40, labels=False)
    add_top_bar(ws, start_col=2, end_col=11)
    set_column_widths(ws, {1: 24, 2: 2, 3: 13, 4: 13, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13, 10: 12})
    _title(ws, f"{ICON['estoque']}  DASHBOARD ESTOQUE")

    paint_kpi_card(ws, 7, 3, "Produtos", "='BI_BASE'!E23", False, 2)
    paint_kpi_card(ws, 7, 5, "Estoque baixo", "='BI_BASE'!E24", False, 2)
    paint_kpi_card(ws, 7, 7, "Valor estoque", "='BI_BASE'!E25", True, 2)
    paint_kpi_card(ws, 7, 9, "Semáforo", "='BI_BASE'!H5", False, 2)

    ws.merge_cells("C12:G12")
    ws["C12"] = "ALERTAS DE ESTOQUE MÍNIMO"
    ws["C12"].font = font_section
    ws["C12"].fill = fill_brand
    for i, h in enumerate(["Produto", "Qtd", "Mínimo", "Status"], start=3):
        ws.cell(row=13, column=i, value=h).fill = fill_gold
        ws.cell(row=13, column=i).font = font_header

    for i in range(8):
        r = 14 + i
        ws.cell(row=r, column=3, value=f"='09_ESTOQUE'!B{6+i}")
        ws.cell(row=r, column=4, value=f"='09_ESTOQUE'!D{6+i}")
        ws.cell(row=r, column=5, value=f"='09_ESTOQUE'!E{6+i}")
        ws.cell(row=r, column=6, value=f"='09_ESTOQUE'!L{6+i}")
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin

    ws["C24"] = "Dica: abra 09_ESTOQUE para entradas/saídas. O BI consolida alertas automaticamente."
    ws["C24"].font = Font(name="Calibri", size=9, color="666666")


def build_dash_equipamentos_bi(wb) -> None:
    ws = wb.create_sheet("19_DASH_EQUIPAMENTOS")
    style_sheet_tab(ws, GOLD)
    paint_canvas(ws, rows=40, cols=12)
    add_sidebar(ws, active="10_EQUIPAMENTOS", rows=40, labels=False)
    add_top_bar(ws, start_col=2, end_col=11)
    set_column_widths(ws, {1: 24, 2: 2, 3: 13, 4: 13, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13, 10: 12})
    _title(ws, f"{ICON['estoque']}  DASHBOARD EQUIPAMENTOS")

    paint_kpi_card(ws, 7, 3, "Equipamentos", "='BI_BASE'!E26", False, 2)
    paint_kpi_card(ws, 7, 5, "Em manutenção", "='BI_BASE'!E27", False, 2)
    paint_kpi_card(ws, 7, 7, "Próx. revisão", '=COUNTIF(\'10_EQUIPAMENTOS\'!H:H,"<=30")', False, 2)
    paint_kpi_card(ws, 7, 9, "% OK", "=IFERROR(1-IFERROR('BI_BASE'!E27/'BI_BASE'!E26,0),1)", False, 2)
    ws.cell(row=8, column=9).number_format = "0%"

    ws.merge_cells("C12:F12")
    ws["C12"] = "STATUS DOS EQUIPAMENTOS"
    ws["C12"].font = font_section
    ws["C12"].fill = fill_brand
    for i, h in enumerate(["Equipamento", "Status", "Dias p/ revisão"], start=3):
        ws.cell(row=13, column=i, value=h).fill = fill_gold
        ws.cell(row=13, column=i).font = font_header
    for i in range(8):
        r = 14 + i
        ws.cell(row=r, column=3, value=f"='10_EQUIPAMENTOS'!A{6+i}")
        ws.cell(row=r, column=4, value=f"='10_EQUIPAMENTOS'!G{6+i}")
        ws.cell(row=r, column=5, value=f"='10_EQUIPAMENTOS'!H{6+i}")
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill_panel
            ws.cell(row=r, column=c).border = border_thin
